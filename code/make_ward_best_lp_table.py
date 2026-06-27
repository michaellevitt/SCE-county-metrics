"""make_ward_best_lp_table.py

For each of the 100 XDE clusters, extract the representative metric with the
strongest correlation, ranked by max |CC| (|max_signed_cc|) descending.
Output: Excel tab "SEM_Best_CC" with the columns:

  super_cluster_id | super_cluster_name | Ward100 | cluster_label |
  metric | explain | max_signed_cc | min_lp | sum_abs_lp |
  n_sig_years | SI | SI_near | sign | age | dominant_age

The output FILENAME is kept as sem_best_lp_*.xlsx for pipeline compatibility;
only the internal sheet name / headline column / ranking are |CC|-based.

Usage (single line):
  python3 code/make_ward_best_lp_table.py --master-xlsx master_sem_clusters_2745.xlsx --output sem_best_lp.xlsx
"""

import argparse, sys, os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

def _tee(path):
    """Print 'Saved <relpath>' to stdout and stderr."""
    _p = str(path)
    try:
        _rel = os.path.relpath(_p)
    except ValueError:
        _rel = _p
    msg = 'Saved ' + _rel
    print(msg)
    print(msg, file=sys.stderr, flush=True)


COLS = [
    'super_cluster_id', 'super_cluster_name', 'Ward100', 'cluster_label',
    'metric', 'explain', 'max_signed_cc', 'min_lp', 'sum_abs_lp',
    'n_sig_years', 'SI', 'SI_near', 'sign', 'age', 'dominant_age',
]

NAVY  = '1F3864'; MID   = '2E75B6'; WHITE = 'FFFFFF'
DKRED = '8B0000'; DKBLUE = '1F4E79'
GOLD  = 'FFD700'
YELLOW_HL = 'FFFF00'

thin  = Side(style='thin',   color='CCCCCC')
thick = Side(style='medium', color='888888')

def brd(c, top=False, bot=False):
    c.border = Border(left=thin, right=thin,
                      top=(thick if top else thin),
                      bottom=(thick if bot else thin))
    return c

def cell(ws, r, c, val, bg=WHITE, fg='000000', bold=False,
         left=False, wrap=False, courier=False, size=9):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(
        name='Courier New' if courier else 'Arial',
        bold=bold, color=fg, size=size
    )
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(
        horizontal='left' if left else 'center',
        vertical='center', wrap_text=wrap
    )
    return brd(cl)


def load_master(xlsx_path):
    """Load Master tab; fall back to Multi_Year if Master absent.
    Handles both legacy (A1) and styled (B2) workbook layouts."""
    from _workbook_style import read_styled_excel
    from openpyxl import load_workbook
    wb_peek = load_workbook(xlsx_path, read_only=True)
    tab = 'Master' if 'Master' in wb_peek.sheetnames else wb_peek.sheetnames[0]
    wb_peek.close()
    df = read_styled_excel(xlsx_path, tab)
    df.columns = [str(c).strip() for c in df.columns]
    print(f"  Loaded {len(df)} rows from tab '{tab}'", file=sys.stderr)
    return df


def compute_best_lp(df):
    """Return one row per XDE100 cluster: the metric with the largest |CC|
    (|max_signed_cc|).  Final table is ranked by max |CC| descending."""
    required = ['Ward100', 'metric', 'max_signed_cc']
    for c in required:
        if c not in df.columns:
            sys.exit(f"ERROR: column '{c}' not found in master. Columns: {list(df.columns)}")

    df = df.copy()
    df['max_signed_cc'] = pd.to_numeric(df['max_signed_cc'], errors='coerce')
    if 'min_lp' in df.columns:
        df['min_lp'] = pd.to_numeric(df['min_lp'], errors='coerce')
    df['abs_cc'] = df['max_signed_cc'].abs()

    # For each Ward100, pick the row with the largest |CC| (representative metric)
    best = (df.sort_values('abs_cc', ascending=False)
              .groupby('Ward100', sort=False)
              .first()
              .reset_index())

    # Ensure all COLS present (fill missing with '')
    for c in COLS:
        if c not in best.columns:
            best[c] = ''

    best['abs_cc'] = best['max_signed_cc'].abs()
    best = best[COLS + ['abs_cc']].copy()
    # Rank by max |CC| descending (strongest first)
    best = best.sort_values('abs_cc', ascending=False).reset_index(drop=True)
    best = best.drop(columns=['abs_cc'])
    print(f"  {len(best)} XDE clusters with best-|CC| metric", file=sys.stderr)
    return best


def write_excel(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'SEM_Best_CC'

    # -- column widths (matching screenshot proportions) --
    widths = {
        'super_cluster_id':   5,
        'super_cluster_name': 30,
        'Ward100':            7,
        'cluster_label':      24,
        'metric':             10,
        'explain':            40,
        'max_signed_cc':       9,
        'min_lp':              9,
        'sum_abs_lp':          9,
        'n_sig_years':         6,
        'SI':                  8,
        'SI_near':            10,
        'sign':                8,
        'age':                 8,
        'dominant_age':        9,
    }
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 12)

    ri = 1
    # Title
    ws.row_dimensions[ri].height = 20
    for ci in range(1, len(COLS)+1):
        cl = ws.cell(ri, ci,
            f'XDE cluster best-|CC| metric  --  {len(df)} XDE clusters  '
            f'(one metric per cluster, largest max |CC|; ranked by max |CC| descending)'
            if ci == 1 else '')
        cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        cl.fill = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    # Column headers -- rotated 60 deg to match screenshot
    ws.row_dimensions[ri].height = 72
    for ci, col in enumerate(COLS, 1):
        cl = ws.cell(ri, ci, col)
        cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=MID)
        cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                  text_rotation=60, wrap_text=False)
        brd(cl)
    ri += 1

    ws.freeze_panes = 'A3'

    # headline |CC| highlight column index
    hl_col = COLS.index('max_signed_cc') + 1  # 1-based

    # SC colour palette (alternating two shades per SC)
    SC_LIGHT = ['FFF3E0','FFF9C4','E8F5E9','E3F2FD','F3E5F5',
                'FCE4D6','E0F7FA','F1F8E9','EDE7F6','FBE9E7',
                'E8EAF6','F9FBE7']
    SC_DARK  = ['FFE0B2','FFF176','C8E6C9','BBDEFB','E1BEE7',
                'FFCCBC','B2EBF2','DCEDC8','D1C4E9','FFCCBC',
                'C5CAE9','F0F4C3']

    prev_sc = None
    shade_toggle = 0
    for _, row in df.iterrows():
        ws.row_dimensions[ri].height = 30
        sc = int(row['super_cluster_id']) if pd.notna(row['super_cluster_id']) else 0
        if sc != prev_sc:
            shade_toggle = 0
            prev_sc = sc
        else:
            shade_toggle = 1 - shade_toggle
        pal = SC_LIGHT if shade_toggle == 0 else SC_DARK
        bg = pal[(sc - 1) % len(pal)]

        for ci, col in enumerate(COLS, 1):
            val = row[col]
            if pd.isna(val):
                val = ''
            elif col in ('max_signed_cc', 'min_lp', 'sum_abs_lp'):
                try:
                    val = round(float(val), 3)
                except Exception:
                    pass
            elif col in ('Ward100', 'super_cluster_id', 'n_sig_years'):
                try:
                    val = int(val)
                except Exception:
                    pass

            is_hl = (ci == hl_col)
            is_left = col in ('super_cluster_name', 'cluster_label', 'explain')
            is_cour = col in ('SI', 'SI_near', 'sign', 'age')
            is_bold = col in ('Ward100', 'min_lp', 'max_signed_cc', 'SI')
            fg = DKRED if (col == 'max_signed_cc' and isinstance(val, float) and val > 0) else \
                 DKBLUE if (col == 'max_signed_cc' and isinstance(val, float) and val < 0) else \
                 DKRED  if col == 'min_lp' else '000000'

            cell(ws, ri, ci, val,
                 bg=YELLOW_HL if is_hl else bg,
                 fg=fg, bold=is_bold, left=is_left,
                 wrap=(col == 'explain'), courier=is_cour, size=9)
        ri += 1

    wb.save(out_path)
    _tee(out_path)
    try:
        _r_out_path = os.path.relpath(out_path)
    except ValueError:
        _r_out_path = out_path
    print('Saved ' + _r_out_path, file=sys.stderr, flush=True)
    print(f"Saved: {out_path}", flush=True)


def main():
    p = argparse.ArgumentParser(description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--master-xlsx', required=True,
                   help='master_sem_clusters_2745.xlsx')
    p.add_argument('--output', default='sem_best_lp.xlsx',
                   help='Output xlsx (default: sem_best_lp.xlsx)')
    args = p.parse_args()

    print("Loading master xlsx...", file=sys.stderr)
    df = load_master(args.master_xlsx)
    best = compute_best_lp(df)
    write_excel(best, args.output)
    print(f"\n{len(best)} rows written (one per XDE cluster)", flush=True)


if __name__ == '__main__':
    main()
