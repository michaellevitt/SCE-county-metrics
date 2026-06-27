#!/usr/bin/env python3
"""
ward_medoid_ccbs_v4.py

Combined pipeline:
  1. Load square CC matrix
  2. Run XDE clustering at --ward-k clusters, find medoids
  3. Filter to medoids with >=1 significant CC (LP <= --lp-threshold)
     against the 15 pandemic death measures (2020-2024 x All/GE65/LT65)
  4. Compute medoid-level Ward linkage and plot dendrogram (PNG)
  5. Run CC_BS sig-pattern analysis on the significant medoids (Excel)

Outputs (in --out-dir):
  xde{k}_reps.csv               -- all medoids with cluster size and best CC
  xde{k}_assignments.csv        -- per-metric cluster assignments (NEW in v4)
  medoid_list_xde{k}.json       -- ordered medoid list per k (NEW in v4)
  Z_med_xde{k}.npy             -- medoid linkage matrix per k (NEW in v4)
  ward{k}_sig_medoids.csv        -- significant medoids only
  Z_ward.npy                     -- full Ward linkage (reused across k values)
  cc_sub.npy                     -- metric CC submatrix (reused across k values)
  metrics_list.json              -- metric names + best_cc (reused across k values)
  medoid_list.json               -- full ordered medoid list
  medoid_dendrogram_{desc|k400}.png
  top{N}_sig_patterns_ward{k}.xlsx

Usage:
  python3 ward_medoid_ccbs_v4.py --cc-matrix full_cc_ase0_p=0.0.csv.gz
      --death-cc metric_x_death_cc=0.0.csv
      --explain BEN_MERGED_MEASURES_imputed_20s_v1_31_GG_Add2024.explain
      --hub hub_combined_k400_labels.csv
      --ward-k 323 --out-dir out/

Arguments:
  --cc-matrix FILE      Square CC matrix CSV or CSV.GZ (required)
  --death-cc FILE       metric x death CC/LP CSV (required)
  --explain FILE        Metric explain file (optional)
  --hub FILE            hub_combined_k400_labels.csv (optional)
  --ward-k INT          Number of XDE clusters / medoids (default: 323)
  --lp-threshold FLOAT  LP threshold for significance (default: -14.0)
  --top-n INT           Top-N sig patterns for Excel (default: 10)
  --nsuper INT          Super-clusters for dendrogram colouring (default: 12)
  --linthresh FLOAT     Symlog linear threshold for dendrogram x-axis (default: 0.5)
  --k400                Use k400 cluster labels in dendrogram (default: metric desc)
  --out-dir DIR         Output directory (default: .)
  --skip-precompute     Skip slow full XDE clustering; reload Z_ward.npy and recut at --ward-k
"""

import argparse
import json
import os
import sys
import time

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Patch, FancyBboxPatch
from matplotlib.lines import Line2D
import numpy as np
import pandas as pd
from scipy.cluster.hierarchy import linkage, fcluster, dendrogram
from scipy.spatial.distance import squareform
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _tee(path):
    """Print 'Saved <relpath>' to stdout and stderr."""
    _p = str(path)
    try:
        _rel = os.path.relpath(_p)
    except ValueError:
        _rel = _p
    msg = 'Saved ' + _rel
    print(msg)
    print(msg, file=sys.stderr, flush=True)


# -- pandemic death columns (15-bit) ------------------------------------------
PANDEMIC_YEARS = ['2020', '2021', '2022', '2023', '2024']
PANDEMIC_AGES  = ['All', 'GE65', 'LT65']
PANDEMIC_PAIRS = []
for _yr in PANDEMIC_YEARS:
    for _age in PANDEMIC_AGES:
        _sfx = '' if _age == 'All' else f'_{_age}'
        PANDEMIC_PAIRS.append((f'asedx_p_{_yr}{_sfx}', f'LP_asedx_p_{_yr}{_sfx}'))

# Significance mode: 'lp' (p-value, default) or 'cc' (|CC| > CC_SIG magnitude band).
SIG_MODE = 'lp'
CC_SIG   = 0.3

def _sig_test(ccv, lpv, lp_threshold):
    """Per-pair significance under the active mode (lp: LP<=thr; cc: |CC|>CC_SIG)."""
    if SIG_MODE == 'cc':
        try:
            return pd.notna(ccv) and abs(float(ccv)) > CC_SIG
        except (TypeError, ValueError):
            return False
    try:
        return pd.notna(lpv) and float(lpv) <= lp_threshold
    except (TypeError, ValueError):
        return False

PALETTE = [
    '#E41A1C', '#377EB8', '#4DAF4A', '#FF7F00', '#984EA3',
    '#A65628', '#F781BF', '#999999', '#1B9E77', '#D95F02',
    '#7570B3', '#E7298A',
]

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)

def clean_desc(desc):
    if '=' in desc:
        desc = desc[:desc.index('=')]
    return desc.strip().rstrip('_').rstrip('-').strip()

# -- Step 1: Full XDE clustering (slow -- run once) --------------------------
def compute_ward_full(cc_matrix_path, out_dir):
    """Load CC matrix, compute full Ward linkage, save Z_ward.npy + cc_sub.npy + metrics_list.json."""
    log(f"Loading CC matrix: {cc_matrix_path}")
    cc_df   = pd.read_csv(cc_matrix_path, index_col=0)
    all_ids = list(cc_df.index)
    death_set = {m for m in all_ids if m.startswith('asedx_p_')}
    metrics   = [m for m in all_ids if m not in death_set]
    log(f"  {len(metrics)} non-death metrics")

    cc_sub = cc_df.loc[metrics, metrics].values.astype(float)
    np.fill_diagonal(cc_sub, 1.0)
    cc_sub = np.where(np.isfinite(cc_sub), cc_sub, 0.0)

    # best_cc against any death column (metric-level, not k-dependent)
    death_cols = [c for c in cc_df.columns if c.startswith('asedx_p_')]
    if death_cols:
        cc_death     = cc_df.loc[metrics, death_cols].values.astype(float)
        cc_death     = np.where(np.isfinite(cc_death), cc_death, 0.0)
        best_col     = np.abs(cc_death).argmax(axis=1)
        best_cc_vals = cc_death[np.arange(len(metrics)), best_col]
        best_cc_list = best_cc_vals.tolist()
    else:
        best_cc_list = [float('nan')] * len(metrics)

    log("Computing Ward linkage on full matrix (this is the slow step)...")
    dist = 1.0 - np.abs(cc_sub)
    dist = (dist + dist.T) / 2.0
    np.fill_diagonal(dist, 0.0)
    Z_ward = linkage(squareform(dist), method='ward', optimal_ordering=True)
    log("  Done.")

    np.save(os.path.join(out_dir, 'Z_ward.npy'), Z_ward)
    np.save(os.path.join(out_dir, 'cc_sub.npy'), cc_sub)
    with open(os.path.join(out_dir, 'metrics_list.json'), 'w') as f:
        json.dump({'metrics': metrics, 'best_cc': best_cc_list}, f)
    log(f"  Saved Z_ward.npy, cc_sub.npy, metrics_list.json")
    return Z_ward, cc_sub, metrics, best_cc_list

# -- Step 2: Cut at k, find medoids, compute medoid linkage (fast) -------------
def compute_medoids_from_ward(Z_ward, cc_sub, metrics, best_cc_list, k, out_dir):
    """Cut Ward tree at k, find medoids, compute kxk medoid linkage. Fast -- rerun freely."""
    log(f"Cutting Ward tree at k={k}...")
    labels = fcluster(Z_ward, t=k, criterion='maxclust')
    log(f"  {len(set(labels))} clusters")

    metric_idx  = {m: i for i, m in enumerate(metrics)}
    best_cc_map = dict(zip(metrics, best_cc_list))
    cluster_map = {}
    for m, lbl in zip(metrics, labels):
        cluster_map.setdefault(lbl, []).append(m)

    log("Finding medoids...")
    rep_rows, medoid_list = [], []
    for cid in sorted(cluster_map.keys()):
        members = cluster_map[cid]
        idxs    = [metric_idx[m] for m in members]
        sub     = cc_sub[np.ix_(idxs, idxs)]
        medoid  = members[int(np.abs(sub).mean(axis=1).argmax())]
        medoid_list.append(medoid)
        rep_rows.append({
            f'Ward{k}': cid,
            'medoid':       medoid,
            'cluster_size': len(members),
            'best_cc':      round(float(best_cc_map.get(medoid, float('nan'))), 4),
        })
    log(f"  {len(medoid_list)} medoids")

    log("Computing medoid linkage matrix...")
    med_idxs = [metric_idx[m] for m in medoid_list]
    cc_med   = cc_sub[np.ix_(med_idxs, med_idxs)]
    np.fill_diagonal(cc_med, 1.0)
    dist_med = 1.0 - np.abs(cc_med)
    dist_med = (dist_med + dist_med.T) / 2.0
    np.fill_diagonal(dist_med, 0.0)
    Z_med = linkage(squareform(dist_med), method='ward', optimal_ordering=True)

    reps_path = os.path.join(out_dir, f'xde{k}_reps.csv')
    pd.DataFrame(rep_rows).to_csv(reps_path, index=False)
    _tee(reps_path)
    log(f"  Saved {reps_path}")

    # Save per-metric cluster assignments (needed by make_master_excel)
    assignments_path = os.path.join(out_dir, f'xde{k}_assignments.csv')
    pd.DataFrame({'metric': metrics, f'Ward{k}': labels}).to_csv(
        assignments_path, index=False)
    log(f"  Saved {assignments_path}")

    # Save medoid_list.json in format expected by make_master_excel
    medoid_list_path = os.path.join(out_dir, f'medoid_list_xde{k}.json')
    with open(medoid_list_path, 'w') as f:
        json.dump(medoid_list, f)
    log(f"  Saved {medoid_list_path}")

    # Save Z_med for dendrogram plotting
    z_med_path = os.path.join(out_dir, f'Z_med_xde{k}.npy')
    np.save(z_med_path, Z_med)
    _tee(z_med_path)
    log(f"  Saved {z_med_path}")

    return Z_med, medoid_list, pd.DataFrame(rep_rows)

# -- Step 3: Filter significant medoids ---------------------------------------
def filter_sig_medoids(medoid_list, death_cc_df, lp_threshold, out_dir, k):
    pairs = [(cc, lp) for cc, lp in PANDEMIC_PAIRS
             if lp in death_cc_df.columns and cc in death_cc_df.columns]
    lp_cols = [lp for _, lp in pairs]
    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df

    sig_medoids = []
    for m in medoid_list:
        if m not in dc.index:
            continue
        row = dc.loc[m]
        if any(_sig_test(row.get(cc), row.get(lp), lp_threshold) for cc, lp in pairs):
            sig_medoids.append(m)

    _crit = f"|CC|>{CC_SIG}" if SIG_MODE == 'cc' else f"LP<={lp_threshold}"
    log(f"  {len(sig_medoids)} of {len(medoid_list)} medoids have >=1 sig CC ({_crit})")

    sig_path = os.path.join(out_dir, f'ward{k}_sig_medoids.csv')
    rows = []
    for m in sig_medoids:
        row = dc.loc[m]
        n_sig    = sum(1 for cc, lp in pairs if _sig_test(row.get(cc), row.get(lp), lp_threshold))
        best_lp  = min((row.get(c) for c in lp_cols if pd.notna(row.get(c))), default=float('nan'))
        rows.append({'metric': m, 'n_sig': n_sig, 'best_lp': round(best_lp, 2)})
    pd.DataFrame(rows).to_csv(sig_path, index=False)
    log(f"  Saved {sig_path}")
    return sig_medoids

# -- Step 4: Dendrogram --------------------------------------------------------
def build_node_colors(Z, n, super_labels, palette):
    node_color = {}
    for i in range(n):
        node_color[i] = palette[(super_labels[i] - 1) % len(palette)]
    for step, row in enumerate(Z):
        node_id = n + step
        l, r = int(row[0]), int(row[1])
        lc = node_color.get(l, '#555555')
        rc = node_color.get(r, '#555555')
        node_color[node_id] = lc if lc == rc else '#555555'
    return node_color

def plot_dendrogram(Z_med, medoid_list, reps_df, death_cc_df, explain, k400_map,
                    lp_threshold, nsuper, linthresh, use_k400, k, out_dir):
    n = len(medoid_list)
    ward_col = [c for c in reps_df.columns if c.startswith('Ward')][0]
    medoid_to_wid  = dict(zip(reps_df['medoid'], reps_df[ward_col]))
    ward_size      = dict(zip(reps_df[ward_col], reps_df['cluster_size']))
    best_cc_map    = dict(zip(reps_df['medoid'], reps_df['best_cc'])) if 'best_cc' in reps_df.columns else {}

    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
    pairs = [(cc, lp) for cc, lp in PANDEMIC_PAIRS if lp in dc.columns and cc in dc.columns]

    bit_count = {}
    for m in medoid_list:
        if m in dc.index:
            row = dc.loc[m]
            bit_count[m] = sum(1 for cc, lp in pairs if _sig_test(row.get(cc), row.get(lp), lp_threshold))
        else:
            bit_count[m] = 0

    labels = []
    for m in medoid_list:
        wid    = medoid_to_wid.get(m, '?')
        sz     = ward_size.get(wid, 1)
        prefix = f"W{wid}(n={sz})  "
        cc_val = best_cc_map.get(m, float('nan'))
        cc_str = f"  {cc_val:+.3f}" if np.isfinite(float(cc_val)) else ""
        if use_k400:
            primary = k400_map.get(m, clean_desc(explain.get(m, m)))
            label   = f"{prefix}{primary}  [{m}]{cc_str}"
        else:
            desc  = clean_desc(explain.get(m, m))
            label = f"{prefix}{desc}{cc_str}"
        labels.append(label)

    super_labels = fcluster(Z_med, t=nsuper, criterion='maxclust')
    node_color   = build_node_colors(Z_med, n, super_labels, PALETTE)

    def link_color_func(link_id):
        return node_color.get(link_id, '#555555')

    LEAF_FONT   = 6.5 * 1.4
    LEG_FONT    = 8   * 1.4
    XLABEL_FONT = 10  * 1.4
    TICK_FONT   = 8   * 1.4
    fig_w = (22 if use_k400 else 18) * 1.4
    fig_h = max(40, n * 0.13) * 0.40
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=130)

    dn = dendrogram(
        Z_med, labels=labels, orientation='left', leaf_font_size=LEAF_FONT,
        link_color_func=link_color_func, above_threshold_color='#555555', ax=ax,
    )
    for line in ax.get_lines():
        line.set_linewidth(line.get_linewidth() * 2.5)

    ax.set_xscale('symlog', linthresh=linthresh)
    ax.invert_xaxis()
    ax.yaxis.tick_left()
    ax.yaxis.set_tick_params(pad=6, length=0)

    font_pt  = LEG_FONT
    lines    = 7
    fig_h_in = fig.get_size_inches()[1]
    gap_frac = (lines * font_pt / 72) / fig_h_in
    pos = ax.get_position()
    ax.set_position([pos.x0 + pos.width * 0.55, pos.y0,
                     pos.width * 0.25, pos.height - gap_frac])

    ax.set_xlabel('Ward linkage distance  (1 - |CC|)  [symlog scale]', fontsize=XLABEL_FONT)
    ax.tick_params(axis='x', labelsize=TICK_FONT)

    fig.canvas.draw()
    ylbls      = ax.get_yticklabels()
    leaf_order = dn['leaves']
    n_red = n_black = 0
    for pos_i, leaf_idx in enumerate(leaf_order):
        m   = medoid_list[leaf_idx]
        col = '#CC0000' if bit_count.get(m, 0) > 0 else '#000000'
        ylbls[pos_i].set_color(col)
        if col == '#CC0000': n_red   += 1
        else:                n_black += 1
    log(f"  Dendrogram: {n_red} red (sig), {n_black} black labels")

    mode_str = 'k400 label + metric ID' if use_k400 else 'metric description'
    branch_handles = [Patch(facecolor=PALETTE[i % len(PALETTE)], label=f'SC-{i+1}')
                      for i in range(nsuper)]
    sig_handles = [
        Line2D([0], [0], color='#000000', lw=2, label='Black: not sig'),
        Line2D([0], [0], color='#CC0000', lw=2,
               label=(f'Red: |CC|>{CC_SIG}' if SIG_MODE == 'cc' else f'Red: LP<={lp_threshold}')),
    ]
    leg = ax.legend(
        handles=branch_handles + sig_handles,
        loc='lower center', bbox_to_anchor=(0.5, 1.0), fontsize=LEG_FONT,
        framealpha=0.88, ncol=4,
        title=(f'Ward Dendrogram  |  {n} medoids  |  {nsuper} super-clusters  |  '
               f'Labels: {mode_str}  |  linthresh={linthresh}'),
        title_fontsize=LEG_FONT,
        labelspacing=0.2, handletextpad=0.4, borderpad=0.5, columnspacing=1.0,
    )
    leg.get_title().set_multialignment('center')
    fig.canvas.draw()
    ax_pos = ax.get_position()
    leg.set_bbox_to_anchor((0.5, ax_pos.y0 + ax_pos.height), transform=fig.transFigure)

    mode_tag = 'k400' if use_k400 else 'desc'
    outpath  = os.path.join(out_dir, f'medoid_dendrogram_{mode_tag}_ward{k}.png')
    fig.savefig(outpath, dpi=130, bbox_inches='tight')
    plt.close()
    log(f"  Saved: {outpath}")


# -- Step 6: Full medoid roster sheet -----------------------------------------
def make_medoid_roster_sheet(ws, medoid_list, reps_df, death_cc_df, explain,
                              k400_map, lp_threshold, k):
    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
    cc_lp_pairs = [(cc, lp) for cc, lp in PANDEMIC_PAIRS
                   if cc in dc.columns and lp in dc.columns]

    ward_col   = [c for c in reps_df.columns if c.startswith('Ward')][0]
    size_map   = dict(zip(reps_df['medoid'], reps_df['cluster_size']))
    wid_map    = dict(zip(reps_df['medoid'], reps_df[ward_col]))
    bestcc_map = dict(zip(reps_df['medoid'], reps_df['best_cc'])) if 'best_cc' in reps_df.columns else {}

    mono     = Font(name='Courier New', size=9)
    norm     = Font(name='Arial', size=9)
    bold_hdr = Font(name='Arial', bold=True, size=9)
    red_bold = Font(name='Arial', size=9, color='CC0000', bold=True)
    thin     = Side(style='thin', color='CCCCCC')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', start_color='DDEEFF')
    sig_fill = PatternFill('solid', start_color='FFF0F0')
    alt_fill = PatternFill('solid', start_color='F5F8FF')
    cc_fill  = PatternFill('solid', start_color='D0EED0')
    lp_fill  = PatternFill('solid', start_color='EED0D0')

    FIXED_HDRS = [
        ('ward_id',      6,  'center'),
        ('cluster_size', 8,  'center'),
        ('metric',       14, 'left'),
        ('description',  50, 'left'),
        ('k400_label',   30, 'left'),
        ('n_sig',         6, 'center'),
        ('SU_pattern',   22, 'left'),
        ('best_cc',       8, 'center'),
    ]
    n_fixed = len(FIXED_HDRS)

    year_short = {'2020': '20', '2021': '21', '2022': '22', '2023': '23', '2024': '24'}
    age_short  = {'': 'A', '_GE65': 'G', '_LT65': 'L'}
    cc_hdrs, lp_hdrs = [], []
    for cc_col, lp_col in cc_lp_pairs:
        base = cc_col.replace('asedx_p_', '')
        for yr, ys in year_short.items():
            if base.startswith(yr):
                sfx = base[len(yr):]
                ag  = age_short.get(sfx, sfx)
                cc_hdrs.append(f'{ys}{ag}')
                lp_hdrs.append(f'LP{ys}{ag}')
                break
        else:
            cc_hdrs.append(cc_col)
            lp_hdrs.append(lp_col)

    cc_start = n_fixed + 1
    cc_end   = n_fixed + len(cc_lp_pairs)
    lp_start = cc_end + 1
    lp_end   = cc_end + len(cc_lp_pairs)

    # Row 1: group headers
    for ci, (hdr, width, align) in enumerate(FIXED_HDRS):
        col_idx = ci + 1
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font = bold_hdr; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    c = ws.cell(row=1, column=cc_start,
                value='Correlation Coefficient (CC) -- 2020-2024 x All/GE65/LT65')
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = cc_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    if len(cc_lp_pairs) > 1:
        ws.merge_cells(start_row=1, start_column=cc_start, end_row=1, end_column=cc_end)

    c = ws.cell(row=1, column=lp_start,
                value='Log10 P-value (LP) -- red/bold = significant')
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = lp_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    if len(cc_lp_pairs) > 1:
        ws.merge_cells(start_row=1, start_column=lp_start, end_row=1, end_column=lp_end)

    # Row 2: individual column headers
    for ci, hdr in enumerate(cc_hdrs):
        col_idx = cc_start + ci
        c = ws.cell(row=2, column=col_idx, value=hdr)
        c.font = bold_hdr; c.fill = cc_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws.column_dimensions[get_column_letter(col_idx)].width = 6
    for ci, hdr in enumerate(lp_hdrs):
        col_idx = lp_start + ci
        c = ws.cell(row=2, column=col_idx, value=hdr)
        c.font = bold_hdr; c.fill = lp_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_i, m in enumerate(medoid_list):
        excel_row = row_i + 3
        n_sig, su_bits, cc_vals, lp_vals = 0, '', [], []
        if m in dc.index:
            row_data = dc.loc[m]
            for cc_col, lp_col in cc_lp_pairs:
                ccv = float(row_data.get(cc_col, float('nan')))
                lpv = float(row_data.get(lp_col, float('nan')))
                cc_vals.append(ccv); lp_vals.append(lpv)
                if _sig_test(ccv, lpv, lp_threshold):
                    n_sig += 1; su_bits += 'S'
                else:
                    su_bits += 'U'
        else:
            cc_vals = [float('nan')] * len(cc_lp_pairs)
            lp_vals = [float('nan')] * len(cc_lp_pairs)
            su_bits = 'U' * len(cc_lp_pairs)

        is_sig   = n_sig > 0
        row_fill = sig_fill if is_sig else (alt_fill if row_i % 2 == 1 else None)

        fixed_vals = [
            wid_map.get(m, ''), size_map.get(m, ''), m,
            clean_desc(explain.get(m, m)), k400_map.get(m, ''),
            n_sig, su_bits,
            round(float(bestcc_map.get(m, float('nan'))), 4),
        ]
        for ci, val in enumerate(fixed_vals):
            col_idx = ci + 1
            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.font = red_bold if (ci == 5 and is_sig) else norm
            c.alignment = Alignment(horizontal=FIXED_HDRS[ci][2], vertical='center')
            c.border = bdr
            if row_fill: c.fill = row_fill

        for ci, (ccv, (cc_col, lp_col)) in enumerate(zip(cc_vals, cc_lp_pairs)):
            col_idx  = cc_start + ci
            lpv      = lp_vals[ci]
            cell_sig = _sig_test(ccv, lpv, lp_threshold)
            display  = round(ccv, 3) if np.isfinite(ccv) else None
            c = ws.cell(row=excel_row, column=col_idx, value=display)
            c.font = Font(name='Courier New', size=9, bold=cell_sig,
                          color=('CC0000' if cell_sig and ccv > 0 else
                                 '0000CC' if cell_sig and ccv < 0 else '000000'))
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr
            c.fill = (PatternFill('solid', start_color='FFE8E8') if (cell_sig and ccv > 0) else
                      PatternFill('solid', start_color='E8E8FF') if (cell_sig and ccv < 0) else
                      row_fill if row_fill else PatternFill())

        for ci, lpv in enumerate(lp_vals):
            col_idx  = lp_start + ci
            cell_sig = _sig_test(cc_vals[ci], lpv, lp_threshold)
            display  = round(lpv, 1) if np.isfinite(lpv) else None
            c = ws.cell(row=excel_row, column=col_idx, value=display)
            c.font = Font(name='Courier New', size=9, bold=cell_sig,
                          color='CC0000' if cell_sig else '000000')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr
            c.fill = (PatternFill('solid', start_color='FFE8E8') if cell_sig else
                      row_fill if row_fill else PatternFill())

        ws.row_dimensions[excel_row].height = 14

    ws.freeze_panes = 'A3'
    log(f"  Roster sheet: {len(medoid_list)} medoids, {len(cc_lp_pairs)} CC/LP pairs")



# -- Pattern clustering helpers -----------------------------------------------
def cluster_su_patterns(patterns, counts, n_clusters=None):
    """
    Cluster SU patterns using Jaccard distance + Ward linkage.
    Jaccard ignores shared-U positions and clusters only on shared S positions:
      d(i,j) = 1 - |intersect(S_i, S_j)| / |union(S_i, S_j)|
    Two patterns that are significant in exactly the same positions get d=0.
    Auto-cut at largest dendrogram gap if n_clusters not specified.
    Returns: Z (linkage), labels (1-indexed cluster per pattern), n_auto
    """
    n = len(patterns)
    if n <= 1:
        return None, np.ones(n, dtype=int), 1

    X = np.array([[1 if c == 'S' else 0 for c in p] for p in patterns], dtype=float)

    # Jaccard distance: ignores shared-U positions, clusters only on shared S
    # d(i,j) = 1 - |intersect(S)| / |union(S)|
    dist_mat = np.zeros((n, n))
    for i in range(n):
        for j in range(i + 1, n):
            intersect = float(np.sum((X[i] == 1) & (X[j] == 1)))
            union     = float(np.sum((X[i] == 1) | (X[j] == 1)))
            jac = 1.0 - (intersect / union if union > 0 else 1.0)
            dist_mat[i, j] = dist_mat[j, i] = jac
    np.fill_diagonal(dist_mat, 0.0)

    Z = linkage(squareform(dist_mat), method='ward', optimal_ordering=True)

    # Auto-cut: largest gap in merge distances (excluding the final merge)
    dists = Z[:, 2]
    gaps  = np.diff(dists)
    auto_k = n - int(np.argmax(gaps[:-1])) - 1  # exclude last gap (trivial 2-cluster)
    # clamp to sensible range
    auto_k = max(2, min(auto_k, min(20, n // 3)))

    k_use = n_clusters if n_clusters and n_clusters >= 2 else auto_k
    k_use = min(k_use, n)
    labels = fcluster(Z, t=k_use, criterion='maxclust')
    return Z, labels, auto_k

# -- Step 7: CC_BS pattern analysis sheet ------------------------------------
def make_pattern_analysis_sheet(ws, medoid_list, death_cc_df, lp_threshold, k, n_pattern_clusters=None):
    """Full CC_BS pattern analysis for the medoids at this k."""
    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
    cc_lp_pairs = [(cc, lp) for cc, lp in PANDEMIC_PAIRS
                   if cc in dc.columns and lp in dc.columns]

    year_short = {'2020': '20', '2021': '21', '2022': '22', '2023': '23', '2024': '24'}
    age_short  = {'': 'A', '_GE65': 'G', '_LT65': 'L'}
    col_labels = []
    for cc_col, _ in cc_lp_pairs:
        base = cc_col.replace('asedx_p_', '')
        for yr, ys in year_short.items():
            if base.startswith(yr):
                ag = age_short.get(base[len(yr):], base[len(yr):])
                col_labels.append(f'{ys}{ag}')
                break
        else:
            col_labels.append(cc_col)

    # Build per-medoid records
    records = []
    for m in medoid_list:
        if m not in dc.index:
            continue
        row_data = dc.loc[m]
        su, pm = '', ''
        n_sig, lp_sig, cc_sig = 0, [], []
        for (cc_col, lp_col), lbl in zip(cc_lp_pairs, col_labels):
            lp  = float(row_data.get(lp_col, float('nan')))
            ccv = float(row_data.get(cc_col, float('nan')))
            sig = _sig_test(ccv, lp, lp_threshold)
            su += 'S' if sig else 'U'
            pm += ('+' if ccv > 0 else '-') if sig else '.'
            if sig:
                n_sig += 1; lp_sig.append(lp); cc_sig.append(ccv)
        rec = {
            'metric': m, 'su': su, 'pm': pm, 'n_sig': n_sig,
            'best_lp':     min(lp_sig) if lp_sig else np.nan,
            'sum_abs_lp':  sum(abs(l) for l in lp_sig),
            'mean_abs_cc': float(np.mean([abs(c) for c in cc_sig])) if cc_sig else np.nan,
            'pct_pos':     100.0 * sum(1 for c in cc_sig if c > 0) / len(cc_sig) if cc_sig else np.nan,
        }
        for i, (sig_ch, lbl) in enumerate(zip(su, col_labels)):
            rec[f'S_{lbl}'] = 1 if sig_ch == 'S' else 0
        records.append(rec)

    df_r  = pd.DataFrame(records)
    n_tot = len(df_r)
    sig   = df_r[df_r['n_sig'] > 0].copy()
    n_sig_any = len(sig)
    pos_cols  = [f'S_{l}' for l in col_labels]

    # Pattern summary table
    if n_sig_any > 0:
        pat = sig.groupby('su').agg(
            n_metrics   = ('metric', 'count'),
            max_n_sig   = ('n_sig', 'max'),
            mean_n_sig  = ('n_sig', 'mean'),
            best_lp     = ('best_lp', 'min'),
            mean_sum_lp = ('sum_abs_lp', 'mean'),
            mean_abs_cc = ('mean_abs_cc', 'mean'),
            pct_pos     = ('pct_pos', 'mean'),
        ).sort_values('n_metrics', ascending=False).reset_index()
        pat['cum_pct'] = (pat['n_metrics'].cumsum() / pat['n_metrics'].sum() * 100).round(1)
        n_patterns = len(pat)
        n50 = int((pat['cum_pct'] >= 50).idxmax()) + 1
        n80 = int((pat['cum_pct'] >= 80).idxmax()) + 1
        pos_freq = sig[pos_cols].mean() * 100
        # cluster the patterns
        pat_counts = dict(zip(pat['su'], pat['n_metrics']))
        pat_list   = pat['su'].tolist()
        Z_pat, pat_cluster_labels, auto_k = cluster_su_patterns(pat_list, pat_counts, n_pattern_clusters)
        pat['pattern_cluster'] = pat_cluster_labels
        pat['pattern_cluster_auto_k'] = auto_k
        used_k = int(pat_cluster_labels.max()) if len(pat_cluster_labels) else 0
    else:
        pat = pd.DataFrame()
        n_patterns = n50 = n80 = 0
        pos_freq = pd.Series(0.0, index=pos_cols)
        Z_pat = None
        pat_cluster_labels = np.array([], dtype=int)
        auto_k = 0
        used_k = 0

    # n_sig distribution
    nsig_dist = df_r['n_sig'].value_counts().sort_index()

    # --- Styles ---
    norm     = Font(name='Arial', size=9)
    bold     = Font(name='Arial', bold=True, size=9)
    mono     = Font(name='Courier New', size=9)
    bold_hdr = Font(name='Arial', bold=True, size=10)
    red_bold = Font(name='Arial', bold=True, size=9, color='CC0000')
    thin     = Side(style='thin', color='CCCCCC')
    med      = Side(style='medium', color='888888')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    sect_bdr = Border(left=med,  right=med,  top=med,  bottom=med)
    hdr_fill = PatternFill('solid', start_color='DDEEFF')
    grn_fill = PatternFill('solid', start_color='D0EED0')
    yel_fill = PatternFill('solid', start_color='FFFACC')
    alt_fill = PatternFill('solid', start_color='F5F8FF')
    red_fill = PatternFill('solid', start_color='FFE8E8')

    def write_cell(row, col, value, font=None, fill=None, align='left',
                   border=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font   = font   or norm
        c.fill   = fill   or PatternFill()
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        c.border = border or bdr
        if num_fmt: c.number_format = num_fmt
        return c

    def section_title(row, col, text, span=8):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name='Arial', bold=True, size=11)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = sect_bdr
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + span - 1)
        ws.row_dimensions[row].height = 20
        return row + 1

    crow = 1  # current row pointer

    # ---- SECTION 1: Summary statistics ----
    crow = section_title(crow, 1, f'CC_BS Pattern Analysis  --  Ward k={k}  --  LP threshold={lp_threshold}', span=12)
    pairs_summary = [
        ('Total medoids (k)',        n_tot),
        ('Medoids with >=1 sig CC',  n_sig_any),
        ('% significant',            f'{100*n_sig_any/n_tot:.1f}%' if n_tot else 'N/A'),
        ('Unique SU patterns',        n_patterns),
        ('Pattern clusters (used)',   used_k),
        ('Pattern clusters (auto-cut suggestion)', auto_k),
        ('Patterns covering 50% of sig medoids', n50),
        ('Patterns covering 80% of sig medoids', n80),
        ('Max n_sig (any medoid)',    int(df_r['n_sig'].max()) if n_tot else 0),
        ('Mean n_sig (sig medoids only)', f"{sig['n_sig'].mean():.2f}" if n_sig_any else 'N/A'),
    ]
    for label, value in pairs_summary:
        write_cell(crow, 1, label, font=bold, border=bdr)
        write_cell(crow, 2, value, font=norm, align='right', border=bdr)
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 14
        crow += 1

    crow += 1  # blank row

    # ---- SECTION 2: n_sig distribution ----
    crow = section_title(crow, 1, 'Distribution of n_sig (number of significant positions per medoid)', span=6)
    write_cell(crow, 1, 'n_sig', font=bold_hdr, fill=hdr_fill, align='center')
    write_cell(crow, 2, 'n_medoids', font=bold_hdr, fill=hdr_fill, align='center')
    write_cell(crow, 3, 'cum_%', font=bold_hdr, fill=hdr_fill, align='center')
    ws.column_dimensions['C'].width = 10
    crow += 1
    cum = 0
    for ns, cnt in nsig_dist.items():
        cum += cnt
        fill = red_fill if ns > 0 else alt_fill
        write_cell(crow, 1, int(ns),  font=mono, align='center', fill=fill)
        write_cell(crow, 2, int(cnt), font=norm, align='right',  fill=fill)
        write_cell(crow, 3, f'{100*cum/n_tot:.1f}%', font=norm, align='right', fill=fill)
        crow += 1

    crow += 1

    # ---- SECTION 3: Position S-frequency heatmap ----
    crow = section_title(crow, 1, 'S-frequency per position (% of sig medoids that are S at each year-age)', span=len(col_labels)+2)
    write_cell(crow, 1, 'position', font=bold_hdr, fill=hdr_fill, align='center')
    write_cell(crow, 2, '% S',      font=bold_hdr, fill=hdr_fill, align='center')
    crow += 1
    for lbl, pct in zip(col_labels, pos_freq.values):
        intensity = min(int(pct * 2.55), 255)
        hex_color = f'FF{255-intensity:02X}{255-intensity:02X}'
        cell_fill = PatternFill('solid', start_color=hex_color)
        write_cell(crow, 1, lbl,           font=mono,      align='center', fill=cell_fill)
        write_cell(crow, 2, round(pct, 1), font=norm, align='right',  fill=cell_fill)
        crow += 1

    crow += 1

    # ---- SECTION 4: Pattern frequency table ----
    crow = section_title(crow, 1,
        f'Pattern frequency table ({n_patterns} unique SU patterns among {n_sig_any} sig medoids)',
        span=10)
    if n_sig_any > 0:
        PAT_HDRS = [
            ('pat_cluster', 8, 'center', norm),
            ('SU_pattern', 18, 'left',    mono),
            ('n_medoids',   9, 'right',   norm),
            ('cum_%',       7, 'right',   norm),
            ('max_n_sig',   8, 'center',  norm),
            ('mean_n_sig',  9, 'center',  norm),
            ('best_LP',     8, 'right',   mono),
            ('mean_sum|LP|',10, 'right',  norm),
            ('mean|CC|',    8, 'right',   norm),
            ('%_pos_CC',    8, 'right',   norm),
        ]
        for ci, (hdr, width, align, _) in enumerate(PAT_HDRS):
            col_idx = ci + 1
            write_cell(crow, col_idx, hdr, font=bold_hdr, fill=hdr_fill, align='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        crow += 1

        for ridx, prow in pat.iterrows():
            fill = alt_fill if ridx % 2 == 1 else None
            # colour by n_sig: more S -> more red tint
            ns = prow['max_n_sig']
            intensity = min(int(ns / 15 * 200), 200)
            row_fill = PatternFill('solid', start_color=f'FF{255-intensity:02X}{255-intensity:02X}') if ns >= 3 else (fill or PatternFill())
            vals = [
                int(prow['pattern_cluster']),
                prow['su'],
                int(prow['n_metrics']),
                f"{prow['cum_pct']:.1f}%",
                int(prow['max_n_sig']),
                round(prow['mean_n_sig'], 1),
                round(prow['best_lp'], 1),
                round(prow['mean_sum_lp'], 1),
                round(prow['mean_abs_cc'], 3),
                f"{prow['pct_pos']:.0f}%",
            ]
            for ci, (val, (_, _, align, fnt)) in enumerate(zip(vals, PAT_HDRS)):
                write_cell(crow, ci+1, val, font=fnt, align=align, fill=row_fill)
            crow += 1

    ws.freeze_panes = 'A1'
    log(f"  Analysis sheet: {n_patterns} patterns, {n_sig_any}/{n_tot} sig medoids")

# -- Step 5: Excel CC_BS sig-pattern analysis ----------------------------------
def make_ccbs_excel(wb, sig_medoids, death_cc_df, explain, k400_map,
                    lp_threshold, top_n, k, out_dir):
    if not sig_medoids:
        log("  No significant medoids -- skipping Excel output")
        return

    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
    lp_cols = [lp for _, lp in PANDEMIC_PAIRS if lp in dc.columns]
    cc_cols = [cc for cc, lp in PANDEMIC_PAIRS if cc in dc.columns]

    # Build sig/dir bit strings for significant medoids only
    records = []
    for m in sig_medoids:
        if m not in dc.index:
            continue
        row = dc.loc[m]
        s, d = '', ''
        for cc_col, lp_col in PANDEMIC_PAIRS:
            if lp_col not in dc.columns:
                continue
            lp  = row.get(lp_col)
            ccv = row.get(cc_col, 0.0)
            sig = '1' if _sig_test(ccv, lp, lp_threshold) else '0'
            s  += sig
            d  += ('1' if ccv > 0 else '0') if sig == '1' else '0'
        records.append({'metric': m, 'sig_bits': s, 'dir_bits': d})

    df_bits = pd.DataFrame(records)
    top_sigs = df_bits['sig_bits'].value_counts().head(top_n).index.tolist()

    def to_su(s):
        return ''.join('S' if c == '1' else 'U' for c in s)

    def to_pm(sig_raw, d):
        return ''.join(
            ('+' if c == '1' else '-') if s == '1' else '.'
            for s, c in zip(sig_raw, d)
        )

    rows = []
    for sig in top_sigs:
        sub  = df_bits[df_bits['sig_bits'] == sig]
        n_p  = len(sub)
        dirs = sub['dir_bits'].value_counts()
        dir_lines = [f"{to_pm(sig, dv)} (n={cnt})" for dv, cnt in dirs.items()]

        metrics = sub['metric'].tolist()
        names   = [clean_desc(explain.get(m, m)) for m in metrics]
        k400_labels = [k400_map.get(m, '') for m in metrics]

        patterns_cell = '\n'.join([to_su(sig)] + dir_lines)
        rows.append({
            'patterns':     patterns_cell,
            'n':            n_p,
            'metric_names': ' | '.join(names),
            'k400_labels':  ' | '.join(dict.fromkeys(k for k in k400_labels if k)),
        })

    out_df = pd.DataFrame(rows)

    ws = wb.create_sheet(title=f'CC_BS k={k}')

    mono     = Font(name='Courier New', size=10)
    norm     = Font(name='Arial', size=10)
    bold_hdr = Font(name='Arial', bold=True, size=10)
    thin     = Side(style='thin', color='AAAAAA')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', start_color='DDEEFF')
    alt_fill = PatternFill('solid', start_color='F5F8FF')

    COLS = [
        (1, 'patterns',     22),
        (2, 'n',             6),
        (3, 'metric_names', 80),
        (4, 'k400_labels',  60),
    ]
    for cidx, hdr, width in COLS:
        c = ws.cell(row=1, column=cidx, value=hdr)
        c.font = bold_hdr
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        ws.column_dimensions[get_column_letter(cidx)].width = width
    ws.row_dimensions[1].height = 20

    for ridx, row in out_df.iterrows():
        excel_row = ridx + 2
        fill      = alt_fill if ridx % 2 == 1 else None
        n_lines   = row['patterns'].count('\n') + 1

        c = ws.cell(row=excel_row, column=1, value=row['patterns'])
        c.font = mono
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.border = bdr
        if fill: c.fill = fill

        for cidx, attr, halign in [(2, 'n', 'right'), (3, 'metric_names', 'left'), (4, 'k400_labels', 'left')]:
            val = row[attr]
            c = ws.cell(row=excel_row, column=cidx, value=val)
            c.font = norm
            c.alignment = Alignment(horizontal=halign, vertical='top', wrap_text=True)
            c.border = bdr
            if fill: c.fill = fill
        ws.row_dimensions[excel_row].height = max(40, n_lines * 14 + 6)

    ws.freeze_panes = 'A2'
    log(f"  CC_BS sheet: {len(out_df)} pattern rows")


# -- Step 8: Pattern cluster dendrogram PNG -----------------------------------
def plot_pattern_dendrogram(medoid_list, death_cc_df, explain, k400_map,
                             lp_threshold, n_pattern_clusters, k, out_dir):
    """Plot dendrogram of the unique SU patterns, coloured by pattern cluster."""
    dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
    cc_lp_pairs = [(cc, lp) for cc, lp in PANDEMIC_PAIRS
                   if cc in dc.columns and lp in dc.columns]

    year_short = {'2020': '20', '2021': '21', '2022': '22', '2023': '23', '2024': '24'}
    age_short  = {'': 'A', '_GE65': 'G', '_LT65': 'L'}
    col_labels = []
    for cc_col, _ in cc_lp_pairs:
        base = cc_col.replace('asedx_p_', '')
        for yr, ys in year_short.items():
            if base.startswith(yr):
                ag = age_short.get(base[len(yr):], base[len(yr):])
                col_labels.append(f'{ys}{ag}')
                break
        else:
            col_labels.append(cc_col)

    # Build pattern counts from medoid list
    counts = {}
    for m in medoid_list:
        if m not in dc.index:
            continue
        su = ''.join('S' if _sig_test(dc.loc[m].get(cc_col), dc.loc[m].get(lp), lp_threshold)
                     else 'U'
                     for cc_col, lp in cc_lp_pairs)
        if 'S' in su:
            counts[su] = counts.get(su, 0) + 1

    if len(counts) < 2:
        log("  Too few patterns for dendrogram -- skipping")
        return

    patterns = sorted(counts.keys())
    Z_pat, pat_labels, auto_k = cluster_su_patterns(patterns, counts, n_pattern_clusters)
    used_k = int(pat_labels.max())

    # Build dendrogram labels: SU string + count
    dn_labels = [f"(n={counts[p]:3d}) {p}" for p in patterns]

    # Colour branches by cluster
    node_color = build_node_colors(Z_pat, len(patterns), pat_labels, PALETTE)
    def link_color_func(link_id):
        return node_color.get(link_id, '#555555')

    n = len(patterns)
    fig_h = max(6, n * 0.22)
    fig, ax = plt.subplots(figsize=(18, fig_h), dpi=130)

    dn = dendrogram(Z_pat, labels=dn_labels, orientation='left',
                    leaf_font_size=8, link_color_func=link_color_func,
                    above_threshold_color='#555555', ax=ax)
    for line in ax.get_lines():
        line.set_linewidth(line.get_linewidth() * 2.0)

    ax.set_xlabel('Ward linkage distance (Jaccard on S-positions)', fontsize=11)
    ax.tick_params(axis='x', labelsize=9)

    # Colour leaf labels by cluster
    fig.canvas.draw()
    ylbls      = ax.get_yticklabels()
    leaf_order = dn['leaves']
    for pos_i, leaf_idx in enumerate(leaf_order):
        cid = pat_labels[leaf_idx]
        col = PALETTE[(cid - 1) % len(PALETTE)]
        ylbls[pos_i].set_color(col)

    # Legend
    cluster_handles = [Patch(facecolor=PALETTE[(i) % len(PALETTE)],
                             label=f'Cluster {i+1}')
                       for i in range(used_k)]
    ax.legend(handles=cluster_handles, loc='lower right', fontsize=9,
              title=f'{used_k} pattern clusters (auto={auto_k})',
              title_fontsize=9, framealpha=0.85)

    ax.set_title(f'SU Pattern Dendrogram  |  {n} unique patterns  |  '
                 f'Ward k={k} medoids  |  LP<={lp_threshold}  |  '
                 f'Jaccard distance on S-positions',
                 fontsize=11, pad=10)

    plt.tight_layout()
    outpath = os.path.join(out_dir, f'pattern_dendrogram_ward{k}.png')
    fig.savefig(outpath, dpi=130, bbox_inches='tight')
    plt.close()
    log(f"  Saved: {outpath}")


# -- Year pattern ranks sheet (multi-k comparison) ----------------------------
def make_year_pattern_sheet(ws, k_results, lp_threshold, top_n=10):
    """
    k_results: dict of k -> {'medoid_list': [...], 'death_cc_df': df}
    Builds a side-by-side table of top-N year patterns for each k.
    """
    from openpyxl.utils import get_column_letter

    def get_year_counts(medoid_list, death_cc_df):
        dc = death_cc_df.set_index('metric') if 'metric' in death_cc_df.columns else death_cc_df
        counts = {}
        for m in medoid_list:
            if m not in dc.index: continue
            row = dc.loc[m]
            su15 = ''.join(
                'S' if (np.isfinite(float(row.get(lp, float('nan')))) and
                        float(row.get(lp, float('nan'))) <= lp_threshold) else 'U'
                for _, lp in PANDEMIC_PAIRS if lp in dc.columns
            )
            if 'S' not in su15: continue
            yr_pat = ''.join('S' if 'S' in su15[yi*3:(yi+1)*3] else 'U' for yi in range(5))
            counts[yr_pat] = counts.get(yr_pat, 0) + 1
        return counts

    norm     = Font(name='Arial', size=10)
    mono     = Font(name='Courier New', size=10)
    bold_hdr = Font(name='Arial', bold=True, size=10)
    bold_ttl = Font(name='Arial', bold=True, size=12)
    thin     = Side(style='thin', color='CCCCCC')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', start_color='DDEEFF')
    s_fill   = PatternFill('solid', start_color='FFE0E0')
    alt_fill = PatternFill('solid', start_color='F5F8FF')
    k_fills  = [
        PatternFill('solid', start_color='E8F5E9'),
        PatternFill('solid', start_color='E3F2FD'),
        PatternFill('solid', start_color='FFF3E0'),
        PatternFill('solid', start_color='F3E5F5'),
        PatternFill('solid', start_color='FCE4EC'),
    ]

    def wc(row, col, val, font=None, fill=None, align='center', border=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = font or norm
        c.fill   = fill or PatternFill()
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border = border or bdr
        return c

    COL_W   = [5, 10, 5, 6, 6, 4, 4, 4, 4, 4]
    NCOLS   = len(COL_W)
    GAP     = 1
    k_list  = sorted(k_results.keys())

    # Row 1: title
    total_span = len(k_list) * (NCOLS + GAP) - GAP
    c = ws.cell(row=1, column=1,
        value=f'Year Pattern Occupancy -- Top {top_n} Patterns by k  (LP threshold = {lp_threshold})')
    c.font = bold_ttl
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_span)
    ws.row_dimensions[1].height = 22

    block_starts = [1 + bi * (NCOLS + GAP) for bi in range(len(k_list))]

    # Row 2: k group headers
    all_top10 = {}
    for bi, k in enumerate(k_list):
        ml   = k_results[k]['medoid_list']
        dcdf = k_results[k]['death_cc_df']
        counts = get_year_counts(ml, dcdf)
        total  = sum(counts.values())
        n_uniq = len(counts)
        sorted_pats = sorted(counts.items(), key=lambda x: -x[1])
        rows = []
        cum = 0
        for rank, (pat, n) in enumerate(sorted_pats[:top_n], 1):
            cum += n
            rows.append({
                'rank': rank, 'pattern': pat, 'n': n,
                'pct': round(100*n/total, 1),
                'cum_pct': round(100*cum/total, 1),
                'y20': 'S' if pat[0]=='S' else '.',
                'y21': 'S' if pat[1]=='S' else '.',
                'y22': 'S' if pat[2]=='S' else '.',
                'y23': 'S' if pat[3]=='S' else '.',
                'y24': 'S' if pat[4]=='S' else '.',
                'total_sig': total, 'n_unique': n_uniq,
            })
        all_top10[k] = rows
        bs = block_starts[bi]
        kf = k_fills[bi % len(k_fills)]
        c = ws.cell(row=2, column=bs,
            value=f'k={k}  |  sig: {total}  |  unique: {n_uniq}/32')
        c.font = Font(name='Arial', bold=True, size=10)
        c.fill = kf
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=2, start_column=bs, end_row=2, end_column=bs+NCOLS-1)
        ws.row_dimensions[2].height = 18
        # col headers row 3
        for ci, hdr in enumerate(['rank','pattern','n','%','cum%','2020','2021','2022','2023','2024']):
            wc(3, bs+ci, hdr, font=bold_hdr, fill=hdr_fill)
        for ci, w in enumerate(COL_W):
            ws.column_dimensions[get_column_letter(bs+ci)].width = w
        # gap col
        if bi < len(k_list)-1:
            ws.column_dimensions[get_column_letter(bs+NCOLS)].width = 2
    ws.row_dimensions[3].height = 16

    # Data rows
    for row_i in range(top_n):
        excel_row = row_i + 4
        base_fill = alt_fill if row_i % 2 == 1 else None
        for bi, k in enumerate(k_list):
            rows = all_top10[k]
            if row_i >= len(rows): continue
            r = rows[row_i]
            bs = block_starts[bi]
            bf = base_fill or PatternFill()
            wc(excel_row, bs+0, r['rank'],          font=norm, fill=bf, align='center')
            wc(excel_row, bs+1, r['pattern'],        font=mono, fill=bf, align='left')
            wc(excel_row, bs+2, r['n'],              font=norm, fill=bf, align='right')
            wc(excel_row, bs+3, f"{r['pct']}%",     font=norm, fill=bf, align='right')
            wc(excel_row, bs+4, f"{r['cum_pct']}%", font=norm, fill=bf, align='right')
            for yi, yk in enumerate(['y20','y21','y22','y23','y24']):
                cf = s_fill if r[yk]=='S' else bf
                wc(excel_row, bs+5+yi, r[yk], font=mono, fill=cf, align='center')
        ws.row_dimensions[excel_row].height = 14

    ws.freeze_panes = 'A4'
    log(f"  Year pattern sheet: {len(k_list)} k values, top {top_n} patterns each")

# -- Main ----------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Ward medoid CC_BS analysis pipeline')
    parser.add_argument('--cc-matrix',      required=True,  help='Square CC matrix CSV or CSV.GZ')
    parser.add_argument('--death-cc',       required=True,  help='metric x death CC/LP CSV')
    parser.add_argument('--explain',        default=None,   help='Metric explain file (optional)')
    parser.add_argument('--hub',            default=None,   help='hub_combined_k400_labels.csv (optional)')
    parser.add_argument('--ward-k',         type=int,   default=323,
                        help='XDE clusters for single run (default: 323)')
    parser.add_argument('--ward-k-list',    type=str,   default=None,
                        help='Comma-separated list of k values for multi-k run, e.g. 150,323,2639')
    parser.add_argument('--lp-threshold',   type=float, default=-14.0, help='LP significance threshold (default: -14.0)')
    parser.add_argument('--sig-mode',       choices=['lp', 'cc'], default='cc',
                        help="Significance definition: 'cc' (|CC|>--cc-sig, default) or 'lp' (legacy)")
    parser.add_argument('--cc-sig',         type=float, default=0.3, help='|CC| Significant threshold (cc mode)')
    parser.add_argument('--top-n',          type=int,   default=10,  help='Top-N sig patterns for Excel (default: 10)')
    parser.add_argument('--nsuper',         type=int,   default=12,  help='Super-clusters for dendrogram (default: 12)')
    parser.add_argument('--linthresh',      type=float, default=0.5, help='Symlog linthresh (default: 0.5)')
    parser.add_argument('--k400',           action='store_true', help='Use k400 labels in dendrogram')
    parser.add_argument('--out-dir',        default='.',  help='Output directory (default: .)')
    parser.add_argument('--pattern-clusters', type=int, default=None,
                        help='Number of SU pattern clusters (default: auto from dendrogram gap)')
    parser.add_argument('--skip-precompute', action='store_true',
                        help='Skip full XDE clustering; load Z_ward.npy + cc_sub.npy + metrics_list.json and recut at --ward-k')
    args = parser.parse_args()

    global SIG_MODE, CC_SIG
    SIG_MODE = args.sig_mode
    CC_SIG   = args.cc_sig
    if SIG_MODE == 'cc':
        print(f"Significance mode: cc  (|CC|>{CC_SIG})")

    if args.lp_threshold >= 0:
        print("ERROR: --lp-threshold must be negative (e.g. -14.0)")
        raise SystemExit(1)

    os.makedirs(args.out_dir, exist_ok=True)

    # -- Parse k list ----------------------------------------------------------
    if args.ward_k_list:
        k_list = [int(x.strip()) for x in args.ward_k_list.split(',')]
    else:
        k_list = [args.ward_k]
    log(f"k values to run: {k_list}")

    # -- Load death CC file ----------------------------------------------------
    log("Loading death CC/LP file...")
    death_cc_df = pd.read_csv(args.death_cc)
    log(f"  {len(death_cc_df)} metrics x {len(death_cc_df.columns)} columns")

    # -- Load optional metadata ------------------------------------------------
    if args.explain and os.path.isfile(args.explain):
        exp_df  = pd.read_csv(args.explain)
        explain = dict(zip(exp_df.iloc[:, 0].str.strip(), exp_df.iloc[:, 1]))
    else:
        explain = {}
        log("  No explain file -- metric codes will be used as labels")

    if args.hub and os.path.isfile(args.hub):
        hub_df = pd.read_csv(args.hub)
        if 'member metric' in hub_df.columns:
            hub_df = hub_df.rename(columns={'member metric': 'metric'})
        k400_map = dict(zip(hub_df['metric'], hub_df['label_k400']))
    else:
        k400_map = {}
        log("  No hub file -- k400 labels will be empty")

    # -- Step 1: Full XDE clustering (run once) --------------------------------
    if args.skip_precompute:
        log("--skip-precompute: loading Z_ward.npy, cc_sub.npy, metrics_list.json...")
        Z_ward = np.load(os.path.join(args.out_dir, 'Z_ward.npy'))
        cc_sub = np.load(os.path.join(args.out_dir, 'cc_sub.npy'))
        with open(os.path.join(args.out_dir, 'metrics_list.json')) as f:
            ml = json.load(f)
        metrics      = ml['metrics']
        best_cc_list = ml['best_cc']
        log(f"  Loaded {len(metrics)} metrics.")
    else:
        log("Step 1: Full XDE clustering (slow -- runs once)...")
        Z_ward, cc_sub, metrics, best_cc_list = compute_ward_full(args.cc_matrix, args.out_dir)

    # -- Multi-k workbook: one workbook with all k values ----------------------
    multi_k_wb   = Workbook()
    multi_k_wb.remove(multi_k_wb.active)
    k_results    = {}   # for year pattern sheet

    # -- Loop over k values ----------------------------------------------------
    for k in k_list:
        log(f"=== k={k} ===")

        # Step 2: cut + medoids
        log(f"  Step 2: Cut at k={k}, find medoids...")
        Z_med, medoid_list, reps_df = compute_medoids_from_ward(
            Z_ward, cc_sub, metrics, best_cc_list, k, args.out_dir)

        # Step 3: filter sig
        log(f"  Step 3: Filtering significant medoids...")
        sig_medoids = filter_sig_medoids(medoid_list, death_cc_df,
                                          args.lp_threshold, args.out_dir, k)

        # Step 4: medoid dendrogram PNG
        log(f"  Step 4: Plotting medoid dendrogram...")
        plot_dendrogram(Z_med, medoid_list, reps_df, death_cc_df, explain, k400_map,
                        args.lp_threshold, args.nsuper, args.linthresh,
                        args.k400, k, args.out_dir)

        # Step 5: CC_BS sheet
        log(f"  Step 5: CC_BS sig-pattern sheet...")
        make_ccbs_excel(multi_k_wb, sig_medoids, death_cc_df, explain, k400_map,
                        args.lp_threshold, args.top_n, k, args.out_dir)

        # Step 6: roster sheet
        log(f"  Step 6: Full medoid roster sheet...")
        ws_roster = multi_k_wb.create_sheet(title=f'Roster k={k}')
        make_medoid_roster_sheet(ws_roster, medoid_list, reps_df, death_cc_df,
                                  explain, k400_map, args.lp_threshold, k)

        # Step 7: pattern analysis sheet
        log(f"  Step 7: Pattern analysis sheet...")
        ws_analysis = multi_k_wb.create_sheet(title=f'Patterns k={k}')
        make_pattern_analysis_sheet(ws_analysis, medoid_list, death_cc_df,
                                     args.lp_threshold, k, args.pattern_clusters)

        # Step 8: pattern cluster dendrogram PNG
        log(f"  Step 8: Pattern cluster dendrogram...")
        plot_pattern_dendrogram(medoid_list, death_cc_df, explain, k400_map,
                                 args.lp_threshold, args.pattern_clusters, k, args.out_dir)

        # stash for year pattern sheet
        k_results[k] = {'medoid_list': medoid_list, 'death_cc_df': death_cc_df}

    # -- Step 9: Year pattern ranks (multi-k comparison) -----------------------
    if len(k_list) > 1:
        log("Step 9: Year pattern ranks comparison sheet...")
        ws_yrpat = multi_k_wb.create_sheet(title='Year Pattern Ranks')
        make_year_pattern_sheet(ws_yrpat, k_results, args.lp_threshold, args.top_n)

    # -- Save combined workbook ------------------------------------------------
    k_tag    = '_'.join(str(k) for k in k_list)
    out_path = os.path.join(args.out_dir, f'medoid_analysis_k{k_tag}.xlsx')
    multi_k_wb.save(out_path)
    _tee(out_path)
    log(f"Saved: {out_path}")
    log(f"Sheets: {multi_k_wb.sheetnames}")
    log("Done.")

if __name__ == '__main__':
    main()
