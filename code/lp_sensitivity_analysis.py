import os
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
#!/usr/bin/env python3
"""
lp_sensitivity_analysis.py
===========================
LP threshold sensitivity analysis for the SCE_USCounties project.

For each LP threshold in a range, computes the number of metrics that are
significant across each of the 15 pandemic-era death measures (5 years x 3
age groups: All, GE65, LT65).

Also produces a summary comparing key findings at LP=-11, -13, -15.

Outputs:
  lp_threshold_counts.csv         -- n significant per threshold per death measure
  fig_lp_threshold_sensitivity.png -- line plot of counts vs threshold
  lp_sensitivity_summary.txt      -- text summary of key findings at 3 thresholds

Usage:
  python3 code/lp_sensitivity_analysis.py --cc-file metric_x_death_cc_0_0_25_1.csv >& lp_sensitivity_analysis.log
  python3 code/lp_sensitivity_analysis.py --cc-file metric_x_death_cc_0_0_25_1.csv --master-xlsx master_xde_clusters_2745.xlsx >& lp_sensitivity_analysis.log
"""

import argparse
import os
import sys
import numpy as np
import pandas as pd
import warnings
import matplotlib
warnings.filterwarnings("ignore", category=UserWarning)
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from scipy import stats
from scipy.optimize import brentq

def save_fig(fig, path, dpi=200):
    """Save figure, adding the filename as a small label in the top margin."""
    import matplotlib.pyplot as plt
    fname = os.path.basename(path)
    fig.text(0.5, 0.995, fname, ha='center', va='top',
             fontsize=7, color='#888888', fontfamily='monospace',
             transform=fig.transFigure)
    fig.savefig(path, dpi=dpi, bbox_inches='tight', facecolor='white')
    try:
        _rel = os.path.relpath(path)
    except ValueError:
        _rel = path
    print('Saved ' + _rel, file=sys.stderr, flush=True)
    plt.close(fig)



YEARS    = ['2020', '2021', '2022', '2023', '2024']
COLS_15  = [
    'LP_asedx_p_2020',      'LP_asedx_p_2020_GE65', 'LP_asedx_p_2020_LT65',
    'LP_asedx_p_2021',      'LP_asedx_p_2021_GE65', 'LP_asedx_p_2021_LT65',
    'LP_asedx_p_2022',      'LP_asedx_p_2022_GE65', 'LP_asedx_p_2022_LT65',
    'LP_asedx_p_2023',      'LP_asedx_p_2023_GE65', 'LP_asedx_p_2023_LT65',
    'LP_asedx_p_2024',      'LP_asedx_p_2024_GE65', 'LP_asedx_p_2024_LT65',
]
SHORT_15 = [
    '2020 All', '2020 GE65', '2020 LT65',
    '2021 All', '2021 GE65', '2021 LT65',
    '2022 All', '2022 GE65', '2022 LT65',
    '2023 All', '2023 GE65', '2023 LT65',
    '2024 All', '2024 GE65', '2024 LT65',
]
THRESHOLDS = [-5,-6,-7,-8,-9,-10,-11,-12,-13,-14,-15,-16,-17,-18,-19,-20,-21,-22,-25,-30]

YEAR_COLORS = {
    '2020': '#E74C3C', '2021': '#E67E22',
    '2022': '#27AE60', '2023': '#2980B9', '2024': '#8E44AD',
}
LS_MAP = {'All': '-', 'GE65': '--', 'LT65': ':'}


def lp_from_cc(r, n):
    t = abs(r) * np.sqrt(n-2) / np.sqrt(1 - r**2)
    p = 2 * stats.t.sf(t, df=n-2)
    return np.log10(p) if p > 0 else -999


def min_cc_for_lp(lp_target, n=3032):
    return brentq(lambda r: lp_from_cc(r, n) - lp_target, 0.001, 0.999)


def load_master_multiyear(master_xlsx):
    """Load Multi_Year tab from master Excel."""
    from openpyxl import load_workbook
    from _workbook_style import detect_table_offset
    wb   = load_workbook(master_xlsx, data_only=True)
    ws   = wb['Multi_Year']
    hdr_row, hdr_col = detect_table_offset(ws)
    hdr = []
    c = hdr_col
    while c <= ws.max_column:
        v = ws.cell(hdr_row, c).value
        if v is None:
            break
        hdr.append(str(v))
        c += 1
    n = len(hdr)
    rows = [{hdr[i]: ws.cell(r, hdr_col + i).value for i in range(n)}
            for r in range(hdr_row + 1, ws.max_row + 1)]
    df   = pd.DataFrame(rows)
    for col in ['max_signed_cc', 'super_cluster_id', 'Ward100', 'rank_abs_cc']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def get_stats_at_threshold(cc_df, master_df, threshold):
    """Compute key statistics at a given LP threshold using master assignments."""
    LP_COLS = [f'LP_asedx_p_{yr}' for yr in YEARS]

    si_list = []
    for m in master_df['metric']:
        if m in cc_df.index:
            si = ''.join(
                'S' if (lp := cc_df.loc[m, lp_col]) is not None
                       and pd.notna(lp) and lp <= threshold
                else 'I'
                for lp_col in LP_COLS if lp_col in cc_df.columns
            )
        else:
            si = 'IIIII'
        si_list.append(si)

    master2         = master_df.copy()
    master2['SI_t'] = si_list
    master2['n_sig_t'] = [s.count('S') for s in si_list]
    my   = master2[master2['n_sig_t'] > 1]
    pos  = my[my['max_signed_cc'] > 0]
    neg  = my[my['max_signed_cc'] < 0]

    si_vc   = my['SI_t'].value_counts().to_dict()
    sc_sig  = {sc: int(len(my[my['super_cluster_id'] == sc])) for sc in sorted(my['super_cluster_id'].dropna().astype(int).unique())}
    sc_null = [sc for sc in sorted(my['super_cluster_id'].dropna().astype(int).unique()) if sc_sig[sc] == 0]

    # Race groups
    aian = my[my['explain'].str.contains(
        r'American.?Indian|Alaska.?Native', na=False, regex=True, case=False)]
    wnh  = my[my['explain'].str.contains(
        r'White.?non.?Hispanic', na=False, regex=True, case=False)]
    hisp = my[my['explain'].str.contains(
        r'Hispanic|Latino', na=False, regex=True, case=False)]
    hisp = hisp[~hisp['explain'].str.contains(
        r'White.?non.?Hispanic', na=False, regex=True, case=False)]

    return {
        'threshold':  threshold,
        'min_cc':     round(min_cc_for_lp(threshold), 4),
        'n_multiyear':len(my),
        'n_risky':    len(pos),
        'n_protective':len(neg),
        'pct_risky':  round(len(pos)/len(my)*100, 1) if len(my) else 0,
        'mean_cc_risky': round(float(pos['max_signed_cc'].mean()), 3) if len(pos) else None,
        'mean_cc_prot':  round(float(neg['max_signed_cc'].mean()), 3) if len(neg) else None,
        'SSSSS':      int(si_vc.get('SSSSS', 0)),
        'SSIII':      int(si_vc.get('SSIII', 0)),
        'SSSII':      int(si_vc.get('SSSII', 0)),
        'ISSSS':      int(si_vc.get('ISSSS', 0)),
        'ISSII':      int(si_vc.get('ISSII', 0)),
        'sc_sig':     sc_sig,
        'sc_null':    sc_null,
        'n_aian':     len(aian),
        'n_aian_risky':int((aian['max_signed_cc'] > 0).sum()),
        'n_wnh':      len(wnh),
        'n_wnh_prot': int((wnh['max_signed_cc'] < 0).sum()),
        'n_hisp':     len(hisp),
        'n_hisp_risky':int((hisp['max_signed_cc'] > 0).sum()),
    }


def write_summary(results, out_path, sc_names=None):
    """Write text summary comparing key thresholds."""
    lines = []
    lines.append("LP THRESHOLD SENSITIVITY ANALYSIS")
    lines.append("=" * 70)
    lines.append(f"n metrics: {results[list(results.keys())[0]]['n_multiyear'] + 999}")
    lines.append("")

    thresholds = sorted(results.keys())
    w = 12

    def row(label, key, fmt=None, sc_id=None):
        vals = []
        for t in thresholds:
            r = results[t]
            if sc_id is not None:
                v = r['sc_sig'].get(sc_id, '---')
            else:
                v = r.get(key, '---')
            if fmt and v not in ('---', None):
                try: v = fmt.format(v)
                except: pass
            vals.append(str(v) if v is not None else '---')
        lines.append(f"  {label:42s}" + "".join(f"{v:>{w}}" for v in vals))

    header = f"  {'':42s}" + "".join(f"{t:>{w}}" for t in thresholds)
    lines.append(header)
    lines.append("  " + "-" * (42 + w * len(thresholds)))

    lines.append("")
    lines.append("  --- Min |CC| required ---")
    row("Min |CC| for threshold",      'min_cc', '{:.4f}')

    lines.append("")
    lines.append("  --- Overall counts ---")
    row("n significant (any year)",    'n_multiyear')
    row("n risky (+CC)",               'n_risky')
    row("n protective (-CC)",          'n_protective')
    row("% risky",                     'pct_risky', '{:.1f}%')
    row("Mean CC risky",               'mean_cc_risky', '{:+.3f}')
    row("Mean CC protective",          'mean_cc_prot',  '{:+.3f}')

    lines.append("")
    lines.append("  --- SI patterns ---")
    row("SSSSS (all 5 years)",         'SSSSS')
    row("SSIII (2020-21 only)",        'SSIII')
    row("SSSII (2020-22)",             'SSSII')
    row("ISSSS (2021-24 lagged)",      'ISSSS')
    row("ISSII (2021-22)",             'ISSII')

    lines.append("")
    lines.append("  --- Super-cluster breakdown ---")
    sc_name_map = sc_names or {}
    all_scs = sorted(set().union(*[r['sc_sig'].keys() for r in results.values()]))
    for sc in all_scs:
        label = f"SC{sc:02d} {sc_name_map.get(sc, '')}"
        row(label, None, sc_id=sc)

    lines.append("")
    lines.append("  --- Null SCs (0 significant) ---")
    for t in thresholds:
        nulls = results[t]['sc_null']
        lines.append(f"  LP={t:4d}: SC{nulls}")

    lines.append("")
    lines.append("  --- Race groups ---")
    row("n AIAN significant",          'n_aian')
    row("  of which risky",            'n_aian_risky')
    row("n White non-Hisp sig",        'n_wnh')
    row("  of which protective",       'n_wnh_prot')
    row("n Hispanic sig",              'n_hisp')
    row("  of which risky",            'n_hisp_risky')

    text = "\n".join(lines)
    with open(out_path, 'w') as f:
        f.write(text)
    print(text)
    print(f"\nSaved: {out_path}")
    try:
        _r_out_path = os.path.relpath(out_path)
    except ValueError:
        _r_out_path = out_path
    print('Saved ' + _r_out_path, file=sys.stderr, flush=True)


YEAR_ORDER = ['2020', '2021', '2022', '2023', '2024']
AGE_ORDER  = ['All', 'GE65', 'LT65']

def write_excel_tab(df_counts, out_dir, compare_thresholds, n_metrics, bonf_lp):
    """Write LP sensitivity count table to Excel with colour formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LP_sensitivity_counts"

    # -- colour helpers --
    def hex_fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    YEAR_HEX = {'2020':'FADBD8','2021':'FDEBD0','2022':'D5F5E3','2023':'D6EAF8','2024':'E8DAEF'}
    COMPARE_HEX = 'FFD700'  # gold highlight for compare-threshold rows

    thin  = Side(style='thin', color='AAAAAA')
    med   = Side(style='medium', color='555555')
    thick = Side(style='thick', color='000000')
    def bord(top=None, bot=None, left=None, right=None):
        return Border(top=top or thin, bottom=bot or thin,
                      left=left or thin, right=right or thin)

    # -- header rows --
    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + len(SHORT_15))
    title_cell = ws.cell(1, 1)
    title_cell.value = (f"LP Threshold Sensitivity  |  {n_metrics} metrics  |  "
                        f"n=3,032 counties  |  Bonferroni LP={bonf_lp:.1f}")
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')

    # Row 2: blank
    ws.row_dimensions[2].height = 6

    # Row 3: "LP" + year super-headers (merged 3 cols each)
    ws.cell(3, 1).value = "LP"
    ws.cell(3, 1).font = Font(bold=True)
    ws.cell(3, 1).alignment = Alignment(horizontal='center', vertical='center')
    col = 2
    for yr in YEAR_ORDER:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+2)
        c = ws.cell(3, col)
        c.value = yr
        c.font = Font(bold=True, size=11)
        c.fill = hex_fill(YEAR_HEX[yr])
        c.alignment = Alignment(horizontal='center', vertical='center')
        col += 3

    # Row 4: age-group sub-headers
    ws.cell(4, 1).value = "threshold"
    ws.cell(4, 1).font = Font(bold=True, italic=True)
    ws.cell(4, 1).alignment = Alignment(horizontal='center')
    col = 2
    for yr in YEAR_ORDER:
        for ag in AGE_ORDER:
            c = ws.cell(4, col)
            c.value = ag
            c.font = Font(bold=True)
            c.fill = hex_fill(YEAR_HEX[yr])
            c.alignment = Alignment(horizontal='center')
            col += 1

    # Row 5: min |CC| sub-row
    ws.cell(5, 1).value = "min |CC|"
    ws.cell(5, 1).font = Font(italic=True, color='666666')
    ws.cell(5, 1).alignment = Alignment(horizontal='center')
    col = 2
    for lbl in SHORT_15:
        yr, ag = lbl[:4], lbl[5:]
        cc_val = min_cc_for_lp(float(df_counts.index[df_counts.index == -13][0])
                               if -13 in df_counts.index else -13)
        # just leave blank -- will fill per-column below if we want; skip for now
        col += 1

    # Data rows start at row 6
    data_row_start = 6
    for r_i, t in enumerate(df_counts.index):
        row = data_row_start + r_i
        # LP column
        lp_cell = ws.cell(row, 1)
        lp_cell.value = int(t)
        lp_cell.font = Font(bold=(t in compare_thresholds))
        lp_cell.alignment = Alignment(horizontal='center')

        is_compare = (t in compare_thresholds)
        fill_bg = hex_fill('FFF9C4') if is_compare else None  # light yellow

        col = 2
        for lbl in SHORT_15:
            yr = lbl[:4]
            cnt = int(df_counts.loc[t, lbl]) if lbl in df_counts.columns else 0
            c = ws.cell(row, col)
            c.value = cnt
            c.alignment = Alignment(horizontal='right')

            # colour by density: darker = more significant
            # scale: 0=white, max(col)=deep year colour
            col_max = int(df_counts[lbl].max()) if lbl in df_counts.columns else 1
            if col_max > 0:
                intensity = cnt / col_max  # 0..1
                # interpolate from white (FFFFFF) to year colour
                base = YEAR_HEX[yr]
                r_b = int(base[0:2], 16); g_b = int(base[2:4], 16); b_b = int(base[4:6], 16)
                r_f = int(255 + (r_b - 255) * intensity)
                g_f = int(255 + (g_b - 255) * intensity)
                b_f = int(255 + (b_b - 255) * intensity)
                c.fill = PatternFill("solid", fgColor=f"{r_f:02X}{g_f:02X}{b_f:02X}")

            if is_compare:
                c.font = Font(bold=True)
            col += 1

    # Column widths
    ws.column_dimensions['A'].width = 12
    for col_i in range(2, 2 + len(SHORT_15)):
        ws.column_dimensions[get_column_letter(col_i)].width = 7

    # Freeze panes below headers
    ws.freeze_panes = 'B6'

    out_path = os.path.join(out_dir, 'lp_sensitivity_counts.xlsx')
    wb.save(out_path)
    try:
        _r_out_path = os.path.relpath(out_path)
    except ValueError:
        _r_out_path = out_path
    print('Saved ' + _r_out_path, file=sys.stderr, flush=True)
    print(f"Saved: {out_path}")
    try:
        _r_out_path = os.path.relpath(out_path)
    except ValueError:
        _r_out_path = out_path
    print('Saved ' + _r_out_path, file=sys.stderr, flush=True)
    return out_path


def plot_count_heatmap(df_counts, out_dir, compare_thresholds):
    """Plot a 2D heatmap: rows=LP thresholds, cols=15 death measures."""
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import numpy as np

    # Build matrix: rows=thresholds (most negative at bottom), cols=15 measures
    thresholds_plot = sorted(df_counts.index, reverse=True)  # -5 .. -30 top to bottom
    cols_plot = [lbl for lbl in SHORT_15 if lbl in df_counts.columns]
    mat = df_counts.loc[thresholds_plot, cols_plot].values.astype(float)

    fig, ax = plt.subplots(figsize=(14, 10))

    im = ax.imshow(mat, aspect='auto', cmap='YlOrRd',
                   vmin=0, vmax=mat.max(), interpolation='none')

    # Annotate cells
    for i in range(mat.shape[0]):
        for j in range(mat.shape[1]):
            v = int(mat[i, j])
            ax.text(j, i, str(v), ha='center', va='center',
                    fontsize=7, color='black' if v < mat.max() * 0.7 else 'white')

    # x-axis: year groups with colour bands
    ax.set_xticks(range(len(cols_plot)))
    ax.set_xticklabels(cols_plot, rotation=45, ha='right', fontsize=9)

    # y-axis: LP thresholds
    ax.set_yticks(range(len(thresholds_plot)))
    ax.set_yticklabels([str(t) for t in thresholds_plot], fontsize=9)

    # Highlight compare-threshold rows
    for i, t in enumerate(thresholds_plot):
        if t in compare_thresholds:
            ax.axhline(i - 0.5, color='blue', lw=1.5, alpha=0.6)
            ax.axhline(i + 0.5, color='blue', lw=1.5, alpha=0.6)

    # Vertical year-group separators
    for k in range(1, 5):
        ax.axvline(k * 3 - 0.5, color='black', lw=1.5)

    # Year super-labels
    year_colours = {'2020':'#E74C3C','2021':'#E67E22','2022':'#27AE60','2023':'#2980B9','2024':'#8E44AD'}
    for k, yr in enumerate(YEAR_ORDER):
        mid = k * 3 + 1
        ax.text(mid, -1.5, yr, ha='center', va='center',
                fontsize=11, fontweight='bold', color=year_colours[yr],
                transform=ax.get_xaxis_transform())

    cb = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    cb.set_label('n significant metrics', fontsize=10)

    ax.set_xlabel('Death measure (year x age group)', fontsize=11)
    ax.set_ylabel('LP threshold', fontsize=11)
    ax.set_title('Number of significant metrics by LP threshold and death measure\n'
                 '(rows = LP threshold; cols = 5 years x 3 age groups;'
                 ' blue lines = comparison thresholds)',
                 fontsize=11, fontweight='bold')

    fig.tight_layout(rect=[0, 0, 1, 0.95])
    fig_path = os.path.join(out_dir, 'fig_lp_heatmap.png')
    save_fig(fig, fig_path, dpi=200)
    plt.close()
    print(f"Saved: {fig_path}")
    try:
        _r_fig_path = os.path.relpath(fig_path)
    except ValueError:
        _r_fig_path = fig_path
    print('Saved ' + _r_fig_path, file=sys.stderr, flush=True)
    return fig_path




def plot_ge_share(cc_df, thresholds, out_dir):
    YEARS = ['2020', '2021', '2022', '2023', '2024']
    YEAR_COLORS = {
        '2020': '#E74C3C', '2021': '#E67E22',
        '2022': '#27AE60', '2023': '#2980B9', '2024': '#8E44AD',
    }
    C_GE = '#2980B9'; C_LT = '#E74C3C'

    data = {yr: [] for yr in YEARS}
    for t in thresholds:
        for yr in YEARS:
            lp_ge = 'LP_asedx_p_' + yr + '_GE65'
            lp_lt = 'LP_asedx_p_' + yr + '_LT65'
            n_ge = int((cc_df[lp_ge].dropna() <= t).sum()) if lp_ge in cc_df.columns else 0
            n_lt = int((cc_df[lp_lt].dropna() <= t).sum()) if lp_lt in cc_df.columns else 0
            denom = n_ge + n_lt
            share = n_ge / denom if denom > 0 else 0.5
            data[yr].append(share * 100)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    for yr in YEARS:
        ax1.plot(thresholds, data[yr], 'o-', color=YEAR_COLORS[yr], lw=2, ms=6, label=yr)
    ax1.axhline(50, color='black', lw=0.8, ls='--', alpha=0.5)
    ax1.set_xlabel('LP threshold', fontsize=11)
    ax1.set_ylabel('GE65 share of age-significant metrics (%)', fontsize=11)
    ax1.set_title('GE65 share by year vs LP threshold\n(>50% = GE65 dominant)',
                  fontsize=11, fontweight='bold')
    ax1.set_xlim(-5, -30); ax1.invert_xaxis()
    ax1.set_ylim(0, 100); ax1.legend(fontsize=10); ax1.grid(True, alpha=0.25)

    ge_vals, lt_vals = [], []
    for yr in YEARS:
        lp_ge = 'LP_asedx_p_' + yr + '_GE65'
        lp_lt = 'LP_asedx_p_' + yr + '_LT65'
        n_ge = int((cc_df[lp_ge].dropna() <= -13).sum()) if lp_ge in cc_df.columns else 0
        n_lt = int((cc_df[lp_lt].dropna() <= -13).sum()) if lp_lt in cc_df.columns else 0
        ge_vals.append(n_ge); lt_vals.append(n_lt)

    x = list(range(len(YEARS)))
    ax2.bar(x, ge_vals, color=C_GE, alpha=0.85, label='GE65 significant')
    ax2.bar(x, lt_vals, bottom=ge_vals, color=C_LT, alpha=0.85, label='LT65 significant')
    for i, (g, l) in enumerate(zip(ge_vals, lt_vals)):
        if g > 0:
            ax2.text(i, g/2, str(g), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        if l > 0:
            ax2.text(i, g+l/2, str(l), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        pct_str = ('%d%%\nGE65' % (g*100//(g+l))) if (g+l) > 0 else ''
        ax2.text(i, g+l+3, pct_str, ha='center', fontsize=8.5, color=C_GE)
    ax2.set_xticks(x); ax2.set_xticklabels(YEARS, fontsize=11)
    ax2.set_ylabel('n metrics significant', fontsize=11)
    ax2.set_title('GE65 vs LT65 significant metrics by year\n(at LP threshold = -13)',
                  fontsize=11, fontweight='bold')
    ax2.legend(fontsize=10); ax2.grid(True, axis='y', alpha=0.25)

    fig.suptitle('Age group dominance of significant metrics by LP threshold',
                 fontsize=12, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    out = os.path.join(out_dir, 'lp_sensitivity_ge_share.png')
    save_fig(fig, out, dpi=200)
    plt.close()
    print('Saved: ' + out)
    print(('Saved ' + os.path.relpath(out)), file=sys.stderr, flush=True)


def plot_top_si(cc_df, thresholds, out_dir):
    YEARS_LIST = ['2020', '2021', '2022', '2023', '2024']
    TOP_SI = ['SSSSS', 'SSIII', 'SSSII', 'ISSSS', 'ISSII', 'SSSIS']
    SI_COLORS = {
        'SSSSS': '#1A5276', 'SSIII': '#E74C3C', 'SSSII': '#E67E22',
        'ISSSS': '#27AE60', 'ISSII': '#8E44AD', 'SSSIS': '#7D6608',
    }
    lp_cols_avail = ['LP_asedx_p_' + yr for yr in YEARS_LIST if
                     'LP_asedx_p_' + yr in cc_df.columns]

    si_counts = {si: [] for si in TOP_SI}
    si_counts['other'] = []
    for t in thresholds:
        si_list = []
        for m in cc_df.index:
            row = cc_df.loc[m]
            si = ''.join('S' if (not pd.isna(row.get(c)) and row.get(c) <= t)
                         else 'I' for c in lp_cols_avail)
            if si.count('S') > 1:
                si_list.append(si)
        counts = {si: si_list.count(si) for si in TOP_SI}
        total = len(si_list)
        other = total - sum(counts.values())
        for si in TOP_SI:
            si_counts[si].append(counts[si])
        si_counts['other'].append(other)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    for si in TOP_SI:
        ax1.plot(thresholds, si_counts[si], 'o-',
                 color=SI_COLORS[si], lw=2, ms=5, label=si)
    ax1.set_xlabel('LP threshold', fontsize=11)
    ax1.set_ylabel('n significant metrics (any year)', fontsize=11)
    ax1.set_title('Top SI pattern counts vs LP threshold\n(any year significant)',
                  fontsize=11, fontweight='bold')
    ax1.set_xlim(-5, -30); ax1.invert_xaxis()
    ax1.legend(fontsize=10, loc='upper right'); ax1.grid(True, alpha=0.25)

    t13_idx = thresholds.index(-13) if -13 in thresholds else len(thresholds)//2
    all_labels = TOP_SI + ['other']
    all_vals   = [si_counts[si][t13_idx] for si in TOP_SI] + [si_counts['other'][t13_idx]]
    all_colors = [SI_COLORS[si] for si in TOP_SI] + ['#BDC3C7']
    x = list(range(len(all_labels)))
    bars = ax2.bar(x, all_vals, color=all_colors, alpha=0.85, edgecolor='white')
    for bar, v in zip(bars, all_vals):
        if v > 0:
            ax2.text(bar.get_x()+bar.get_width()/2, v+0.5, str(v),
                     ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax2.set_xticks(x)
    ax2.set_xticklabels(all_labels, fontsize=10, fontfamily='monospace',
                        rotation=25, ha='right')
    ax2.set_ylabel('n metrics', fontsize=11)
    ax2.set_title('SI pattern distribution at LP = -13\n(any year significant)',
                  fontsize=11, fontweight='bold')
    ax2.grid(True, axis='y', alpha=0.25)

    fig.suptitle('Temporal significance pattern (SI) distribution vs LP threshold',
                 fontsize=12, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    out = os.path.join(out_dir, 'lp_sensitivity_top_si.png')
    save_fig(fig, out, dpi=200)
    plt.close()
    print('Saved: ' + out)
    print(('Saved ' + os.path.relpath(out)), file=sys.stderr, flush=True)


def plot_pandemic_waves(cc_df, thresholds, out_dir):
    YEARS = ['2020', '2021', '2022', '2023', '2024']
    YEAR_COLORS = {
        '2020': '#E74C3C', '2021': '#E67E22',
        '2022': '#27AE60', '2023': '#2980B9', '2024': '#8E44AD',
    }
    C_GE = '#2980B9'; C_LT = '#E74C3C'

    age_data = {}
    for yr in YEARS:
        lp_ge = 'LP_asedx_p_' + yr + '_GE65'
        lp_lt = 'LP_asedx_p_' + yr + '_LT65'
        n_ge = int((cc_df[lp_ge].dropna() <= -13).sum()) if lp_ge in cc_df.columns else 0
        n_lt = int((cc_df[lp_lt].dropna() <= -13).sum()) if lp_lt in cc_df.columns else 0
        age_data[yr] = {'GE65': n_ge, 'LT65': n_lt}

    x = list(range(len(YEARS)))
    ge = [age_data[yr]['GE65'] for yr in YEARS]
    lt = [age_data[yr]['LT65'] for yr in YEARS]
    tot = [ge[i]+lt[i] for i in range(len(YEARS))]
    ge_pct = [ge[i]/tot[i]*100 if tot[i] else 0 for i in range(len(YEARS))]
    lt_pct = [lt[i]/tot[i]*100 if tot[i] else 0 for i in range(len(YEARS))]

    ge_share_by_t = {yr: [] for yr in YEARS}
    for tt in thresholds:
        for yr in YEARS:
            lp_ge = 'LP_asedx_p_' + yr + '_GE65'
            lp_lt = 'LP_asedx_p_' + yr + '_LT65'
            n_ge = int((cc_df[lp_ge].dropna() <= tt).sum()) if lp_ge in cc_df.columns else 0
            n_lt = int((cc_df[lp_lt].dropna() <= tt).sum()) if lp_lt in cc_df.columns else 0
            denom = n_ge + n_lt
            ge_share_by_t[yr].append(n_ge/denom*100 if denom else 50)

    fig = plt.figure(figsize=(18, 7))
    gs  = fig.add_gridspec(1, 3, width_ratios=[1.2, 1, 1], wspace=0.3)
    ax1 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1])
    ax3 = fig.add_subplot(gs[2])

    w = 0.6
    ax1.bar(x, ge, w, color=C_GE, alpha=0.85, label='GE65 significant')
    ax1.bar(x, lt, w, bottom=ge, color=C_LT, alpha=0.85, label='LT65 significant')
    for i, (g, l) in enumerate(zip(ge, lt)):
        if g > 0:
            ax1.text(i, g/2, str(g), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        if l > 0:
            ax1.text(i, g+l/2, str(l), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        ax1.text(i, g+l+3, 'n=%d' % (g+l), ha='center', va='bottom', fontsize=8.5)
    ax1.set_xticks(x); ax1.set_xticklabels(YEARS, fontsize=11)
    ax1.set_ylabel('n metrics significant (LP<=-13)', fontsize=10)
    ax1.set_title('GE65 vs LT65 by year\n(at LP=-13)', fontsize=11, fontweight='bold')
    ax1.legend(fontsize=9); ax1.grid(True, axis='y', alpha=0.25)
    for xa, xb, label, col in [(-0.4, 0.4, 'COVID-19\n2020', 'grey'),
                                 (0.6,  1.4, 'Delta\n2021',    '#B03A2E'),
                                 (1.6,  2.4, 'Omicron\n2022', '#1A5276')]:
        ax1.axvspan(xa, xb, alpha=0.06, color=col)

    for yr in YEARS:
        ax2.plot(thresholds, ge_share_by_t[yr], 'o-',
                 color=YEAR_COLORS[yr], lw=2, ms=5, label=yr)
    ax2.axhline(50, color='black', lw=0.8, ls='--', alpha=0.5)
    ax2.set_xlabel('LP threshold', fontsize=10)
    ax2.set_ylabel('GE65 share (%)', fontsize=10)
    ax2.set_title('GE65 share vs LP threshold\nby year', fontsize=11, fontweight='bold')
    ax2.set_xlim(-5, -30); ax2.invert_xaxis()
    ax2.set_ylim(0, 100); ax2.legend(fontsize=9); ax2.grid(True, alpha=0.25)

    ax3.plot(x, ge_pct, 'o-', color=C_GE, lw=2.5, ms=9, label='GE65 %')
    ax3.plot(x, lt_pct, 's-', color=C_LT, lw=2.5, ms=9, label='LT65 %')
    ax3.axhline(50, color='black', lw=0.8, ls='--', alpha=0.4)
    for i in range(len(YEARS)):
        ax3.annotate('%d%%' % ge_pct[i], (i, ge_pct[i]),
                     textcoords='offset points', xytext=(0, 9),
                     ha='center', fontsize=9, color=C_GE, fontweight='bold')
        if lt_pct[i] > 0:
            ax3.annotate('%d%%' % lt_pct[i], (i, lt_pct[i]),
                         textcoords='offset points', xytext=(0, -14),
                         ha='center', fontsize=9, color=C_LT, fontweight='bold')
    ax3.set_xticks(x); ax3.set_xticklabels(YEARS, fontsize=11)
    ax3.set_ylabel('% of age-specific significant', fontsize=10)
    ax3.set_title('Age dominance shift\nacross pandemic waves (LP=-13)',
                  fontsize=11, fontweight='bold')
    ax3.set_ylim(0, 110); ax3.legend(fontsize=9); ax3.grid(True, alpha=0.25)

    fig.suptitle('Pandemic wave age dominance: LP threshold sensitivity',
                 fontsize=12, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    out = os.path.join(out_dir, 'lp_sensitivity_pandemic_waves.png')
    save_fig(fig, out, dpi=200)
    plt.close()
    print('Saved: ' + out)
    print(('Saved ' + os.path.relpath(out)), file=sys.stderr, flush=True)


def write_summary_excel_tab(results, xlsx_path, sc_names=None):
    """Add LP_sensitivity_summary tab to existing or new xlsx."""
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thresholds = sorted(results.keys())
    sc_name_map = sc_names or {}

    NAVY  = 'FF1F3864'; MID   = 'FF2E75B6'; WHITE = 'FFFFFFFF'
    GOLD  = 'FFFFD700'; GREY  = 'FFF2F2F2'
    DKRED = 'FF8B0000'; DKBLUE= 'FF1F4E79'
    SEC_COLORS = {
        'Min |CC|':       'FFECE9F1',
        'Overall':        'FFFDEDEC',
        'SI patterns':    'FFEAF4FB',
        'Super-cluster':  'FFE9F7EF',
        'Null SCs':       'FFFFF8E7',
        'Race groups':    'FFFDEDEC',
    }
    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color='888888')
    def brd(c, top=False, bot=False):
        c.border = Border(left=thin, right=thin,
                          top=(thick if top else thin),
                          bottom=(thick if bot else thin))
        return c

    def hdr_cell(ws, r, c, val, bg=NAVY, fg='FFFFFF', bold=True, size=10):
        cl = ws.cell(r, c, val)
        cl.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cl.fill = PatternFill('solid', fgColor=bg)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        return brd(cl)

    def dat_cell(ws, r, c, val, bg=WHITE, bold=False, left=False,
                 fg='000000', size=9, compare=False):
        cl = ws.cell(r, c, val)
        cl.font = Font(name='Arial', bold=bold or compare, color=fg, size=size)
        cl.fill = PatternFill('solid', fgColor=bg if not compare else GOLD)
        cl.alignment = Alignment(horizontal='left' if left else 'center',
                                  vertical='center')
        return brd(cl)

    def sec_hdr(ws, ri, label, bg, ncols):
        ws.row_dimensions[ri].height = 14
        for ci in range(1, ncols+1):
            cl = ws.cell(ri, ci, label if ci == 1 else '')
            cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            cl.fill = PatternFill('solid', fgColor=bg)
            cl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            brd(cl, top=True)

    # Load or create workbook
    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)

    TAB = 'LP_sensitivity_summary'
    if TAB in wb.sheetnames:
        wb.remove(wb[TAB])
    ws = wb.create_sheet(TAB)

    ncols = 1 + len(thresholds)

    # Column widths
    ws.column_dimensions['A'].width = 36
    for ci in range(2, ncols+1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    compare_set = {-11, -13, -15}

    ri = 1
    # Title row
    ws.row_dimensions[ri].height = 20
    for ci in range(1, ncols+1):
        hdr_cell(ws, ri, ci,
                 'LP Threshold Sensitivity Summary' if ci == 1 else '',
                 bg=NAVY, size=12)
    ri += 1

    # Column headers (LP threshold values)
    ws.row_dimensions[ri].height = 18
    hdr_cell(ws, ri, 1, 'Metric', bg='FF37474F', size=9)
    for i, t in enumerate(thresholds):
        is_c = t in compare_set
        hdr_cell(ws, ri, 2+i, str(t),
                 bg='FF1F3864' if not is_c else 'FFB7950B', size=10)
    ri += 1
    ws.freeze_panes = 'B3'

    def data_row(label, vals, section_bg, bold_row=False):
        nonlocal ri
        ws.row_dimensions[ri].height = 14
        dat_cell(ws, ri, 1, label, bg=section_bg, left=True, bold=bold_row, size=9)
        for i, (t, v) in enumerate(zip(thresholds, vals)):
            is_c = t in compare_set
            dat_cell(ws, ri, 2+i, v, bg=section_bg, bold=bold_row,
                     compare=is_c, size=9)
        ri += 1

    def spacer():
        nonlocal ri
        ws.row_dimensions[ri].height = 4
        for ci in range(1, ncols+1):
            ws.cell(ri, ci).fill = PatternFill('solid', fgColor='FFF0F0F0')
        ri += 1

    # --- Min CC ---
    sec_hdr(ws, ri, '  Min |CC| required at each LP threshold', 'FF6A0572', ncols); ri += 1
    bg = SEC_COLORS['Min |CC|']
    data_row('Min |CC| for significance',
             ['{:.4f}'.format(results[t]['min_cc']) for t in thresholds], bg)
    spacer()

    # --- Overall counts ---
    sec_hdr(ws, ri, '  Overall significant metrics (any year)', 'FF8B0000', ncols); ri += 1
    bg = SEC_COLORS['Overall']
    data_row('n significant (any year)',
             [results[t]['n_multiyear'] for t in thresholds], bg, bold_row=True)
    data_row('n risky (+CC)',
             [results[t]['n_risky'] for t in thresholds], bg)
    data_row('n protective (-CC)',
             [results[t]['n_protective'] for t in thresholds], bg)
    data_row('% risky',
             ['{:.1f}%'.format(results[t]['pct_risky']) for t in thresholds], bg)
    data_row('Mean CC risky',
             ['{:+.3f}'.format(results[t]['mean_cc_risky']) if results[t]['mean_cc_risky'] else '---'
              for t in thresholds], bg)
    data_row('Mean CC protective',
             ['{:+.3f}'.format(results[t]['mean_cc_prot']) if results[t]['mean_cc_prot'] else '---'
              for t in thresholds], bg)
    spacer()

    # --- SI patterns ---
    sec_hdr(ws, ri, '  SI temporal patterns', 'FF1F4E79', ncols); ri += 1
    bg = SEC_COLORS['SI patterns']
    for si, desc in [('SSSSS','all 5 years'), ('SSIII','2020-21 only'),
                     ('SSSII','2020-22'), ('ISSSS','2021-24 lagged'),
                     ('ISSII','2021-22')]:
        data_row('%s  (%s)' % (si, desc),
                 [results[t].get(si, 0) for t in thresholds], bg)
    spacer()

    # --- Super-cluster breakdown ---
    sec_hdr(ws, ri, '  Super-cluster breakdown', 'FF1E8449', ncols); ri += 1
    bg = SEC_COLORS['Super-cluster']
    all_scs = sorted(set().union(*[r['sc_sig'].keys() for r in results.values()]))
    for sc in all_scs:
        label = 'SC%02d  %s' % (sc, sc_name_map.get(sc, ''))
        data_row(label,
                 [results[t]['sc_sig'].get(sc, '---') for t in thresholds], bg)
    spacer()

    # --- Null SCs ---
    sec_hdr(ws, ri, '  Null SCs (0 significant metrics)', 'FF7D6608', ncols); ri += 1
    bg = SEC_COLORS['Null SCs']
    data_row('SC IDs with 0 significant',
             [str(results[t]['sc_null']) for t in thresholds], bg)
    spacer()

    # --- Race groups ---
    sec_hdr(ws, ri, '  Race / ethnicity groups', 'FF6C3483', ncols); ri += 1
    bg = SEC_COLORS['Race groups']
    for label, key in [
        ('n AIAN significant',      'n_aian'),
        ('  of which risky',        'n_aian_risky'),
        ('n White non-Hisp sig',    'n_wnh'),
        ('  of which protective',   'n_wnh_prot'),
        ('n Hispanic sig',          'n_hisp'),
        ('  of which risky',        'n_hisp_risky'),
    ]:
        data_row(label, [results[t].get(key, '---') for t in thresholds], bg)

    wb.save(xlsx_path)
    try:
        _r_xlsx_path = os.path.relpath(xlsx_path)
    except ValueError:
        _r_xlsx_path = xlsx_path
    print('Saved ' + _r_xlsx_path, file=sys.stderr, flush=True)
    print('Saved: ' + xlsx_path + ' (tab: ' + TAB + ')')
    print(('Saved ' + os.path.relpath(xlsx_path) + '  (tab: ' + TAB + ')'), file=sys.stderr, flush=True)


def plot_pandemic_waves_grid(cc_df, thresholds, out_dir):
    """6-panel grid: one panel per LP threshold, GE65% vs LT65% by year with wave shading."""
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np

    YEARS = ['2020', '2021', '2022', '2023', '2024']
    C_GE  = '#2980B9'
    C_LT  = '#C0392B'

    WAVE_BANDS = [
        (-0.5, 0.5,  'COVID-19', 'grey',    '#F2F2F2'),
        ( 0.5, 1.5,  'Delta',    '#C0392B',  '#FDEDEC'),
        ( 1.5, 2.5,  'Omicron',  '#2980B9',  '#EBF5FB'),
    ]

    # Use 6 representative thresholds spread across the sweep range
    plot_thresholds = sorted([t for t in thresholds
                              if t in [-7,-10,-13,-16,-19,-22]],
                             reverse=False)
    if not plot_thresholds:
        # fallback: pick 6 evenly spaced
        plot_thresholds = sorted(thresholds)[::max(1, len(thresholds)//6)][:6]

    ncols = 3
    nrows = (len(plot_thresholds) + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(18, 5.5 * nrows),
                             sharex=True, sharey=True)
    axes = np.array(axes).reshape(-1)

    for ax_i, t in enumerate(plot_thresholds):
        ax = axes[ax_i]

        ge_pct_list = []
        lt_pct_list = []
        n_total = 0

        for yr in YEARS:
            lp_ge = 'LP_asedx_p_' + yr + '_GE65'
            lp_lt = 'LP_asedx_p_' + yr + '_LT65'
            n_ge = int((cc_df[lp_ge].dropna() <= t).sum()) if lp_ge in cc_df.columns else 0
            n_lt = int((cc_df[lp_lt].dropna() <= t).sum()) if lp_lt in cc_df.columns else 0
            denom = n_ge + n_lt
            ge_pct_list.append(n_ge / denom * 100 if denom else 0)
            lt_pct_list.append(n_lt / denom * 100 if denom else 0)
            n_total += denom

        x = list(range(len(YEARS)))

        # Wave background shading
        for xa, xb, label, col, bg in WAVE_BANDS:
            ax.axvspan(xa, xb, alpha=0.18, color=bg, zorder=0)
            ax.text((xa + xb) / 2, 107, label, ha='center', va='bottom',
                    fontsize=8, color=col,
                    fontweight='bold' if label == 'Delta' else 'normal')

        # Lines
        ax.plot(x, ge_pct_list, 'o-', color=C_GE, lw=2.5, ms=8,
                label='GE65 dominant', zorder=3)
        ax.plot(x, lt_pct_list, 's-', color=C_LT, lw=2.5, ms=8,
                label='LT65 dominant', zorder=3)

        # Annotate points
        for i, (gp, lp_v) in enumerate(zip(ge_pct_list, lt_pct_list)):
            ax.annotate('%d' % round(gp),  (i, gp),
                        xytext=(0,  6), textcoords='offset points',
                        ha='center', fontsize=9, color=C_GE, fontweight='bold')
            ax.annotate('%d' % round(lp_v), (i, lp_v),
                        xytext=(0, -13), textcoords='offset points',
                        ha='center', fontsize=9, color=C_LT)

        n_sig_all = 0
        for yr in YEARS:
            lp_all = 'LP_asedx_p_' + yr
            if lp_all in cc_df.columns:
                n_sig_all += int((cc_df[lp_all].dropna() <= t).sum())

        ax.set_title('LP <= %d   (total sig = %d)' % (t, n_sig_all),
                     fontsize=11, fontweight='normal', color='#333333')
        ax.set_xticks(x)
        ax.set_xticklabels(YEARS, fontsize=10)
        ax.set_ylim(-5, 115)
        ax.set_xlim(-0.5, 4.5)
        ax.set_ylabel('% of age-specific sig metrics', fontsize=9)
        ax.grid(True, axis='y', alpha=0.2)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        if ax_i == 0:
            ax.legend(fontsize=9, loc='lower left')

    # Hide unused axes
    for ax_i in range(len(plot_thresholds), len(axes)):
        axes[ax_i].set_visible(False)

    fig.suptitle('LP threshold sensitivity: GE65 vs LT65 dominance shift across pandemic waves',
                 fontsize=13, fontweight='bold', y=1.01)
    fig.tight_layout(rect=[0, 0, 1, 1])

    out = os.path.join(out_dir, 'lp_sensitivity_pandemic_waves_grid.png')
    save_fig(fig, out, dpi=150)
    print(('Saved ' + os.path.relpath(out)), file=sys.stderr, flush=True)


def main():
    parser = argparse.ArgumentParser(description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('--cc-file',     required=True)
    parser.add_argument('--master-xlsx', default=None)
    parser.add_argument('--out-dir',     default='.')
    parser.add_argument('--compare',     default='-11,-13,-15')
    parser.add_argument('--n-counties',  type=int, default=3032)
    parser.add_argument('--sc-names',    default=None,
                        help='sem_sc_names.csv')
    # Extract --compare manually before argparse (argparse mis-parses negative numbers)
    import sys as _sys
    _argv = list(_sys.argv[1:])
    _compare_val = '-11,-13,-15'
    _clean_argv = []
    _skip = False
    for _idx, _tok in enumerate(_argv):
        if _skip:
            _skip = False
            continue
        if _tok == '--compare' and _idx + 1 < len(_argv):
            _compare_val = _argv[_idx + 1].strip('"\'')
            _skip = True
        else:
            _clean_argv.append(_tok)
    args = parser.parse_args(_clean_argv)
    args.compare = _compare_val

    os.makedirs(args.out_dir, exist_ok=True)
    compare_thresholds = [int(x) for x in args.compare.split(',')]

    # ---- Load CC file ----
    print(f"Loading CC file: {args.cc_file}")
    cc_df = pd.read_csv(args.cc_file).set_index('metric')
    n_metrics = len(cc_df)
    print(f"  {n_metrics} metrics  {cc_df.shape[1]} columns")

    # Check which COLS_15 are present
    avail_15 = [c for c in COLS_15 if c in cc_df.columns]
    missing  = [c for c in COLS_15 if c not in cc_df.columns]
    if missing:
        print(f"  WARNING: {len(missing)} LP columns not found: {missing[:5]}")
    print(f"  {len(avail_15)} of 15 pandemic LP columns available")

    # ---- Min CC table ----
    print("\nMin |CC| required at each LP threshold (n=3032 counties):")
    print(f"  {'LP':>6}  {'min |CC|':>10}  {'Bonferroni context'}")
    n_tests = n_metrics * 15
    bonf_lp = np.log10(0.05 / n_tests)
    print(f"  {'Bonf':>6}  {min_cc_for_lp(bonf_lp):>10.4f}  "
          f"(alpha=0.05 / {n_tests:,} tests = p={0.05/n_tests:.2e}, LP={bonf_lp:.1f})")
    for t in THRESHOLDS:
        r = min_cc_for_lp(t, args.n_counties)
        print(f"  {t:>6}  {r:>10.4f}")

    # ---- Count table ----
    print("\nBuilding count table across thresholds and death measures...")
    data = {}
    for col, lbl in zip(avail_15, SHORT_15[:len(avail_15)]):
        lp_vals = cc_df[col].dropna()
        counts  = [int((lp_vals <= t).sum()) for t in THRESHOLDS]
        data[lbl] = counts

    df_counts = pd.DataFrame(data, index=THRESHOLDS)
    df_counts.index.name = 'LP_threshold'

    counts_path = os.path.join(args.out_dir, 'lp_threshold_counts.csv')
    df_counts.to_csv(counts_path)
    print(f"Saved: {counts_path}")
    try:
        _r_counts_path = os.path.relpath(counts_path)
    except ValueError:
        _r_counts_path = counts_path
    print('Saved ' + _r_counts_path, file=sys.stderr, flush=True)

    # ---- Excel table tab ----
    write_excel_tab(df_counts, args.out_dir, compare_thresholds,
                    n_metrics, bonf_lp)

    # ---- Heatmap figure ----
    plot_count_heatmap(df_counts, args.out_dir, compare_thresholds)

    # ---- GE share / top SI / pandemic waves figures ----
    plot_ge_share(cc_df, THRESHOLDS, args.out_dir)
    plot_top_si(cc_df, THRESHOLDS, args.out_dir)
    plot_pandemic_waves(cc_df, THRESHOLDS, args.out_dir)
    plot_pandemic_waves_grid(cc_df, THRESHOLDS, args.out_dir)

    # ---- GE share / top SI / pandemic waves figures ----

    # Print key summary
    print("\nKey thresholds (All ages only):")
    print(f"{'LP':>6}  " + "  ".join(f"{yr:>8}" for yr in YEARS))
    for t in [-7, -9, -10, -11, -13, -14, -15, -20]:
        if t in THRESHOLDS:
            vals = [int(df_counts.loc[t, f'{yr} All']) for yr in YEARS]
            print(f"{t:>6}  " + "  ".join(f"{v:>8}" for v in vals))

    # ---- Figure ----
    print("\nPlotting sensitivity figure...")
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    # Left: all 15 lines
    for col, lbl in zip(avail_15, SHORT_15[:len(avail_15)]):
        yr  = lbl[:4]
        ag  = lbl[5:]
        ax1.plot(THRESHOLDS, df_counts[lbl].values,
                 color=YEAR_COLORS.get(yr, '#888'),
                 linestyle=LS_MAP.get(ag, '-'),
                 lw=2.2 if ag == 'All' else 1.2,
                 alpha=0.9 if ag == 'All' else 0.60,
                 label=lbl if ag == 'All' else None)

    for t in compare_thresholds:
        ax1.axvline(t, color='black', lw=1.2, ls='--', alpha=0.5)
        ax1.text(t, ax1.get_ylim()[1]*0.02 if ax1.get_ylim()[1] > 0 else 10,
                 f'LP={t}', fontsize=7.5, ha='center', va='bottom')

    ax1.set_xlabel('LP threshold', fontsize=11)
    ax1.set_ylabel('n metrics significant', fontsize=11)
    ax1.set_title('All 15 death measures\n(solid=All ages, dashed=GE65, dotted=LT65)',
                  fontsize=11, fontweight='bold')
    ax1.set_xlim(-5, -30); ax1.invert_xaxis()
    ax1.grid(True, alpha=0.25)

    legend_elems = []
    for yr, col in YEAR_COLORS.items():
        legend_elems.append(mlines.Line2D([0],[0], color=col, lw=2, label=yr))
    legend_elems.append(mlines.Line2D([0],[0], color='black', lw=2,   ls='-',  label='All ages'))
    legend_elems.append(mlines.Line2D([0],[0], color='black', lw=1.5, ls='--', label='GE65'))
    legend_elems.append(mlines.Line2D([0],[0], color='black', lw=1.5, ls=':',  label='LT65'))
    ax1.legend(handles=legend_elems, fontsize=8.5, loc='upper right', framealpha=0.9)

    # Right: All ages only with annotations at compare thresholds
    for yr in YEARS:
        col  = f'{yr} All'
        vals = df_counts[col].values
        ax2.plot(THRESHOLDS, vals,
                 color=YEAR_COLORS[yr], lw=2.5, marker='o', ms=5, label=yr)
        for t in compare_thresholds:
            if t in THRESHOLDS:
                idx = THRESHOLDS.index(t)
                ax2.annotate(str(vals[idx]),
                             (t, vals[idx]),
                             textcoords='offset points', xytext=(6, 2),
                             fontsize=8, color=YEAR_COLORS[yr], fontweight='bold')

    for t in compare_thresholds:
        ax2.axvline(t, color='black', lw=1.2, ls='--', alpha=0.5)

    ax2.set_xlabel('LP threshold', fontsize=11)
    ax2.set_ylabel('n metrics significant', fontsize=11)
    ax2.set_title('All ages only\n(annotated at comparison thresholds)',
                  fontsize=11, fontweight='bold')
    ax2.set_xlim(-5, -30); ax2.invert_xaxis()
    ax2.legend(fontsize=10, framealpha=0.9)
    ax2.grid(True, alpha=0.25)

    fig.suptitle(
        f'LP threshold sensitivity  |  n={n_metrics} metrics  |  n=3,032 counties  |  '
        f'Bonferroni LP={bonf_lp:.1f} for {n_tests:,} tests',
        fontsize=11, fontweight='bold')
    plt.tight_layout()

    fig_path = os.path.join(args.out_dir, 'fig_lp_threshold_sensitivity.png')
    save_fig(fig, fig_path, dpi=200)
    plt.close()
    print(f"Saved: {fig_path}")
    try:
        _r_fig_path = os.path.relpath(fig_path)
    except ValueError:
        _r_fig_path = fig_path
    print('Saved ' + _r_fig_path, file=sys.stderr, flush=True)

    # ---- Detailed comparison at compare_thresholds (needs master) ----
    if args.master_xlsx:
        print(f"\nLoading master Excel for detailed comparison: {args.master_xlsx}")
        master_df = load_master_multiyear(args.master_xlsx)
        print(f"  {len(master_df)} significant metrics any year (at LP=-13)")

        # Load SC names from sem_sc_names.csv
        SC_NAMES = {}
        if args.sc_names and os.path.exists(args.sc_names):
            _sc_df = pd.read_csv(args.sc_names)
            for _, _r in _sc_df.iterrows():
                SC_NAMES[int(_r['sc_id'])] = str(_r.get('sc_name', '')).strip()
            print(f'  Loaded {len(SC_NAMES)} SC names from {args.sc_names}',
                  file=sys.stderr)

        # Build per-threshold stats dict for summary
        print("  Computing stats at comparison thresholds...")
        compare_results = {}
        for t in compare_thresholds:
            compare_results[t] = get_stats_at_threshold(cc_df, master_df, t)

        txt_path = os.path.join(args.out_dir, 'lp_sensitivity_summary.txt')
        write_summary(compare_results, txt_path, SC_NAMES)
        xlsx_summary = os.path.join(args.out_dir, 'lp_sensitivity_counts.xlsx')
        write_summary_excel_tab(compare_results, xlsx_summary, SC_NAMES)
    else:
        print("\nSkipping detailed SC/race comparison (no --master-xlsx provided)")

    print("\nDone.")


if __name__ == '__main__':
    main()
