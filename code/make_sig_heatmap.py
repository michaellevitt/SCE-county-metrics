#!/usr/bin/env python3
"""
make_sig_heatmap.py

Produce a significance-colored copy of sc_label_heatmap.png.
Cluster labels and SC labels are colored red if the cluster/SC contains
at least one metric with LP <= lp_thresh in any all-ages death column
(i.e. n_sig_years >= 1). Age-stratified columns (GE65/LT65) are NOT
counted -- they would inflate the red count by ~7 clusters whose strongest
signal lives only in an age split.

Usage:
  python3 code/make_sig_heatmap.py \
      --sem-clusters ward_sem_2745/ward_sem_metrics.csv \
      --sc-assignments ward_xde_2745/sem_sc_assignments.csv \
      --sc-names ward_xde_2745/sem_sc_names.csv \
      --sem100-labels ward_sem_2745/sem100_labels.csv \
      --master-xlsx master_sem_clusters_2745.xlsx \
      --embeddings-cache ward_xde_2745/sem_sc_label_embeddings.npy \
      --lp-thresh -13.0 \
      --out ward_xde_2745/sc_label_heatmap_sig.png
"""
import argparse, os, sys
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import squareform
from openpyxl import load_workbook

from _workbook_style import detect_table_offset


def _rel(p):
    try:
        return os.path.relpath(p)
    except ValueError:
        return p

def _tee(path):
    msg = 'Saved ' + _rel(path)
    print(msg); print(msg, file=sys.stderr, flush=True)


def main():
    p = argparse.ArgumentParser(description='Significance-colored SC label heatmap')
    p.add_argument('--sem-clusters',     required=True,  help='ward_sem_metrics.csv')
    p.add_argument('--sc-assignments',   required=True,  help='sem_sc_assignments.csv')
    p.add_argument('--sc-names',         required=True,  help='sem_sc_names.csv')
    p.add_argument('--sem100-labels',    required=True,  help='sem100_labels.csv')
    p.add_argument('--master-xlsx',      required=True,  help='master_sem_clusters_2745.xlsx')
    p.add_argument('--embeddings-cache', required=True,  help='sem_sc_label_embeddings.npy')
    p.add_argument('--lp-thresh',        type=float, default=-13.0)
    p.add_argument('--cc-sig',           type=float, default=0.3,
                   help='|CC| mode (default): Significant |CC|>cc-sig. Set None via --lp-label for legacy LP labels.')
    p.add_argument('--cc-verysig',       type=float, default=0.45,
                   help='|CC| Very Significant threshold (dark-red tier; needs n_verysig_years in master)')
    p.add_argument('--lp-label', action='store_true', default=False,
                   help='Legacy: label by LP threshold instead of |CC| bands')
    p.add_argument('--min-sig-metrics',  type=int,   default=1,
                   help='Min number of sig metrics required to flag cluster (default 1)')
    p.add_argument('--min-sig-years',    type=int,   default=1,
                   help='A metric counts as "sig" only if n_sig_years >= this (default 1)')
    p.add_argument('--out',              required=True,  help='output png path (flat heatmap)')
    p.add_argument('--out-dendro',       default=None,
                   help='Optional second output PNG: dendrogram + heatmap '
                        '(ordered by Ward linkage on cosine distance).')
    args = p.parse_args()

    # ---- Load SEM cluster medoid explain texts ----
    sem_df = pd.read_csv(args.sem_clusters)
    cl_col = next((c for c in ['cluster_id','semantic_cluster_id','semantic_cluster']
                   if c in sem_df.columns), None)
    if cl_col is None:
        cl_col = next(c for c in sem_df.columns if 'cluster' in c.lower())
    if 'is_medoid' in sem_df.columns:
        reps = sem_df[sem_df['is_medoid'] == True].copy()
    else:
        reps = sem_df.groupby(cl_col).first().reset_index()
    reps        = reps.sort_values(cl_col).reset_index(drop=True)
    cluster_ids = reps[cl_col].tolist()
    n           = len(cluster_ids)
    print(f'  {n} SEM clusters loaded')

    # ---- Load cached embeddings ----
    emb = np.load(args.embeddings_cache)
    # L2-normalise
    norms = np.linalg.norm(emb, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    emb_n = emb / norms
    # Cosine similarity matrix
    sim_mat = emb_n @ emb_n.T
    print(f'  Embeddings loaded: {emb.shape}')

    # ---- SC assignments ----
    assign_raw = pd.read_csv(args.sc_assignments)
    id_col = 'sem100' if 'sem100' in assign_raw.columns else assign_raw.columns[0]
    assign_map = dict(zip(assign_raw[id_col].astype(int), assign_raw['sc_id'].astype(int)))

    # ---- SC names ----
    names_df  = pd.read_csv(args.sc_names)
    sc_name_map = dict(zip(names_df['sc_id'].astype(int), names_df['sc_name'].astype(str)))

    # ---- Short LLM labels ----
    sl_df   = pd.read_csv(args.sem100_labels)
    _id_col = next((c for c in ['sem100','cluster_id','Ward100'] if c in sl_df.columns), sl_df.columns[0])
    _lb_col = next((c for c in ['label_proposed','label','cluster_label','short_label'] if c in sl_df.columns), sl_df.columns[-1])
    short_label_map = dict(zip(sl_df[_id_col].astype(int), sl_df[_lb_col].astype(str)))
    print(f'  Short labels loaded: {len(short_label_map)} (col={_lb_col!r})')

    # ---- Build assign_df ----
    assign_df = pd.DataFrame({
        'sem100':   cluster_ids,
        'sc_id':    [assign_map.get(int(c), 0) for c in cluster_ids],
        'cluster_label': [short_label_map.get(int(c), str(c)) for c in cluster_ids],
    })

    # ---- Load significance from master xlsx ----
    # A metric is "sig" iff n_sig_years >= --min-sig-years (all-ages columns
    # only; age-stratified GE65/LT65 are excluded since those are not part of
    # the n_sig_years count).
    # A cluster/SC is "significant" iff it contains >= --min-sig-metrics such
    # metrics. Default (1, 1) reproduces the lenient "any cluster with any
    # all-ages-sig metric" behaviour.
    K_min = int(args.min_sig_metrics)
    Y_min = int(args.min_sig_years)
    cluster_counts = {}   # ward100 -> count of sig metrics
    sc_counts      = {}   # sc_id   -> count of sig metrics
    wb  = load_workbook(args.master_xlsx, read_only=True, data_only=True)
    ws  = wb['Master']
    hdr_row, hdr_col = detect_table_offset(ws)
    hdr = [ws.cell(hdr_row, c).value for c in range(hdr_col, ws.max_column + 1)]
    # Column index (1-based) within the worksheet for each named header
    def col_for(name):
        return hdr.index(name) + hdr_col if name in hdr else None
    w_idx    = col_for('Ward100')
    sc_idx   = col_for('super_cluster_id')
    nsig_idx = col_for('n_sig_years')
    nvery_idx = col_for('n_verysig_years')   # present only in |CC| (cc-mode) master
    vcluster_counts = {}   # ward100 -> count of very-sig metrics
    vsc_counts      = {}   # sc_id   -> count of very-sig metrics
    print(f'  Ward100={w_idx}  SC={sc_idx}  n_sig_years={nsig_idx}  n_verysig_years={nvery_idx}')
    if w_idx and nsig_idx:
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            nsig_val = row[nsig_idx-1]
            nvery_val = row[nvery_idx-1] if nvery_idx else None
            w = int(row[w_idx-1]) if row[w_idx-1] is not None else None
            s = int(row[sc_idx-1]) if (sc_idx and row[sc_idx-1] is not None) else None
            if nsig_val is not None and int(nsig_val) >= Y_min:
                if w is not None: cluster_counts[w] = cluster_counts.get(w, 0) + 1
                if s is not None: sc_counts[s] = sc_counts.get(s, 0) + 1
            if nvery_val is not None and int(nvery_val) >= Y_min:
                if w is not None: vcluster_counts[w] = vcluster_counts.get(w, 0) + 1
                if s is not None: vsc_counts[s] = vsc_counts.get(s, 0) + 1
    wb.close()
    sig_clusters = {w for w, c in cluster_counts.items() if c >= K_min}
    sig_scs      = {s for s, c in sc_counts.items()      if c >= K_min}
    verysig_clusters = {w for w, c in vcluster_counts.items() if c >= K_min}
    verysig_scs      = {s for s, c in vsc_counts.items()      if c >= K_min}
    print(f'  Sig clusters: {len(sig_clusters)}  Sig SCs: {len(sig_scs)}  '
          f'(>= {K_min} metric(s) with n_sig_years >= {Y_min})')

    # ---- Sort by sc_id then sem100 ----
    sorted_df = assign_df.sort_values(['sc_id', 'sem100']).reset_index(drop=True)
    idx_map   = {v: i for i, v in enumerate(assign_df['sem100'].tolist())}
    row_order = sorted_df['sem100'].map(idx_map).tolist()
    sim_ord   = sim_mat[np.ix_(row_order, row_order)]

    sc_ord    = sorted_df['sc_id'].tolist()
    boundaries = [i for i in range(1, n) if sc_ord[i] != sc_ord[i-1]]
    bdry_ext   = [0] + boundaries + [n]
    sc_mids    = [(bdry_ext[i] + bdry_ext[i+1]) / 2.0 for i in range(len(bdry_ext)-1)]
    sc_ids_unique = [sc_ord[bdry_ext[i]] for i in range(len(bdry_ext)-1)]

    clust_labels_full = [
        short_label_map.get(int(r['sem100']), str(r['cluster_label']))
        for _, r in sorted_df.iterrows()
    ]
    sc_mid_labels_full = [f'SC{sc:02d}: {sc_name_map.get(sc, "")}'
                          for sc in sc_ids_unique]

    # ---- Colors ----
    SIG_RED      = '#CC0000'   # Significant tier
    VERYSIG_DARK = '#7A0000'   # Very Significant tier (cc mode)
    BASE_COL = 'black'
    clust_colors = [VERYSIG_DARK if int(r['sem100']) in verysig_clusters
                    else (SIG_RED if int(r['sem100']) in sig_clusters else BASE_COL)
                    for _, r in sorted_df.iterrows()]
    sc_colors    = [VERYSIG_DARK if sc in verysig_scs
                    else (SIG_RED if sc in sig_scs else BASE_COL)
                    for sc in sc_ids_unique]

    # ---- Draw figure ----
    fig2 = plt.figure(figsize=(24, 20))
    ax2  = fig2.add_axes([0.30, 0.03, 0.52, 0.90])
    ax2.imshow(sim_ord, aspect='equal', cmap='RdBu_r',
               vmin=-0.2, vmax=1.0, interpolation='none')
    for b in boundaries:
        ax2.axhline(b - 0.5, color='black', lw=2.0)
        ax2.axvline(b - 0.5, color='black', lw=2.0)
    ax2.set_yticks(range(n))
    ax2.set_yticklabels(clust_labels_full, fontsize=8.75, fontweight='bold')
    ax2.tick_params(axis='y', length=2, pad=2)
    ax2.set_xticks([])
    if not args.lp_label:
        _sig_txt = (f'red = >= {args.min_sig_metrics} variable(s) with |CC|>{args.cc_sig:g} '
                    f'in >= {args.min_sig_years} all-ages year(s); '
                    f'dark red = same with |CC|>{args.cc_verysig:g} (Strong)')
    else:
        _sig_txt = (f'red = >= {args.min_sig_metrics} variable(s) sig in >= {args.min_sig_years} '
                    f'all-ages year(s) (LP<={args.lp_thresh})')
    ax2.set_title(
        f'Cluster label cosine similarity ({len(sorted_df)}x{len(sorted_df)})\n'
        f'ordered by super-cluster assignment  |  ' + _sig_txt,
        fontsize=11, fontweight='bold', pad=8)
    fig2.canvas.draw()
    # Apply colors AFTER canvas.draw() so tick labels are fully initialized
    for tick, col in zip(ax2.get_yticklabels(), clust_colors):
        tick.set_color(col)
    fig2.canvas.draw()  # second draw to lock colors before savefig
    trans = ax2.transData + fig2.transFigure.inverted()
    for mid, lbl, col in zip(sc_mids, sc_mid_labels_full, sc_colors):
        _, fy = trans.transform((0, mid))
        fig2.text(0.83, fy, lbl, va='center', ha='left',
                  fontsize=8.5, fontweight='bold', color=col,
                  transform=fig2.transFigure)
    fig2.savefig(args.out, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    _tee(args.out)

    # ---- Optional: dendrogram + heatmap combined view ----
    if args.out_dendro:
        _render_dendro_heatmap(
            sim_mat, cluster_ids, sig_clusters, sig_scs,
            assign_map, sc_name_map, short_label_map,
            args, K_min, Y_min)


def _render_dendro_heatmap(sim_mat, cluster_ids, sig_clusters, sig_scs,
                           assign_map, sc_name_map, short_label_map,
                           args, K_min, Y_min):
    """Combined dendrogram + heatmap, ordered by Ward linkage on cosine distance.

    Parallels the look of combined_xde_centroid.png:
      - left panel: horizontal dendrogram (leaves on right)
      - middle: heatmap, rows/cols ordered by dendrogram leaves
      - leaf labels colored red if cluster is significant
      - SC names listed on the far right (in dendrogram order, with breaks)
    """
    import matplotlib.gridspec as gridspec

    n        = len(cluster_ids)
    SIG_RED  = '#CC0000'
    BASE_COL = 'black'

    # ---- Linkage on cosine distance ----
    dist = 1.0 - sim_mat
    np.fill_diagonal(dist, 0.0)
    dist = np.clip(dist, 0.0, None)
    dist = (dist + dist.T) / 2.0       # enforce symmetry
    cond = squareform(dist, checks=False)
    Z    = linkage(cond, method='ward')   # ward gives balanced, contiguous SC groups

    # ---- Get leaf order without plotting ----
    dn_ref     = dendrogram(Z, no_plot=True)
    leaf_order = dn_ref['leaves']
    leaf_cl    = [cluster_ids[i] for i in leaf_order]
    leaf_sc    = [assign_map.get(int(c), 0) for c in leaf_cl]
    leaf_lbl   = [short_label_map.get(int(c), str(c)) for c in leaf_cl]
    leaf_color = [SIG_RED if int(c) in sig_clusters else BASE_COL for c in leaf_cl]

    # Reorder similarity matrix
    sim_d = sim_mat[np.ix_(leaf_order, leaf_order)]

    # ---- SC group runs in dendrogram order (may be non-contiguous) ----
    runs = []  # list of (start, end_exclusive, sc_id)
    s = 0
    for i in range(1, n):
        if leaf_sc[i] != leaf_sc[i-1]:
            runs.append((s, i, leaf_sc[i-1]))
            s = i
    runs.append((s, n, leaf_sc[-1]))
    boundaries = [r[0] for r in runs[1:]]   # row index where each new run begins

    # ---- Figure layout ----
    # 4-column gridspec, left -> right:
    #   [cluster labels | dendrogram | square heatmap | SC labels]
    # We size the figure so the heatmap's gridspec cell is naturally square
    # (cell_w == cell_h):
    #   cell_w = ratio_h * (right - left) * fig_w
    #   cell_h = (top  - bottom) * fig_h
    # With ratios [0.18, 0.18, 0.50, 0.14], margins (0.02, 0.98, 0.94, 0.05),
    # cell_w == cell_h gives fig_h ≈ fig_w * 0.539.
    fig_w = 28.0
    fig_h = 15.5
    fig = plt.figure(figsize=(fig_w, fig_h))
    gs   = gridspec.GridSpec(
        1, 4,
        width_ratios=[0.18, 0.18, 0.50, 0.14],
        wspace=0.02, left=0.02, right=0.98, top=0.94, bottom=0.05,
    )
    ax_l = fig.add_subplot(gs[0, 0])
    ax_d = fig.add_subplot(gs[0, 1])
    ax_h = fig.add_subplot(gs[0, 2])
    ax_r = fig.add_subplot(gs[0, 3])
    ax_l.axis('off')
    ax_r.axis('off')

    # ---- Per-super-cluster distinguishable palette + link colouring ----
    # 11+ qualitative colours ordered so adjacent super-clusters contrast.
    SC_PALETTE = ['#e6194B', '#3cb44b', '#4363d8', '#f58231', '#911eb4',
                  '#42d4f4', '#f032e6', '#9A6324', '#469990', '#808000',
                  '#000075', '#e6beff', '#aaffc3', '#ffd8b1']
    leaf_sc_orig = [assign_map.get(int(c), 0) for c in cluster_ids]   # SC per ORIGINAL leaf
    sc_color = {sc: SC_PALETTE[i % len(SC_PALETTE)]
                for i, sc in enumerate(sorted(set(leaf_sc_orig)))}
    # SC of each linkage node (None if it spans >1 super-cluster -> drawn grey)
    members = {i: {i} for i in range(n)}
    node_sc = {}
    for i in range(len(Z)):
        a_, b_ = int(Z[i, 0]), int(Z[i, 1]); nid = n + i
        members[nid] = members[a_] | members[b_]
        scs = {leaf_sc_orig[l] for l in members[nid]}
        node_sc[nid] = next(iter(scs)) if len(scs) == 1 else None
    def _link_color(k):
        return sc_color.get(node_sc.get(k), '#999999')

    # ---- Dendrogram (leaves on right, NO leaf labels; coloured by super-cluster) ----
    dn = dendrogram(
        Z, orientation='left', no_labels=True,
        link_color_func=_link_color,
        ax=ax_d,
    )
    for line in ax_d.get_lines():
        line.set_linewidth(line.get_linewidth() * 1.6)
    ax_d.set_xscale('symlog', linthresh=0.5)   # same non-linear cosine-distance scale as before
    ax_d.invert_xaxis()
    ax_d.invert_yaxis()                   # leaf 0 at top
    ax_d.set_xlabel('Cosine distance (1 - similarity)  [Ward linkage, symlog]',
                    fontsize=11)
    ax_d.tick_params(axis='x', labelsize=9)
    ax_d.set_yticks([])
    for spine in ('top', 'right'):
        ax_d.spines[spine].set_visible(False)

    # ---- Heatmap (fills its square cell; aspect='auto' lets it stretch) ----
    ax_h.imshow(sim_d, aspect='auto', cmap='RdBu_r',
                vmin=-0.2, vmax=1.0, interpolation='none')
    for b in boundaries:
        ax_h.axhline(b - 0.5, color='black', lw=1.2, alpha=0.6)
        ax_h.axvline(b - 0.5, color='black', lw=1.2, alpha=0.6)
    ax_h.set_xticks([])
    ax_h.set_yticks([])
    ax_h.set_title(
        f'Cluster label cosine similarity ({len(cluster_ids)}x{len(cluster_ids)})  |  '
        f'red = >= {K_min} variable(s) with |CC|>{args.cc_sig:g} in >= {Y_min} all-ages year(s)\n'
        f'Rows/cols ordered by Ward linkage on cosine distance  '
        f'|  black lines = super-cluster boundaries (dendrogram cut, K={len(set(leaf_sc))})',
        fontsize=11, fontweight='bold', pad=8)

    # ---- Align dendrogram leaves with heatmap rows ----
    # scipy puts leaves at y = 5, 15, ..., 10n-5  (after invert: same set, top-down)
    # Heatmap rows are at y = 0, 1, ..., n-1
    # Setting matching ylim guarantees leaves and rows sit at the same
    # fractional vertical positions across all four panels.
    ax_d.set_ylim(10 * n, 0)
    ax_h.set_ylim(n - 0.5, -0.5)
    ax_l.set_xlim(0, 1)
    ax_l.set_ylim(n - 0.5, -0.5)

    # ---- Cluster labels on far left (right-aligned, flush against dendrogram) ----
    for i, (lbl, col) in enumerate(zip(leaf_lbl, leaf_color)):
        ax_l.text(0.99, i, lbl, ha='right', va='center',
                  fontsize=8.75, fontweight='bold', color=col,
                  transform=ax_l.transData)
    # ---- Super-cluster separator lines across the left label column ----
    for b in boundaries:
        ax_l.axhline(b - 0.5, color='#555555', lw=1.0)

    # ---- SC labels on far right, anchored to centre of each contiguous run ----
    fig.canvas.draw()
    trans = ax_h.transData + fig.transFigure.inverted()
    sc_x_fig = ax_r.get_position().x0 + 0.005
    for (s0, s1, sc_id) in runs:
        mid = (s0 + s1 - 1) / 2.0
        _, fy = trans.transform((0, mid))
        col = SIG_RED if sc_id in sig_scs else BASE_COL
        fig.text(sc_x_fig, fy,
                 f'SC{int(sc_id):02d}: {sc_name_map.get(int(sc_id), "")}',
                 va='center', ha='left',
                 fontsize=9, fontweight='bold', color=col)

    fig.savefig(args.out_dendro, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    _tee(args.out_dendro)


if __name__ == '__main__':
    main()
