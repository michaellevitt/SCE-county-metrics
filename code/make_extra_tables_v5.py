import os
import sys
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
#!/usr/bin/env python3
"""
make_extra_tables.py
====================
Build extra_tables.xlsx containing six tables:

  Table_1_and_2  -- Top 20 (+CC) + Top 10 (−CC) metrics
  Table_3        -- All super-clusters and XDE clusters (two-column layout)
  Table_4        -- XDE cluster x year CC heatmap (conditional formatting)
  Table_5        -- Super-cluster significance summary
  Table_Race     -- Race/ethnicity group findings

Usage:
  python3 make_extra_tables.py \
    --my-file        Multi_Year_data.csv          (or use --master-xlsx) \
    --all-file       Master_data.csv              (or use --master-xlsx) \
    --master-xlsx    master_xde_clusters.xlsx \
    --cc-file        metric_x_death_cc_0_0_25_1.csv \
    --output         extra_tables.xlsx
"""

import argparse
import re
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from _workbook_style import clean_style_workbook, detect_table_offset

# Significance mode: 'cc' (|CC| bands, default) or 'lp' (legacy p-value).
# In cc mode the headline "Significant" = |CC|>CC_SIG; the stricter
# "Very Significant" = |CC|>CC_VERYSIG. |CC| = max over the five all-age
# yearly correlations asedx_p_2020..asedx_p_2024.
SIG_MODE   = 'cc'
CC_SIG     = 0.3
CC_VERYSIG = 0.45

def _tee(path):
    """Print 'Saved <relpath>' to stdout and stderr."""
    _p = str(path)
    try:
        _rel = os.path.relpath(_p)
    except ValueError:
        _rel = _p
    msg = 'Saved ' + _rel
    print(msg)
    print(msg, file=sys.stderr, flush=True)


def save_fig(fig, path, dpi=200):
    """Save figure, adding the filename as a small label in the top margin."""
    import matplotlib.pyplot as plt
    fname = os.path.basename(path)
    fig.text(0.5, 0.995, fname, ha='center', va='top',
             fontsize=7, color='#888888', fontfamily='monospace',
             transform=fig.transFigure)
    """Save figure with filename in top margin."""
    fname = os.path.basename(path)
    fig.text(0.5, 0.995, fname, ha='center', va='top',
             fontsize=7, color='#888888', fontfamily='monospace',
             transform=fig.transFigure)
    fig.savefig(path, dpi=dpi, bbox_inches='tight', facecolor='white')
    try:
        _rel = os.path.relpath(path)
    except ValueError:
        _rel = path
    print('Saved ' + _rel, file=sys.stderr, flush=True)
    plt.close(fig)
    plt.close(fig)


YEARS = ['2020', '2021', '2022', '2023', '2024']

# ============================================================
# Style constants
# ============================================================
NAVY   = '1F3864'; MID    = '2E75B6'; WHITE  = 'FFFFFF'
GREEN  = 'E2EFDA'; AMBER  = 'FFF2CC'; RED_BG = 'FCE4D6'
LIGHT  = 'D6E4F0'; GREY   = 'F2F2F2'; DKGREY = '37474F'
DKRED  = '8B0000'; DKBLUE = '1F4E79'
LIGHT_RED  = 'FEF0EE'; LIGHT_BLUE = 'EBF5FB'

thin  = Side(style='thin',   color='CCCCCC')
thick = Side(style='medium', color='888888')

# SC_NAMES_FULL is populated dynamically from the master Excel in load_data()
# Default values are placeholders only -- overwritten at runtime
SC_NAMES_FULL = {}  # populated dynamically from master xlsx


# ============================================================
# Cell helpers
# ============================================================
def brd(c, top=False):
    c.border = Border(left=thin, right=thin,
                      top=(thick if top else thin), bottom=thin)
    return c

def cell(ws, r, c, val, bg=WHITE, fg='000000', bold=False,
         left=False, size=9, wrap=False, courier=False):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = Font(name='Courier New' if courier else 'Times New Roman',
                   bold=bold, color=fg, size=size)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(horizontal='left' if left else 'center',
                              vertical='center', wrap_text=wrap)
    return brd(cl)

def title_row(ws, ri, text, bg, size=12, height=22, ncols=7):
    ws.row_dimensions[ri].height = height
    for ci in range(1, ncols+1):
        cl = ws.cell(ri, ci, text if ci == 1 else '')
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=size)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)

def subtitle_row(ws, ri, text, bg, ncols=7):
    ws.row_dimensions[ri].height = 13
    for ci in range(1, ncols+1):
        cl = ws.cell(ri, ci, text if ci == 1 else '')
        cl.font = Font(name='Times New Roman', italic=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)

def spacer(ws, ri, ncols=7, height=5):
    ws.row_dimensions[ri].height = height
    for ci in range(1, ncols+1):
        ws.cell(ri, ci).fill = PatternFill('solid', start_color='F0F0F0')

def col_hdr_row(ws, ri, labels, bgs, height=45):
    ws.row_dimensions[ri].height = height
    for ci, (lbl, bg) in enumerate(zip(labels, bgs), 1):
        cl = ws.cell(ri, ci, lbl)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                  wrap_text=True)
        brd(cl)

def shorten_sc(sc):
    return f"SC{int(sc):02d}: {SC_NAMES_FULL.get(int(sc), '')}"

def clean_explain(v):
    return str(v).replace('_', ' ').strip()

# ============================================================
# Load data
# ============================================================
def load_data(args):
    print("Loading data...")
    if args.master_xlsx:
        wb = load_workbook(args.master_xlsx, data_only=True)

        def read_ws(name):
            ws = wb[name]
            hdr_row, hdr_col = detect_table_offset(ws)
            hdr = []
            c = hdr_col
            while c <= ws.max_column:
                v = ws.cell(hdr_row, c).value
                if v is None:
                    break
                hdr.append(str(v))
                c += 1
            n = len(hdr)
            return pd.DataFrame(
                [{hdr[i]: ws.cell(r, hdr_col + i).value for i in range(n)}
                 for r in range(hdr_row + 1, ws.max_row + 1)])

        my_df  = read_ws('Multi_Year')
        all_df = read_ws('Master')
        cl_df  = read_ws('Clusters')
    else:
        my_df  = pd.read_csv(args.my_file)
        all_df = pd.read_csv(args.all_file)
        cl_df  = None

    for df in [my_df, all_df]:
        for col in ['max_signed_cc', 'n_sig_years', 'super_cluster_id',
                    'Ward100', 'rank_abs_cc']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

    cc_raw = pd.read_csv(args.cc_file).set_index('metric')

    # Full cluster labels from Clusters tab
    clusters = {}
    if cl_df is not None:
        for _, r in cl_df.iterrows():
            w = r.get('Ward100')
            if w and pd.notna(w):
                clusters[int(w)] = str(r.get('cluster_label', ''))
    else:
        for _, r in all_df.drop_duplicates('Ward100').iterrows():
            w = r.get('Ward100')
            if w and pd.notna(w):
                clusters[int(w)] = str(r.get('cluster_label', ''))

    print(f"  Multi_Year: {len(my_df)} rows")
    print(f"  Master: {len(all_df)} rows")
    print(f"  Cluster labels: {len(clusters)}")

    # Build SC_NAMES_FULL from master data (overrides hardcoded defaults)
    global SC_NAMES_FULL
    sc_name_rows = all_df[['super_cluster_id','super_cluster_name']].dropna().drop_duplicates()
    for _, r in sc_name_rows.iterrows():
        try:
            sc_id = int(float(r['super_cluster_id']))
            name  = str(r['super_cluster_name']).strip()
            if name and name != 'nan':
                SC_NAMES_FULL[sc_id] = name
        except (ValueError, TypeError):
            pass
    print(f"  SC names loaded from master: {len(SC_NAMES_FULL)}")

    # Build sig1_df: any-year Significant (|CC|>CC_SIG in cc mode; LP<=-13 legacy)
    all_df['n_sig_years'] = pd.to_numeric(all_df.get('n_sig_years', 0), errors='coerce').fillna(0)
    if cc_raw is not None:
        YEARS_L   = ['2020','2021','2022','2023','2024']
        cc_cols_l = [f'asedx_p_{yr}' for yr in YEARS_L]
        lp_cols_l = [f'LP_asedx_p_{yr}' for yr in YEARS_L]
        cc_avail_l = [c for c in cc_cols_l if c in cc_raw.columns]
        lp_avail_l = [c for c in lp_cols_l if c in cc_raw.columns]
        def any_sig(m):
            if m not in cc_raw.index: return False
            if SIG_MODE == 'cc':
                return any(pd.notna(cc_raw.loc[m, c]) and abs(cc_raw.loc[m, c]) > CC_SIG
                           for c in cc_avail_l)
            return any(pd.notna(cc_raw.loc[m, c]) and cc_raw.loc[m, c] <= -13.0
                       for c in lp_avail_l)
        sig1_df = all_df[all_df['metric'].apply(any_sig)].copy()
    else:
        sig1_df = all_df[all_df['n_sig_years'] >= 1].copy()
    sig1_df['max_signed_cc'] = pd.to_numeric(sig1_df['max_signed_cc'], errors='coerce')
    # Convert all string columns to plain Python str to avoid StringArray issues
    for _col in ['SI', 'SI_near', 'sign', 'age', 'dominant_age']:
        if _col in sig1_df.columns:
            sig1_df[_col] = sig1_df[_col].astype(str).replace({'nan':'','None':''})
    sig1_df['min_lp']        = pd.to_numeric(sig1_df['min_lp'],        errors='coerce')
    sig1_df['n_sig_years']   = pd.to_numeric(sig1_df['n_sig_years'],   errors='coerce').fillna(0)
    # rank_abs_cc: rank by |max_signed_cc| descending (1=strongest)
    if 'rank_abs_cc' not in sig1_df.columns:
        sig1_df['rank_abs_cc'] = sig1_df['max_signed_cc'].abs().rank(
            ascending=False, method='first').astype(int)
    # Convert all_df string cols too
    for _col in ['SI', 'SI_near', 'sign', 'age', 'dominant_age']:
        if _col in all_df.columns:
            all_df[_col] = all_df[_col].astype(str).replace({'nan':'','None':''})
    print(f"  sig1_df (n_sig_years>=1): {len(sig1_df)} rows")

    return my_df, sig1_df, all_df, cc_raw, clusters

# ============================================================
# TABLE 1 & 2 -- Top +CC + Top -CC (named metrics)
# ============================================================
def write_table_1_2(wb, sig1_df):
    print("Writing Table_1_and_2...")
    ws = wb.create_sheet('Table_1_and_2')
    NCOLS = 7
    col_widths = [22, 28, 6, 50, 9, 8, 8]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    top_risky = sig1_df[sig1_df['max_signed_cc'] > 0].nsmallest(20, 'rank_abs_cc').copy()
    top_prot  = sig1_df[sig1_df['max_signed_cc'] < 0].nsmallest(10, 'rank_abs_cc').copy()
    for df in [top_risky, top_prot]:
        df['explain'] = df['explain'].apply(clean_explain)

    ri = 1
    title_row(ws, ri, 'Table 1.  Top 20 metrics  -- strongest positive (+CC) associations with pandemic excess mortality', NAVY, ncols=NCOLS); ri += 1
    title_row(ws, ri, 'Standard run: population-weighted (w1.0), no minimum-population floor  |  Kish N_eff = 283  |  significance |CC| > 0.30 (Very Significant |CC| > 0.45)', MID, size=9, height=14, ncols=NCOLS); ri += 1
    spacer(ws, ri, NCOLS); ri += 1
    col_hdr_row(ws, ri, ['Super-cluster', 'Cluster', 'Rank', 'Metric description', 'CC', 'SI', 'Sign'],
                [DKRED]*NCOLS); ri += 1

    for _, r in top_risky.iterrows():
        rk  = int(r['rank_abs_cc'])
        sc  = shorten_sc(r['super_cluster_id']) if pd.notna(r.get('super_cluster_id')) else ''
        lbl = str(r.get('cluster_label', '')).strip()
        bg  = LIGHT_RED if rk <= 10 else RED_BG
        ws.row_dimensions[ri].height = 28
        cell(ws, ri, 1, sc,  bg=bg, left=True, wrap=True)
        cell(ws, ri, 2, lbl, bg=bg, left=True, wrap=True)
        cell(ws, ri, 3, f'#{rk}', bg=bg, bold=True)
        cell(ws, ri, 4, clean_explain(r['explain']), bg=bg, left=True, wrap=True)
        cell(ws, ri, 5, round(float(r['max_signed_cc']), 3), bg=bg, bold=True, fg=DKRED)
        cell(ws, ri, 6, r['SI'],   bg=bg, courier=True, bold=True)
        cell(ws, ri, 7, r['sign'], bg=bg, courier=True)
        ri += 1

    spacer(ws, ri, NCOLS, height=12); ri += 1
    title_row(ws, ri, 'Table 2.  Top 10 metrics  -- strongest negative (−CC) associations with pandemic excess mortality', NAVY, ncols=NCOLS); ri += 1
    spacer(ws, ri, NCOLS, height=5); ri += 1
    col_hdr_row(ws, ri, ['Super-cluster', 'Cluster', 'Rank', 'Metric description', 'CC', 'SI', 'Sign'],
                [DKBLUE]*NCOLS); ri += 1

    for _, r in top_prot.iterrows():
        rk  = int(r['rank_abs_cc'])
        sc  = shorten_sc(r['super_cluster_id']) if pd.notna(r.get('super_cluster_id')) else ''
        lbl = str(r.get('cluster_label', '')).strip()
        bg  = LIGHT_BLUE if rk <= 15 else LIGHT
        ws.row_dimensions[ri].height = 28
        cell(ws, ri, 1, sc,  bg=bg, left=True, wrap=True)
        cell(ws, ri, 2, lbl, bg=bg, left=True, wrap=True)
        cell(ws, ri, 3, f'#{rk}', bg=bg, bold=True)
        cell(ws, ri, 4, clean_explain(r['explain']), bg=bg, left=True, wrap=True)
        cell(ws, ri, 5, round(float(r['max_signed_cc']), 3), bg=bg, bold=True, fg=DKBLUE)
        cell(ws, ri, 6, r['SI'],   bg=bg, courier=True, bold=True)
        cell(ws, ri, 7, r['sign'], bg=bg, courier=True)
        ri += 1

    ws.freeze_panes = 'A1'

# ============================================================
# TABLE 3 -- Super-cluster x XDE cluster (two-column layout)
# ============================================================
def write_table_3(wb, sig1_df, all_df, clusters, cc_df=None):
    print("Writing Table_3...")
    ws = wb.create_sheet('Table_3')

    # Build Very-Significant map: metric -> bool (any year |CC|>CC_VERYSIG in cc mode)
    near_map = {}
    if cc_df is not None:
        CC_COLS_N = [f'asedx_p_{yr}' for yr in ['2020','2021','2022','2023','2024']]
        LP_COLS_N = [f'LP_asedx_p_{yr}' for yr in ['2020','2021','2022','2023','2024']]
        cc_avail_n = [c for c in CC_COLS_N if c in cc_df.columns]
        lp_avail_n = [c for c in LP_COLS_N if c in cc_df.columns]
        for m in all_df['metric']:
            if m in cc_df.index:
                if SIG_MODE == 'cc':
                    near_map[m] = any(pd.notna(cc_df.loc[m, c]) and abs(cc_df.loc[m, c]) > CC_VERYSIG
                                      for c in cc_avail_n)
                else:
                    near_map[m] = any(pd.notna(cc_df.loc[m, c]) and cc_df.loc[m, c] <= -6.5
                                      for c in lp_avail_n)
            else:
                near_map[m] = False

    # Build kept clusters (n>=16 OR n_sig>0)
    all_rows = []
    for sc in sorted(all_df['super_cluster_id'].dropna().astype(int).unique()):
        a_sub = all_df[all_df['super_cluster_id'] == sc]
        s_sub = sig1_df[sig1_df['super_cluster_id'] == sc]
        cs = a_sub.groupby(['Ward100', 'cluster_label']).size().reset_index(name='n_metrics')
        cs['n_sig'] = cs['Ward100'].apply(
            lambda w: int((s_sub['Ward100'] == w).sum()))
        cs['n_near'] = cs['Ward100'].apply(
            lambda w: int(sum(1 for m in a_sub[a_sub['Ward100']==w]['metric']
                              if near_map.get(m, False))))
        cs_kept = cs[(cs['n_sig'] > 0) | (cs['n_metrics'] >= 16)].sort_values('n_metrics', ascending=False)
        n_near_sc = int(sum(near_map.get(m, False) for m in a_sub['metric']))
        all_rows.append(('sc', sc, {
            'sc': sc, 'sc_name': SC_NAMES_FULL.get(sc, ''),
            'n_clusters': len(cs_kept), 'n_clusters_total': len(cs),
            'n_metrics': len(a_sub), 'n_sig': int(s_sub.shape[0]),
            'n_near': n_near_sc,
        }))
        for _, cr in cs_kept.iterrows():
            n_m = int(cr['n_metrics'])
            n_s = int(cr['n_sig'])
            n_n = int(cr['n_near'])
            all_rows.append(('cluster', sc, {
                'sc': sc, 'ward': int(cr['Ward100']),
                'label': clusters.get(int(cr['Ward100']), str(cr['cluster_label'])),
                'n_metrics': n_m, 'n_sig': n_s, 'n_near': n_n,
                'pct_sig': round(n_s/n_m*100, 1) if n_m else 0,
            }))

    # Split at 33 cluster rows
    count = 0
    split_at = None
    for i, r in enumerate(all_rows):
        if r[0] == 'cluster':
            count += 1
            if count == 33:
                split_at = i + 1
                break

    left_rows  = all_rows[:split_at]
    right_rows = all_rows[split_at:]
    if right_rows and right_rows[0][0] == 'cluster':
        sc_of_first = right_rows[0][1]
        sc_hdr = next((r for r in reversed(left_rows)
                       if r[0] == 'sc' and r[1] == sc_of_first), None)
        if sc_hdr:
            right_rows = [sc_hdr] + right_rows

    col_widths = [5, 44, 6, 7, 7, 7, 7,  2,  5, 44, 6, 7, 7, 7, 7]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    NCOLS = 15
    ri = 1
    title_row(ws, ri, 'Table 3.  Super-Clusters and XDE Clusters  -- metrics and significance', NAVY, ncols=NCOLS); ri += 1
    subtitle_row(ws, ri,
        'Clusters with < 16 metrics and 0 significant omitted  |  # Sig = |CC|>0.30  |  # Very Sig = |CC|>0.45  |  Green=sig cluster  |  Amber=some sig  |  Red=SC has 0 sig',
        MID, ncols=NCOLS); ri += 1
    spacer(ws, ri, NCOLS, height=5); ri += 1

    ws.row_dimensions[ri].height = 30
    for offset in [0, 8]:
        for ci, lbl in enumerate(['SC', 'Super-cluster / Cluster', 'Ward', '# Met', '# Sig', '# Very Sig', '% Sig'], 1):
            cl = ws.cell(ri, offset+ci, lbl)
            cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
            cl.fill = PatternFill('solid', start_color=NAVY)
            cl.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            brd(cl)
    ri += 1

    def write_half_row(ws, ri, row_data, offset):
        if row_data is None:
            for ci in range(1, 8):
                ws.cell(ri, offset+ci).fill = PatternFill('solid', start_color='FAFAFA')
            return
        rtype, sc_int, data = row_data

        def mk(col, val, bg, bold=False, left=False, size=9):
            cl = ws.cell(ri, offset+col, val)
            cl.font = Font(name='Times New Roman', bold=bold, color='000000', size=size)
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if left else 'center',
                                      vertical='center', wrap_text=left)
            return brd(cl, top=(rtype == 'sc'))

        if rtype == 'sc':
            bg = RED_BG if data['n_sig'] == 0 else (GREEN if data['n_sig'] >= 20 else AMBER)
            mk(1, f"SC{data['sc']:02d}", bg, bold=True)
            mk(2, f"{data['sc_name']}  [{data['n_clusters']} of {data['n_clusters_total']} shown]",
               bg, bold=True, left=True)
            mk(3, '', bg)
            mk(4, data['n_metrics'], bg, bold=True)
            mk(5, data['n_sig'] if data['n_sig'] > 0 else '---', bg, bold=True)
            mk(6, data.get('n_near', '---') if data.get('n_near',0) > 0 else '---', bg)
            mk(7, '', bg)
        else:
            n_sig = data['n_sig']
            n_near = data.get('n_near', 0)
            pct_sig = data.get('pct_sig', 0)
            bg = GREEN if n_sig >= 10 else (LIGHT if n_sig > 0 else (WHITE if sc_int % 2 == 0 else GREY))
            mk(1, '', bg)
            mk(2, data['label'], bg, left=True, size=8)
            mk(3, data['ward'], bg)
            mk(4, data['n_metrics'], bg)
            mk(5, n_sig   if n_sig   > 0 else '', bg, bold=(n_sig > 0))
            mk(6, n_near  if n_near  > 0 else '', bg)
            mk(7, pct_sig if pct_sig > 0 else '', bg)

    max_rows = max(len(left_rows), len(right_rows))
    for i in range(max_rows):
        ws.row_dimensions[ri].height = 13
        write_half_row(ws, ri, left_rows[i]  if i < len(left_rows)  else None, 0)
        write_half_row(ws, ri, right_rows[i] if i < len(right_rows) else None, 8)
        ri += 1

    ws.freeze_panes = 'A5'

# ============================================================
# TABLE 4 -- XDE cluster x year (Excel with conditional formatting)
# ============================================================
def write_table_4(wb, sig1_df, cc_raw, clusters):
    print("Writing Table_4...")
    ws = wb.create_sheet('Table_4')

    ward_meta = {}
    for w in sig1_df['Ward100'].dropna().unique():
        sub = sig1_df[sig1_df['Ward100'] == w]
        sc  = int(sub['super_cluster_id'].iloc[0]) if pd.notna(sub['super_cluster_id'].iloc[0]) else 0
        ward_meta[int(w)] = {'sc': sc, 'n_sig': len(sub)}

    sig_wards = sorted(ward_meta.keys(),
                       key=lambda w: (ward_meta[w]['sc'], -ward_meta[w]['n_sig']))

    ward_cc   = {}
    ward_nsig = {}
    for w in sig_wards:
        sub     = sig1_df[sig1_df['Ward100'] == w]
        metrics = [m for m in sub['metric'] if m in cc_raw.index]
        ward_cc[w] = {}; ward_nsig[w] = {}
        for yr in YEARS:
            cc_col = f'asedx_p_{yr}'; lp_col = f'LP_asedx_p_{yr}'
            cs = cc_raw.loc[metrics, cc_col].dropna() if cc_col in cc_raw.columns and metrics else pd.Series()
            ls = cc_raw.loc[[m for m in metrics if m in cc_raw.index], lp_col].dropna() \
                 if lp_col in cc_raw.columns and metrics else pd.Series()
            ward_cc[w][yr]   = round(float(cs.mean()), 3) if len(cs) else None
            if SIG_MODE == 'cc':
                # count metrics with |CC| > CC_SIG (Significant) in this year
                ward_nsig[w][yr] = int((cs.abs() > CC_SIG).sum()) if len(cs) else 0
            else:
                ward_nsig[w][yr] = int((ls <= -13).sum()) if len(ls) else 0

    SC_MAP = {1:'SC01', 2:'SC02', 3:'SC03', 4:'SC04', 5:'SC05', 6:'SC06',
              7:'SC07', 8:'SC08', 9:'SC09', 10:'SC10', 11:'SC11'}

    col_widths = [6, 6, 65, 8, 8, 8, 8, 8,  2,  8, 8, 8, 8, 8]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = 1
    NCOLS = 14
    title_row(ws, ri, 'Table 4.  XDE cluster x year: mean CC and n significant metrics', NAVY, ncols=NCOLS); ri += 1
    subtitle_row(ws, ri,
        'Clusters with >= 1 significant metric (|CC|>0.30, any year)  |  CC: red = +CC, blue = −CC  |  n: yellow-orange-red scale',
        MID, ncols=NCOLS); ri += 1

    # Group headers
    ws.row_dimensions[ri].height = 16
    for ci in range(1, NCOLS+1):
        bg = DKRED if ci in range(4, 9) else (DKBLUE if ci in range(10, 15) else NAVY)
        lbl = 'Mean CC (2020-2024)' if ci == 4 else ('n Significant (|CC|>0.30)' if ci == 10 else '')
        cl = ws.cell(ri, ci, lbl)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg if ci != 9 else 'F0F0F0')
        cl.alignment = Alignment(horizontal='center', vertical='center')
        if ci != 9: brd(cl)
    ri += 1

    # Column headers
    ws.row_dimensions[ri].height = 40
    hdr_vals = ['SC', 'Ward', 'Cluster label',
                '2020', '2021', '2022', '2023', '2024', '',
                '2020', '2021', '2022', '2023', '2024']
    hdr_bgs  = [NAVY, NAVY, NAVY,
                DKRED, DKRED, DKRED, DKRED, DKRED, 'F0F0F0',
                DKBLUE, DKBLUE, DKBLUE, DKBLUE, DKBLUE]
    for ci, (h, bg) in enumerate(zip(hdr_vals, hdr_bgs), 1):
        if ci == 9:
            ws.cell(ri, ci).fill = PatternFill('solid', start_color='F0F0F0'); continue
        cl = ws.cell(ri, ci, h)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=90)
        brd(cl)
    ri += 1

    data_start = ri
    prev_sc = None
    for w in sig_wards:
        sc  = ward_meta[w]['sc']
        lbl = clusters.get(w, '')
        bg  = GREY if ri % 2 == 0 else WHITE
        top = (prev_sc is not None and sc != prev_sc)
        ws.row_dimensions[ri].height = 14

        def c4(col, val, bold=False, left=False):
            cl = ws.cell(ri, col, val)
            cl.font = Font(name='Times New Roman', bold=bold, color='000000', size=9)
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='left' if left else 'center', vertical='center')
            return brd(cl, top=top)

        c4(1, SC_MAP.get(sc, ''), bold=True)
        c4(2, w, bold=True)
        c4(3, lbl, left=True)
        for yi, yr in enumerate(YEARS):
            v = ward_cc[w][yr]
            c4(4+yi, round(v, 3) if v is not None else '')
        ws.cell(ri, 9).fill = PatternFill('solid', start_color='F0F0F0'); brd(ws.cell(ri, 9))
        for yi, yr in enumerate(YEARS):
            v = ward_nsig[w][yr]
            c4(10+yi, v if v > 0 else '')
        prev_sc = sc
        ri += 1

    data_end = ri - 1

    # Conditional formatting
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.conditional_formatting.add(f'{col}{data_start}:{col}{data_end}',
            ColorScaleRule(start_type='num', start_value=-0.38, start_color='922B21',
                           mid_type='num',   mid_value=0,       mid_color='FFFFFF',
                           end_type='num',   end_value=0.38,    end_color='1A5276'))
    for col in ['J', 'K', 'L', 'M', 'N']:
        ws.conditional_formatting.add(f'{col}{data_start}:{col}{data_end}',
            ColorScaleRule(start_type='num', start_value=0,   start_color='FFFFFF',
                           mid_type='num',   mid_value=50,    mid_color='FFC000',
                           end_type='num',   end_value=200,   end_color='C0392B'))

    ws.freeze_panes = f'D{data_start}'

# ============================================================
# TABLE 5 -- Super-cluster significance summary
# ============================================================
def write_table_5(wb, sig1_df, all_df, cc_df=None):
    """Table 5: Super-cluster significance summary.
    Columns: SC | Name | Total | Significant (|CC|>0.30) | Very Significant (|CC|>0.45) |
             % Significant | % Very Sig | Mean|CC| | Top SI
    Headline / sort key = % Significant (|CC|>0.30).
    All counts include single-year significant metrics.
    Times New Roman throughout.
    """
    print("Writing Table_5...")

    YEARS     = ['2020','2021','2022','2023','2024']
    LP_PRI    = -13.0
    LP_NEAR   =  -6.5
    TIMES     = 'Times New Roman'

    # ---- Build per-metric sig flags from CC file ----
    # sig_S_map  = headline Significant      (any year |CC|>CC_SIG)
    # sig_V_map  = stricter Very Significant (any year |CC|>CC_VERYSIG)
    sig_S_map = {}
    sig_V_map = {}
    if cc_df is not None:
        lp_cols = [f'LP_asedx_p_{yr}' for yr in YEARS]
        cc_cols = [f'asedx_p_{yr}' for yr in YEARS]
        lp_avail = [c for c in lp_cols if c in cc_df.columns]
        cc_avail = [c for c in cc_cols if c in cc_df.columns]
        for m in all_df['metric']:
            if m in cc_df.index:
                if SIG_MODE == 'cc':
                    cc_vals = [abs(cc_df.loc[m, c]) for c in cc_avail if pd.notna(cc_df.loc[m, c])]
                    sig_S_map[m] = any(v > CC_SIG     for v in cc_vals)  # Significant
                    sig_V_map[m] = any(v > CC_VERYSIG for v in cc_vals)  # Very Significant
                else:
                    lp_vals = [cc_df.loc[m, c] for c in lp_avail if pd.notna(cc_df.loc[m, c])]
                    sig_S_map[m] = any(v <= LP_NEAR for v in lp_vals)
                    sig_V_map[m] = any(v <= LP_PRI  for v in lp_vals)
            else:
                sig_S_map[m] = sig_V_map[m] = False
    else:
        my_set = set(sig1_df['metric'].tolist())
        for m in all_df['metric']:
            sig_S_map[m] = m in my_set
            sig_V_map[m] = m in my_set

    # ---- Build SC data ----
    sc_data = []
    for sc in sorted(all_df['super_cluster_id'].dropna().astype(int).unique()):
        a_sub = all_df[all_df['super_cluster_id'] == sc].copy()
        a_sub['max_signed_cc'] = pd.to_numeric(a_sub['max_signed_cc'], errors='coerce')

        s_mask = a_sub['metric'].apply(lambda m: sig_S_map.get(m, False))
        v_mask = a_sub['metric'].apply(lambda m: sig_V_map.get(m, False))
        s_sig = a_sub[s_mask]   # Significant
        s_ver = a_sub[v_mask]   # Very Significant

        n_all     = len(a_sub)
        n_sig     = len(s_sig)
        n_ver     = len(s_ver)
        pct       = round(n_sig/n_all*100, 1) if n_all else 0
        pct_ver   = round(n_ver/n_all*100, 1) if n_all else 0
        mac       = round(float(s_sig['max_signed_cc'].abs().mean()), 3) if n_sig else None

        # Sign-invariant: dominant temporal SI pattern (the meaningful descriptor).
        top_si = (s_sig['SI'].value_counts().index[0] if n_sig else '---')

        sc_data.append({
            'sc': sc, 'name': SC_NAMES_FULL.get(sc, ''),
            'n_all': n_all, 'n_sig': n_sig, 'n_ver': n_ver,
            'pct': pct, 'pct_ver': pct_ver,
            'mac': mac, 'top_si': top_si,
        })

    sc_sorted = sorted(sc_data, key=lambda x: -x['pct'])

    # ---- Worksheet setup ----
    if 'Table_5' in wb.sheetnames: wb.remove(wb['Table_5'])
    ws = wb.create_sheet('Table_5')
    NCOLS = 9
    # SC | Name | Total | Significant | Very Significant | % Sig | % Very Sig | Mean|CC| | Top SI
    col_widths = [6, 48, 9, 13, 14, 12, 11, 10, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    def tc(ws, r, c, val, bg=WHITE, fg='000000', bold=False,
           left=False, size=10, italic=False, courier=False):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(name='Courier New' if courier else TIMES,
                       bold=bold, italic=italic, color=fg, size=size)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='left' if left else 'center',
                                  vertical='center', wrap_text=False)
        brd(cl); return cl

    ri = 1
    # Title
    ws.row_dimensions[ri].height = 22
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci,
            'Table 5.  Super-cluster significance summary  -- % of metrics significant (any year, |CC| > 0.30)'
            if ci==1 else '')
        cl.font  = Font(name=TIMES, bold=True, color='FFFFFF', size=12)
        cl.fill  = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    # Subtitle
    ws.row_dimensions[ri].height = 13
    sub = ('Sorted by % Significant descending  |  Counts include single-year significant metrics  |  '
           'Significant = |CC|>0.30  |  Very Significant = |CC|>0.45  |  '
           'Green=any significant  |  Red=zero significant')
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci, sub if ci==1 else '')
        cl.font  = Font(name=TIMES, italic=True, color='FFFFFF', size=9)
        cl.fill  = PatternFill('solid', start_color=MID)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    spacer(ws, ri, NCOLS); ri += 1

    # Column headers (tall, bottom-aligned, wrap)
    ws.row_dimensions[ri].height = 55
    hdrs = ['SC', 'Super-cluster name',
            'Total\nmetrics',
            'Significant\n(|CC|>0.30)',
            'Very Significant\n(|CC|>0.45)',
            '%\nSignificant',
            '%\nVery Sig',
            'Mean\n|CC|',
            'Top SI\npattern']
    for ci, h in enumerate(hdrs, 1):
        cl = ws.cell(ri, ci, h)
        cl.font  = Font(name=TIMES, bold=True, color='FFFFFF', size=10)
        cl.fill  = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                  wrap_text=True)
        brd(cl)
    ri += 1

    data_start = ri
    for r in sc_sorted:
        ws.row_dimensions[ri].height = 16
        pct = r['pct']
        # Row background: red=0%, green>=20%, amber=some, white
        bg  = RED_BG if pct == 0 else (GREEN if pct >= 20 else (AMBER if pct > 0 else WHITE))

        tc(ws, ri, 1,  f"SC{r['sc']:02d}",                            bg=bg, bold=True)
        tc(ws, ri, 2,  r['name'],                                       bg=bg, left=True)
        tc(ws, ri, 3,  r['n_all'],                                      bg=bg)
        tc(ws, ri, 4,  r['n_sig'] if r['n_sig'] > 0 else '---',        bg=bg, bold=(r['n_sig']>0))
        tc(ws, ri, 5,  r['n_ver'] if r['n_ver'] > 0 else '---',        bg=bg, bold=(r['n_ver']>0))
        tc(ws, ri, 6,  pct,                                              bg=bg, bold=True)
        tc(ws, ri, 7,  r['pct_ver'] if r['pct_ver'] else '---',         bg=bg)
        tc(ws, ri, 8,  r['mac']     if r['mac']     else '---',         bg=bg)
        tc(ws, ri, 9,  r['top_si'],  bg=bg, courier=True, bold=(r['n_sig']>0))
        ri += 1

    from openpyxl.formatting.rule import ColorScaleRule
    ws.conditional_formatting.add(f'F{data_start}:F{ri-1}',
        ColorScaleRule(start_type='num', start_value=0,  start_color='FCE4D6',
                       mid_type='num',   mid_value=15,   mid_color='FFF2CC',
                       end_type='num',   end_value=40,   end_color='E2EFDA'))
    ws.conditional_formatting.add(f'G{data_start}:G{ri-1}',
        ColorScaleRule(start_type='num', start_value=0,  start_color='FCE4D6',
                       mid_type='num',   mid_value=30,   mid_color='FFF2CC',
                       end_type='num',   end_value=70,   end_color='E2EFDA'))
    ws.freeze_panes = 'A5'


# ============================================================
# TABLE RACE -- Race / ethnicity findings
# ============================================================
def write_table_race(wb, sig1_df):
    print("Writing Table_Race...")
    ws = wb.create_sheet('Table_Race')
    NCOLS = 7
    col_widths = [34, 6, 9, 18, 44, 44, 44]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    race_defs = [
        ('White non-Hispanic',
         r'[Ww]hite.?[Nn]on.?[Hh]ispanic|[Ww]hite_[Nn]on.?[Hh]ispanic',
         None,
         'Counties with high White non-Hispanic income, education, and labour force had substantially lower pandemic mortality throughout 2020-2024'),
        ('Hispanic / Latino',
         r'[Hh]ispanic|[Ll]atino',
         r'[Ww]hite.?[Nn]on.?[Hh]ispanic',
         'Hispanic poverty, low education, and large family size predict higher mortality. Earlier apparent paradox was artefact of White non-Hispanic metrics being incorrectly included'),
        ('AIAN (American Indian / Alaska Native)',
         r'[Aa]merican.?[Ii]ndian|[Aa]laska.?[Nn]ative',
         None,
         'Every AIAN metric is +CC. Concentrated in poverty, income below poverty, and age-sex demographics. Pattern strongest 2020-21 (SSIII / ISSII)'),
        ('Multiracial / Some Other Race',
         r'[Mm]ultiracial|[Ss]ome.?[Oo]ther.?[Rr]ace',
         None,
         'Population size and age-sex distribution metrics; all +CC; moderate effect sizes'),
        ('Asian',
         r'\b[Aa]sian\b',
         r'[Hh]ispanic|[Ll]atino|[Ww]hite|[Bb]lack|[Nn]ative',
         'Asian population share, income, and labour force metrics reach no significance in any year'),
        ('Black / African American',
         r'\b[Bb]lack\b|[Aa]frican.?[Aa]merican',
         None,
         'AHRF Black/AA population metrics do not reach the moderate band (|CC|>0.30) here. Signals appear indirectly via household structure and poverty metrics'),
    ]

    table_rows = []
    for grp, pat, excl, interp in race_defs:
        sub = sig1_df[sig1_df['explain'].str.contains(pat, na=False, regex=True, case=False)]
        if excl:
            sub = sub[~sub['explain'].str.contains(excl, na=False, regex=True, case=False)]
        n   = len(sub)
        np_ = int((sub['max_signed_cc'] > 0).sum())
        nn  = int((sub['max_signed_cc'] < 0).sum())
        mac = round(float(sub['max_signed_cc'].abs().mean()), 3) if n > 0 else None
        top_si = sub['SI'].value_counts().index[0] if n > 0 else '---'
        tr = sub[sub['max_signed_cc'] > 0].nsmallest(1, 'rank_abs_cc') if np_ > 0 else pd.DataFrame()
        tp = sub[sub['max_signed_cc'] < 0].nsmallest(1, 'rank_abs_cc') if nn > 0  else pd.DataFrame()
        tr_str = (clean_explain(tr['explain'].values[0]) + f"  (CC={float(tr['max_signed_cc'].values[0]):+.3f}, {tr['SI'].values[0]})") if len(tr) else '---'
        tp_str = (clean_explain(tp['explain'].values[0]) + f"  (CC={float(tp['max_signed_cc'].values[0]):+.3f}, {tp['SI'].values[0]})") if len(tp) else '---'
        table_rows.append([grp, n if n > 0 else 0,
                           f'{mac:.3f}' if mac is not None else '---',
                           top_si, interp, tr_str, tp_str, np_, nn])

    ri = 1
    title_row(ws, ri, 'Table 6.  Race / ethnicity group findings  -- moderate-association metrics (|CC|>0.30, any year, n_sig_years >= 1)', NAVY, ncols=NCOLS); ri += 1
    subtitle_row(ws, ri,
        'Note: Hispanic metrics exclude White non-Hispanic labels. Black/AA not reaching the moderate band in AHRF metrics.',
        MID, ncols=NCOLS); ri += 1
    spacer(ws, ri, NCOLS); ri += 1
    col_hdr_row(ws, ri,
        ['Race / ethnicity group', 'n moderate\nmetrics',
         'Mean\n|CC|', 'Top SI\npattern', 'Interpretation',
         'Strongest +CC metric', 'Strongest −CC metric'],
        [NAVY]*NCOLS); ri += 1

    for r in table_rows:
        grp, n, mac, si, interp, tr, tp, np_, nn = r
        ws.row_dimensions[ri].height = 45
        if n == 0:                   bg = GREY
        elif nn > np_:               bg = LIGHT
        elif np_ > nn * 2:           bg = RED_BG
        elif np_ > nn:               bg = AMBER
        else:                        bg = GREY
        cell(ws, ri, 1,  grp,      bg=bg, bold=True, left=True)
        cell(ws, ri, 2,  n,        bg=bg, bold=(n > 0))
        cell(ws, ri, 3,  mac,      bg=bg, bold=True)
        cell(ws, ri, 4,  si,       bg=bg)
        cell(ws, ri, 5,  interp,   bg=bg, left=True, wrap=True, size=8)
        cell(ws, ri, 6,  tr,       bg=bg, left=True, wrap=True, size=8)
        cell(ws, ri, 7,  tp,       bg=bg, left=True, wrap=True, size=8)
        ri += 1

    ws.freeze_panes = 'A5'


# ============================================================
# TABLE_SC_CLUSTER_SI -- SI pattern distribution by SC and cluster
# ============================================================
def write_table_sc_cluster_si(wb, sig1_df, all_df):
    print("Writing Table_SC_Cluster_SI...")

    SI_COLS  = ['SSSSS','SSIII','SSSII','ISSSS','ISSII','SSSIS','Other']
    HAS_NEAR = 'SI_near' in sig1_df.columns

    def si_vals(si_cts, n_sig):
        out = []
        for si_pat in SI_COLS:
            if si_pat == 'Other':
                v = int(n_sig - sum(si_cts.get(s, 0) for s in SI_COLS[:-1]))
            else:
                v = int(si_cts.get(si_pat, 0))
            out.append(max(v, 0))
        return out

    if 'Table_SC_Cluster_SI' in wb.sheetnames:
        wb.remove(wb['Table_SC_Cluster_SI'])
    ws = wb.create_sheet('Table_SC_Cluster_SI')

    NCOLS = 6 + len(SI_COLS)
    NCOLS_TOTAL = NCOLS + (1 if HAS_NEAR else 0)
    col_widths = [5, 6, 46, 8, 8, 8] + [8]*len(SI_COLS) + ([9] if HAS_NEAR else [])
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = 1
    NFULL = 6 + len(SI_COLS) + (1 if HAS_NEAR else 0)

    def title_row_ws(ws, ri, text, bg, size=12, height=22, ncols=NCOLS):
        ws.row_dimensions[ri].height = height
        for ci in range(1, ncols+1):
            cl = ws.cell(ri, ci, text if ci==1 else '')
            cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=size)
            cl.fill = PatternFill('solid', start_color=bg)
            cl.alignment = Alignment(horizontal='center', vertical='center')
            brd(cl)

    title_row_ws(ws, ri,
        'Table S-SC.  SI temporal pattern distribution by Super-Cluster and XDE Cluster',
        NAVY, 12, 22, NFULL); ri += 1
    subtitle_row(ws, ri,
        'S = significant (|CC|>0.30)  |  I = not significant  |  includes single-year significant metrics',
        MID, NFULL); ri += 1
    spacer(ws, ri, NFULL); ri += 1

    # Group header for SI cols
    ws.row_dimensions[ri].height = 16
    for ci in range(1, NFULL+1):
        bg2 = MID if ci > 6 else NAVY
        lbl = 'SI pattern counts (n significant metrics)' if ci == 7 else ''
        cl = ws.cell(ri, ci, lbl)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg2)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    # Column headers
    ws.row_dimensions[ri].height = 55
    hdrs = ['SC','Ward','Super-cluster / Cluster name','n total\nmetrics','n sig\nmetrics',
            '% sig'] + SI_COLS + (['n multi-year\nsignificant'] if HAS_NEAR else [])
    for ci, h in enumerate(hdrs, 1):
        bg2 = MID if ci > 6 else NAVY
        cl = ws.cell(ri, ci, h)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=bg2)
        cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                  wrap_text=True, text_rotation=90 if ci > 6 else 0)
        brd(cl)
    ri += 1

    # Data
    sig = sig1_df.copy()
    sig['Ward100'] = pd.to_numeric(sig['Ward100'], errors='coerce')
    sig['max_signed_cc'] = pd.to_numeric(sig['max_signed_cc'], errors='coerce')
    if 'SI_near' not in sig.columns:
        sig['SI_near'] = sig['SI']  # fallback if not present
    # Convert SI_near to plain str to avoid StringArray type issues
    sig['SI_near'] = sig['SI_near'].astype(str)

    for sc in sorted(all_df['super_cluster_id'].dropna().astype(int).unique()):
        a_sc = all_df[all_df['super_cluster_id']==sc]
        s_sc = sig[sig['super_cluster_id']==sc]
        n_all_sc = len(a_sc); n_sig_sc = len(s_sc)
        pct_sc   = round(n_sig_sc/n_all_sc*100, 1) if n_all_sc else 0
        si_sc    = s_sc['SI'].value_counts().to_dict()
        sc_bg    = RED_BG if n_sig_sc==0 else (GREEN if pct_sc>=20 else AMBER)

        # SC summary row
        ws.row_dimensions[ri].height = 16
        cell(ws, ri, 1, f'SC{sc:02d}', bg=sc_bg, bold=True, size=10)
        cell(ws, ri, 2, '',            bg=sc_bg)
        cell(ws, ri, 3, SC_NAMES_FULL.get(sc,''), bg=sc_bg, bold=True, left=True, size=10)
        cell(ws, ri, 4, n_all_sc,      bg=sc_bg, bold=True)
        cell(ws, ri, 5, n_sig_sc if n_sig_sc>0 else '---', bg=sc_bg, bold=True)
        cell(ws, ri, 6, pct_sc,        bg=sc_bg, bold=True)
        for ci_off, v in enumerate(si_vals(si_sc, n_sig_sc), 7):
            cell(ws, ri, ci_off, v if v>0 else '', bg=sc_bg,
                 courier=(v>0), bold=(v>0), size=10)
        if HAS_NEAR:
            n_near = int(s_sc['SI_near'].apply(
                lambda x: (str(x).count('S')+str(x).count('s'))>1
                          if (pd.notna(x) and str(x).strip() != '') else False).sum() or 0)
            cell(ws, ri, NFULL, n_near if n_near>0 else '', bg=sc_bg, fg='555555', bold=(n_near>0))
        for ci in range(1, NFULL+1):
            ws.cell(ri, ci).border = Border(left=thin, right=thin, top=thick, bottom=thin)
        ri += 1

        # Cluster rows
        cl_stats = []
        for w in s_sc['Ward100'].dropna().unique():
            s_sub = s_sc[s_sc['Ward100']==w]
            a_sub = a_sc[a_sc['Ward100']==w]
            cl_stats.append({
                'ward': int(w),
                'label': str(s_sub['cluster_label'].iloc[0]),
                'n_all': len(a_sub), 'n_sig': len(s_sub),
                'n_pos': int((s_sub['max_signed_cc']>0).sum()),
                'n_neg': int((s_sub['max_signed_cc']<0).sum()),
                'si_cts': s_sub['SI'].value_counts().to_dict()
            })
        cl_stats.sort(key=lambda x: -x['n_sig'])

        for cs in cl_stats:
            n_sig_cl = cs['n_sig']
            cl_bg = GREEN if n_sig_cl>=10 else (LIGHT if n_sig_cl>=5 else (GREY if ri%2==0 else WHITE))
            ws.row_dimensions[ri].height = 13
            cell(ws, ri, 1, '',           bg=cl_bg)
            cell(ws, ri, 2, cs['ward'],   bg=cl_bg, bold=True, size=8)
            cell(ws, ri, 3, '  '+cs['label'], bg=cl_bg, left=True, size=8)
            cell(ws, ri, 4, cs['n_all'],  bg=cl_bg, size=8)
            cell(ws, ri, 5, n_sig_cl,     bg=cl_bg, bold=True, size=8)
            cell(ws, ri, 6, '',           bg=cl_bg)
            for ci_off, v in enumerate(si_vals(cs['si_cts'], n_sig_cl), 7):
                cell(ws, ri, ci_off, v if v>0 else '', bg=cl_bg,
                     courier=(v>0), bold=(v>0), size=8)
            if HAS_NEAR:
                si_near_vals = sig[sig['Ward100']==cs['ward']]['SI_near'].tolist()
                n_near_cl = sum(1 for s in si_near_vals
                                if (pd.notna(s) and str(s).strip() != ''
                                    and (str(s).count('S')+str(s).count('s'))>1))
                cell(ws, ri, NFULL, n_near_cl if n_near_cl>0 else '', bg=cl_bg,
                     fg='555555', size=8)
            ri += 1

    # Totals row
    tot_sig   = len(sig)
    si_cts_all = sig['SI'].value_counts().to_dict()
    ws.row_dimensions[ri].height = 16
    cell(ws, ri, 1, 'ALL',  bg=NAVY, fg='FFFFFF', bold=True)
    cell(ws, ri, 2, '',     bg=NAVY)
    cell(ws, ri, 3, 'All super-clusters', bg=NAVY, fg='FFFFFF', bold=True, left=True)
    cell(ws, ri, 4, len(all_df), bg=NAVY, fg='FFFFFF', bold=True)
    cell(ws, ri, 5, tot_sig,     bg=NAVY, fg='FFFFFF', bold=True)
    cell(ws, ri, 6, round(tot_sig/len(all_df)*100,1), bg=NAVY, fg='FFFFFF', bold=True)
    for ci_off, v in enumerate(si_vals(si_cts_all, tot_sig), 7):
        cell(ws, ri, ci_off, v if v>0 else '', bg=NAVY, fg='FFFFFF', bold=(v>0), courier=(v>0))
    if HAS_NEAR:
        n_near_tot = int(sig['SI_near'].apply(
            lambda x: (str(x).count('S')+str(x).count('s'))>1
                      if (pd.notna(x) and str(x).strip() != '') else False).sum() or 0)
        cell(ws, ri, NFULL, n_near_tot, bg=NAVY, fg='FFFFFF', bold=True)
    for ci in range(1, NFULL+1):
        ws.cell(ri, ci).border = Border(left=thin, right=thin, top=thick, bottom=thick)

    ws.freeze_panes = 'A6'


# ============================================================
# TABLE_SI -- SI pattern summary table (Excel)
# ============================================================
def write_table_si(wb, all_df, my_df, sig1_df=None):
    """SI pattern table with interpretation column.
    sig1_df: n_sig_years>=1 (any year); my_df: multi-year for TOTAL row.
    """
    print("Writing Table_SI...")

    sig_df = sig1_df.copy() if sig1_df is not None else all_df[pd.to_numeric(all_df.get('n_sig_years', 0), errors='coerce').fillna(0) >= 1].copy()
    sig_df['max_signed_cc'] = pd.to_numeric(sig_df['max_signed_cc'], errors='coerce')

    pos = sig_df[sig_df['max_signed_cc'] > 0]
    neg = sig_df[sig_df['max_signed_cc'] < 0]
    si_pos = pos['SI'].value_counts()
    si_neg = neg['SI'].value_counts()
    all_si = sorted(set(si_pos.index) | set(si_neg.index),
                    key=lambda x: -(si_pos.get(x, 0)+si_neg.get(x, 0)))

    INTERP = {
        'SSIII': 'Significant in 2020-21 only; most common pattern; signal concentrated in early pandemic',
        'SSSSS': 'Significant all 5 years; strongest effect sizes; persistent socioeconomic gradient',
        'SSSII': 'Significant 2020-22; faded after Omicron',
        'ISSSS': 'Lagged onset 2021-24; predominantly labour force and income metrics',
        'ISSII': 'Lagged onset 2021-22 only; mixed domains',
        'SSSSI': 'Significant 2020-23, recovery in 2024',
        'SIIII': 'Significant 2020 only',
        'ISIII': 'Significant 2021 only',
        'IIIIS': 'Significant 2024 only; late-rising burden metrics',
    }

    if 'Table_SI' in wb.sheetnames: wb.remove(wb['Table_SI'])
    ws = wb.create_sheet('Table_SI')

    NCOLS = 4
    col_widths = [12, 12, 13, 52]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = 1
    ws.row_dimensions[ri].height = 22
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci,
            'Table SI.  Distribution of significant metrics by temporal pattern' if ci==1 else '')
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=12)
        cl.fill = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    ws.row_dimensions[ri].height = 13
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci,
            'S = significant (|CC|>0.30)  |  I = not significant  |  includes single-year significant metrics' if ci==1 else '')
        cl.font = Font(name='Times New Roman', italic=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=MID)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    ws.row_dimensions[ri].height = 5
    for ci in range(1, NCOLS+1):
        ws.cell(ri, ci).fill = PatternFill('solid', start_color='F0F0F0')
    ri += 1

    # Column headers (sign-invariant)
    ws.row_dimensions[ri].height = 40
    hdrs = ['SI Pattern', 'Total\nnumber', 'Mean\n|CC|', 'Interpretation']
    for ci, h in enumerate(hdrs, 1):
        cl = ws.cell(ri, ci, h)
        cl.font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        brd(cl)
    ri += 1

    data_start = ri
    is_main = lambda si, idx: idx < 5 and si in INTERP

    for idx, si in enumerate(all_si):
        sp = pos[pos['SI']==si]; sn = neg[neg['SI']==si]
        tot = len(sp)+len(sn)
        sub_all = sig_df[sig_df['SI']==si]
        mac = round(float(sub_all['max_signed_cc'].abs().mean()), 3) if tot else None
        note = INTERP.get(si, f'Minor pattern (n={tot})')
        minor = (tot <= 7)

        bg   = GREY if minor else (WHITE if idx % 2 == 0 else 'F7F7F7')
        fg_c = '999999' if minor else '000000'
        bold = is_main(si, idx)

        ws.row_dimensions[ri].height = 14
        cell(ws, ri, 1, si,  bg=bg, fg=fg_c, bold=bold, courier=True, size=10)
        cell(ws, ri, 2, tot, bg=bg, fg=fg_c, bold=bold)
        cell(ws, ri, 3, mac if mac is not None else '---', bg=bg, fg=fg_c, bold=bold)
        c = ws.cell(ri, NCOLS, note)
        c.font = Font(name='Times New Roman', size=8.5, color=fg_c, italic=minor)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        brd(c)
        ri += 1

    # Totals row
    s1 = sig1_df if sig1_df is not None else my_df
    s1 = s1.copy(); s1['max_signed_cc'] = pd.to_numeric(s1['max_signed_cc'], errors='coerce')
    ws.row_dimensions[ri].height = 16
    cell(ws, ri, 1, 'TOTAL', bg=NAVY, fg='FFFFFF', bold=True, courier=True, size=10)
    cell(ws, ri, 2, len(sig1_df),  bg=NAVY, fg='FFFFFF', bold=True)
    mac_t = round(float(s1['max_signed_cc'].abs().mean()), 3) if len(s1) else None
    cell(ws, ri, 3, mac_t if mac_t is not None else '---', bg=NAVY, fg='FFFFFF', bold=True)
    c = ws.cell(ri, NCOLS, f'All {len(sig1_df)} significant metrics (any year)')
    c.font = Font(name='Times New Roman', size=9, color='FFFFFF', bold=True)
    c.fill = PatternFill('solid', start_color=NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center')
    brd(c)
    for ci in range(1, NCOLS+1):
        ws.cell(ri, ci).border = Border(left=Side(style='thin', color='CCCCCC'),
                                         right=Side(style='thin', color='CCCCCC'),
                                         top=Side(style='medium', color='888888'),
                                         bottom=Side(style='medium', color='888888'))
    ws.freeze_panes = 'A5'

# ============================================================
# Main
# ============================================================
def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--master-xlsx',    default=None,
                   help='master_xde_clusters.xlsx (reads Multi_Year, Master, Clusters tabs)')
    p.add_argument('--my-file',        default=None, help='Multi_Year CSV (alternative to --master-xlsx)')
    p.add_argument('--all-file',       default=None, help='Master CSV (alternative to --master-xlsx)')
    p.add_argument('--cc-file',        required=True)
    p.add_argument('--output',         default='extra_tables.xlsx')
    p.add_argument('--z-med',          default=None,
                   help='Z_med_xde100.npy for XDE SC12 cut')
    p.add_argument('--medoid-list',    default=None,
                   help='medoid_list_xde100.json')
    p.add_argument('--reps',           default=None,
                   help='xde100_reps.csv')
    p.add_argument('--sem-sc',         default=None,
                   help='sem_sc_assignments.csv')
    p.add_argument('--sem-names',      default=None,
                   help='sem_sc_names.csv')
    p.add_argument('--sig-mode', choices=['lp', 'cc'], default='cc',
                   help="Significance definition: 'cc' (|CC| bands, default) or 'lp' (legacy)")
    p.add_argument('--cc-sig', type=float, default=0.3, help='|CC| Significant (cc mode)')
    p.add_argument('--cc-verysig', type=float, default=0.45, help='|CC| Very Significant (cc mode)')
    return p.parse_args()

def write_table_temporal(wb, sig1_df, cc_df):
    """Temporal trajectory table matching fig2: Year x |CC| stats (sign-invariant).
    The sign of a metric's CC is a coding convention, so we pool ALL significant
    metrics and report |CC| statistics rather than splitting into risky/protective."""
    print("Writing Table_Temporal...")
    ws = wb.create_sheet('Table_Temporal')

    YEARS = ['2020', '2021', '2022', '2023', '2024']

    # Column widths: Year | n sig | Mean |CC| | Min |CC| | Max |CC|
    col_widths = [8, 10, 12, 12, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    NCOLS = 5
    NAVY = '1F3864'; MID = '2E75B6'; WHITE = 'FFFFFF'
    LIGHT = 'F2F6FB'
    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color='888888')

    def brd(c, top=False):
        c.border = Border(left=thin, right=thin,
                          top=(thick if top else thin), bottom=thin)
        return c

    def cell(ws, r, c, val, bg=WHITE, fg='000000', bold=False,
             left=False, size=10):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='left' if left else 'center',
                                  vertical='center')
        return brd(cl)

    ri = 1

    # Title
    ws.row_dimensions[ri].height = 22
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci,
            'Table T2.  Temporal trajectory of |CC| by year  --  all significant metrics (n_sig_years >= 1)'
            if ci == 1 else '')
        cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
        cl.fill = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    ws.row_dimensions[ri].height = 13
    for ci in range(1, NCOLS+1):
        cl = ws.cell(ri, ci,
            'n sig = count with |CC| > 0.30 for asedx_p_{year} (all-ages)  |  '
            'Mean/Min/Max |CC| computed over ALL significant metrics with non-NaN CC '
            '(matches fig2_temporal_trajectory.png)'
            if ci == 1 else '')
        cl.font = Font(name='Arial', italic=True, color='FFFFFF', size=9)
        cl.fill = PatternFill('solid', start_color=MID)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    # Spacer
    ws.row_dimensions[ri].height = 4
    for ci in range(1, NCOLS+1):
        ws.cell(ri, ci).fill = PatternFill('solid', start_color='F0F0F0')
    ri += 1

    # Column header row
    ws.row_dimensions[ri].height = 22
    hdrs = ['Year', 'n sig', 'Mean |CC|', 'Min |CC|', 'Max |CC|']
    for ci, h in enumerate(hdrs, 1):
        cl = ws.cell(ri, ci, h)
        cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cl.fill = PatternFill('solid', start_color=NAVY)
        cl.alignment = Alignment(horizontal='center', vertical='center')
        brd(cl)
    ri += 1

    ws.freeze_panes = 'B5'

    # Data rows -- pooled over all significant metrics, using |CC|
    metrics = [m for m in sig1_df['metric'] if m in cc_df.index]

    for yr in YEARS:
        cc_col = f'asedx_p_{yr}'
        lp_col = f'LP_asedx_p_{yr}'

        if cc_col in cc_df.columns:
            cc_vals = cc_df.loc[metrics, cc_col].dropna().abs()
            if SIG_MODE == 'cc':
                n_sig = int((cc_vals > CC_SIG).sum())
            else:
                lp_vals = cc_df.loc[metrics, lp_col].dropna() if lp_col in cc_df.columns else pd.Series(dtype=float)
                n_sig   = int((lp_vals <= -13).sum())
        else:
            cc_vals = pd.Series(dtype=float)
            n_sig = 0

        mean_v = round(float(cc_vals.mean()), 2) if len(cc_vals) else None
        min_v  = round(float(cc_vals.min()),  2) if len(cc_vals) else None
        max_v  = round(float(cc_vals.max()),  2) if len(cc_vals) else None

        def fmt(v):
            if v is None: return '---'
            return f'{v:.2f}'

        bg = LIGHT if ri % 2 == 1 else WHITE
        ws.row_dimensions[ri].height = 20
        cell(ws, ri, 1, yr,          bg=WHITE, bold=True, size=11)
        cell(ws, ri, 2, n_sig,       bg=bg,    bold=True)
        cell(ws, ri, 3, fmt(mean_v), bg=bg,    bold=True)
        cell(ws, ri, 4, fmt(min_v),  bg=bg)
        cell(ws, ri, 5, fmt(max_v),  bg=bg)
        ri += 1


def write_table_xde_sem_overlap(wb, all_df, z_med_path, medoid_list_path,
                                 reps_path, sem_sc_path, sem_names_path, out_dir):
    """XDE vs SEM super-cluster overlap table + heatmap figure."""
    import json, numpy as np
    from scipy.cluster.hierarchy import fcluster
    from sklearn.metrics import adjusted_rand_score
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors

    print("Writing Table_XDE_SEM_Overlap + figure...")

    # -- Load data ----------------------------------------------------------
    Z_med       = np.load(z_med_path)
    medoid_list = json.load(open(medoid_list_path))
    reps        = pd.read_csv(reps_path)
    sem_assign  = pd.read_csv(sem_sc_path)
    sem_names   = pd.read_csv(sem_names_path)

    ward_col = [c for c in reps.columns if c.startswith('Ward') or c=='xde_cluster'][0]                if any(c.startswith('Ward') or c=='xde_cluster' for c in reps.columns)                else reps.columns[0]
    # robust: find column that has cluster IDs 1-100
    for c in reps.columns:
        if reps[c].between(1,100).all() and reps[c].nunique() >= 90:
            ward_col = c; break

    medoid_to_ward = dict(zip(reps['medoid'], reps[ward_col]))
    n = len(medoid_list)

    # XDE SC12: fcluster cut
    n_sc = int(all_df['super_cluster_id'].dropna().astype(int).nunique())
    xde_sc = fcluster(Z_med, t=n_sc, criterion='maxclust')
    ward_to_xde = {medoid_to_ward.get(medoid_list[i], 0): int(xde_sc[i]) for i in range(n)}

    # SEM SC12
    _sem_col = 'sem100' if 'sem100' in sem_assign.columns else 'Ward100'
    ward_to_sem = dict(zip(sem_assign[_sem_col].astype(int), sem_assign['sc_id'].astype(int)))
    sem_name    = dict(zip(sem_names['sc_id'].astype(int), sem_names['sc_name']))

    xde_ids = sorted(set(ward_to_xde.values()))
    sem_ids = sorted(set(ward_to_sem.values()))

    # Overlap matrix
    mat = pd.DataFrame(0, index=xde_ids, columns=sem_ids)
    for w in reps[ward_col].dropna().astype(int):
        xsc = ward_to_xde.get(int(w), 0)
        ssc = ward_to_sem.get(int(w), 0)
        if xsc and ssc:
            mat.loc[xsc, ssc] += 1

    # Per-XDE-SC stats
    rows = []
    for xsc in xde_ids:
        row   = mat.loc[xsc]
        n_tot = int(row.sum())
        dom   = int(row.idxmax())
        pur   = row.max() / n_tot * 100 if n_tot else 0
        n_sem = int((row > 0).sum())
        rows.append({'xde_sc': xsc, 'n_clusters': n_tot,
                     'dom_sem_sc': dom, 'dom_sem_name': sem_name.get(dom,''),
                     'purity_pct': round(pur, 1), 'n_sem_scs': n_sem})
    stats_df = pd.DataFrame(rows)

    # ARI
    xde_labels = [ward_to_xde.get(int(w), 0) for w in reps[ward_col].dropna().astype(int)]
    sem_labels  = [ward_to_sem.get(int(w), 0) for w in reps[ward_col].dropna().astype(int)]
    ari = adjusted_rand_score(xde_labels, sem_labels)

    # -- Excel table --------------------------------------------------------
    ws = wb.create_sheet('XDE_SEM_Overlap')
    NAVY = '1F3864'; MID = '2E75B6'; WHITE = 'FFFFFF'
    DKRED = '8B0000'; DKBLUE = '1F4E79'
    thin  = Side(style='thin',   color='CCCCCC')
    thick = Side(style='medium', color='888888')

    def brd(c):
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin); return c

    def hcell(ws, r, c, val, bg=NAVY, fg='FFFFFF', bold=True, left=False, size=9):
        cl = ws.cell(r, c, val)
        cl.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='left' if left else 'center', vertical='center')
        return brd(cl)

    def dcell(ws, r, c, val, bg=WHITE, fg='000000', bold=False, left=False,
              size=9, wrap=False):
        cl = ws.cell(r, c, val)
        cl.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal='left' if left else 'center',
                                  vertical='center', wrap_text=(wrap or left))
        return brd(cl)

    # Column widths
    OVERLAP_NCOLS = 13 + 5  # 12 SEM cols + XDE label + stats
    ws.column_dimensions['A'].width = 9
    for ci in range(2, 14):
        ws.column_dimensions[get_column_letter(ci)].width = 5
    for ci, w in zip(range(14,19), [10, 8, 50, 8, 8]):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ri = 1
    # Title
    ws.row_dimensions[ri].height = 20
    ncols = 18
    for ci in range(1, ncols+1):
        hcell(ws, ri, ci,
              f'XDE vs SEM Super-Cluster Overlap  |  '
              f'Adjusted Rand Index = {ari:.3f}  |  '
              f'(0=random, 1=perfect)  |  n=100 XDE Ward clusters'
              if ci==1 else '', bg=NAVY, size=11)
    ri += 1

    # Sub-header
    ws.row_dimensions[ri].height = 14
    for ci in range(1, ncols+1):
        hcell(ws, ri, ci,
              'Rows = XDE-SC (Ward on CC excess-death matrix, k=N cut)  |  '
              f'Cols = SEM-SC (Ward on label embeddings, k={n_sc})  |  '
              'Cell = n Ward100 clusters in intersection'
              if ci==1 else '', bg=MID, size=9)
    ri += 1

    ws.row_dimensions[ri].height = 4; ri += 1  # spacer

    # SEM SC name header row (rotated)
    ws.row_dimensions[ri].height = 80
    hcell(ws, ri, 1, 'XDE SC', bg='FF37474F')
    for j, ssc in enumerate(sem_ids, 2):
        nm = sem_name.get(ssc, 'SEM-SC%02d' % ssc)
        cl = ws.cell(ri, j, 'SEM-SC%02d\n%s' % (ssc, nm))
        cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=8)
        cl.fill = PatternFill('solid', start_color=DKBLUE)
        cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                  text_rotation=60, wrap_text=False)
        brd(cl)
    for ci, h in enumerate(['n XDE clusters','Dom SEM SC','Dom SEM name','Purity %','n SEM SCs'], 14):
        hcell(ws, ri, ci, h, bg='FF37474F', size=8)
    ri += 1

    ws.freeze_panes = 'B6'

    # Data rows
    # colour scale: white -> deep blue by count
    max_val = int(mat.values.max())
    def count_bg(v):
        if v == 0: return 'FFFFFF'
        intensity = v / max_val
        r_c = int(255 + (31  - 255) * intensity)
        g_c = int(255 + (78  - 255) * intensity)
        b_c = int(255 + (121 - 255) * intensity)
        return f'{r_c:02X}{g_c:02X}{b_c:02X}'

    purity_bg = lambda p: ('FFEBEE' if p < 40 else 'FFF9C4' if p < 60
                            else 'E8F5E9' if p < 80 else 'C8E6C9')

    for xsc in xde_ids:
        ws.row_dimensions[ri].height = 16
        hcell(ws, ri, 1, f'XDE-SC{xsc:02d}', bg='FF1F3864', size=9)
        row_vals = mat.loc[xsc]
        for j, ssc in enumerate(sem_ids, 2):
            v = int(row_vals[ssc])
            bg = count_bg(v)
            fg = 'FFFFFF' if v > max_val*0.6 else '000000'
            dcell(ws, ri, j, v if v > 0 else '', bg=bg, fg=fg,
                  bold=(v == int(row_vals.max()) and v > 0), size=9)
        # Stats columns
        st = stats_df[stats_df['xde_sc']==xsc].iloc[0]
        dcell(ws, ri, 14, int(st['n_clusters']), bold=True, size=9)
        dcell(ws, ri, 15, f"SEM-SC{int(st['dom_sem_sc']):02d}", bg='EBF5FB', size=9)
        dcell(ws, ri, 16, st['dom_sem_name'], bg='EBF5FB', left=True, wrap=True, size=8)
        pur = float(st['purity_pct'])
        dcell(ws, ri, 17, f"{pur:.0f}%", bg=purity_bg(pur),
              bold=(pur>=75), fg=(DKRED if pur<40 else '000000'), size=9)
        dcell(ws, ri, 18, int(st['n_sem_scs']), size=9)
        ri += 1

    # Summary row
    ws.row_dimensions[ri].height = 4; ri += 1
    ws.row_dimensions[ri].height = 16
    hcell(ws, ri, 1, 'ARI', bg='FF37474F')
    dcell(ws, ri, 2, f'{ari:.3f}', bold=True, fg=DKRED, size=10)
    for ci in range(3, ncols+1):
        dcell(ws, ri, ci, '')
    ri += 1

    # -- Heatmap figure ----------------------------------------------------
    fig, axes = plt.subplots(1, 2, figsize=(18, 7),
                              gridspec_kw={'width_ratios': [2, 1]})

    # Left: heatmap
    ax = axes[0]
    mat_arr = mat.values.astype(float)
    mat_pct = mat_arr / mat_arr.sum(axis=1, keepdims=True) * 100

    im = ax.imshow(mat_pct, aspect='auto', cmap='YlOrRd',
                   vmin=0, vmax=100, interpolation='none')
    for i in range(len(xde_ids)):
        for j in range(len(sem_ids)):
            v = int(mat_arr[i, j])
            p = mat_pct[i, j]
            if v > 0:
                ax.text(j, i, '%d\n%.0f%%' % (v, p), ha='center', va='center',
                        fontsize=6.5,
                        color='white' if p > 55 else 'black')

    ax.set_xticks(range(len(sem_ids)))
    ax.set_xticklabels(['SEM-SC%02d\n%s' % (s, sem_name.get(s,''))
                        for s in sem_ids],
                        rotation=45, ha='right', fontsize=7)
    ax.set_yticks(range(len(xde_ids)))
    ax.set_yticklabels([f'XDE-SC{x:02d}' for x in xde_ids], fontsize=8)
    ax.set_xlabel('SEM Super-Cluster (Ward on label embeddings)', fontsize=10)
    ax.set_ylabel('XDE Super-Cluster (Ward on CC excess-death matrix)', fontsize=10)
    ax.set_title(f'XDE vs SEM Super-Cluster Overlap\n'
                 f'(cell = n Ward100 clusters; colour = row %; '
                 f'ARI={ari:.3f})',
                 fontsize=11, fontweight='bold')
    cb = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02)
    cb.set_label('% of XDE-SC clusters', fontsize=9)

    # Right: purity bar chart
    ax2 = axes[1]
    purities = stats_df['purity_pct'].values
    colors = ['#C8E6C9' if p>=75 else '#FFF9C4' if p>=60
              else '#FFCCBC' if p>=40 else '#FFCDD2' for p in purities]
    bars = ax2.barh(range(len(xde_ids)), purities, color=colors,
                    edgecolor='grey', linewidth=0.5)
    ax2.axvline(75, color='green',  lw=1.5, ls='--', alpha=0.7, label='75% threshold')
    ax2.axvline(50, color='orange', lw=1.0, ls='--', alpha=0.6, label='50% threshold')
    for i, (p, b) in enumerate(zip(purities, bars)):
        ax2.text(p+1, i, f'{p:.0f}%', va='center', fontsize=8)
    ax2.set_yticks(range(len(xde_ids)))
    ax2.set_yticklabels([f'XDE-SC{x:02d}' for x in xde_ids], fontsize=8)
    # Annotate dominant SEM SC name on each bar
    for i, xsc in enumerate(xde_ids):
        st = stats_df[stats_df['xde_sc']==xsc].iloc[0]
        nm = sem_name.get(int(st['dom_sem_sc']), '')
        ax2.text(1, i, 'SEM-SC%02d: %s' % (int(st['dom_sem_sc']), nm),
                 va='center', fontsize=6.5, color='#333333')
    ax2.set_xlabel('Purity (% clusters from dominant SEM-SC)', fontsize=10)
    ax2.set_title('XDE Super-Cluster Purity\n(dominant SEM-SC match)',
                  fontsize=11, fontweight='bold')
    ax2.set_xlim(0, 110)
    ax2.legend(fontsize=8, loc='lower right')
    ax2.grid(True, axis='x', alpha=0.3)

    fig.suptitle(
        ('XDE vs SEM Super-Cluster Comparison  |  ARI=%.3f  |  '
         'Near-zero ARI confirms the two systems are orthogonal:\n') % ari +
        f'XDE-SCs group metrics by mortality correlation structure; '
        f'SEM-SCs group metrics by thematic content',
        fontsize=10, fontweight='bold'
    )
    fig.tight_layout(rect=[0, 0, 1, 0.92])

    fig_path = os.path.join(out_dir, 'fig_xde_sem_overlap.png')
    save_fig(fig, fig_path, dpi=200)
    plt.close()
    print(f'Saved: {fig_path}')
    print('Saved ' + os.path.relpath(fig_path), file=sys.stderr, flush=True)


if __name__ == '__main__':
    args = parse_args()
    SIG_MODE   = args.sig_mode
    CC_SIG     = args.cc_sig
    CC_VERYSIG = args.cc_verysig
    if not args.master_xlsx and not (args.my_file and args.all_file):
        print("ERROR: provide --master-xlsx or both --my-file and --all-file")
        sys.exit(1)

    sig1_df, sig1_df, all_df, cc_raw, clusters = load_data(args)

    wb = Workbook()
    wb.remove(wb.active)

    write_table_1_2(wb, sig1_df)
    write_table_temporal(wb, sig1_df, cc_raw)
    if all(getattr(args, a, None) for a in
           ['z_med','medoid_list','reps','sem_sc','sem_names']):
        out_dir = os.path.dirname(os.path.abspath(args.output))
        write_table_xde_sem_overlap(wb, all_df,
            args.z_med, args.medoid_list, args.reps,
            args.sem_sc, args.sem_names, out_dir)
    else:
        print('Skipping XDE_SEM_Overlap table (provide --z-med --medoid-list --reps --sem-sc --sem-names)')
    write_table_3(wb, sig1_df, all_df, clusters, cc_df=cc_raw)
    write_table_4(wb, sig1_df, cc_raw, clusters)
    write_table_5(wb, sig1_df, all_df, cc_df=cc_raw)
    write_table_race(wb, sig1_df)
    write_table_sc_cluster_si(wb, sig1_df, all_df)
    write_table_si(wb, all_df, sig1_df, sig1_df=sig1_df)

    # Uniform style pass: shift all sheets to start at B2, Times New Roman
    # 11pt black on white, no fills, no freeze panes (matches master xlsx).
    print("Applying uniform black-on-white Times Roman 11pt style to all tabs...")
    clean_style_workbook(wb)

    wb.save(args.output)
    _tee(args.output)
    try:
        _r_args_output = os.path.relpath(args.output)
    except ValueError:
        _r_args_output = args.output
    print('Saved ' + _r_args_output, file=sys.stderr, flush=True)
    print(f"\nSaved: {args.output}")
    print(f"Tabs: {wb.sheetnames}")
