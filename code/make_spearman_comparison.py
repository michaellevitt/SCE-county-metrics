#!/usr/bin/env python3
"""
make_spearman_comparison.py
---------------------------
Build the Pearson-vs-Spearman comparison for the paper's "Sensitivity analysis on
the correlation method" (Table S8 + supplement workbook).

Inputs (run these first):
  full_w1.0/metric_x_death_cc_1.0_0.csv            Pearson death-CC   (committed / from run_all.sh)
  spearman_results/metric_x_death_cc_1.0_0.csv     Spearman death-CC  (from run_spearman.sh)
  figures_2745/Metric_Super-Cluster_Cluster.csv    variable -> super-cluster map (from run_all.sh)
  data/BEN_MERGED_MEASURES_explain_extended_2745.csv

Statistic (matches Table 1): per variable, maximum |CC| over the ALL-AGE years
2020-2024, on the 2,727 valid-CC variables. Moderate = |CC|>0.30, strong = |CC|>0.45.

Output:
  SCE_Pearson_vs_Spearman_supplement.xlsx   (Summary, TableS8, TopVariables)
Run from the repo root:  python3 code/make_spearman_comparison.py
"""
import os, sys
import numpy as np, pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PEAR = "full_w1.0/metric_x_death_cc_1.0_0.csv"
SPEAR = "spearman_results/metric_x_death_cc_1.0_0.csv"
MAP = "figures_2745/Metric_Super-Cluster_Cluster.csv"
EXPLAIN = "data/BEN_MERGED_MEASURES_explain_extended_2745.csv"
OUT = "SCE_Pearson_vs_Spearman_supplement.xlsx"


def maxabs(path):
    df = pd.read_csv(path)
    cols = [f"asedx_p_{y}" for y in range(2020, 2025) if f"asedx_p_{y}" in df.columns]
    A = df[cols].apply(pd.to_numeric, errors="coerce").abs().values
    allnan = np.all(np.isnan(A), axis=1)
    mx = np.nanmax(np.where(np.isnan(A), -np.inf, A), axis=1); mx[allnan] = np.nan
    return pd.Series(mx, index=df[df.columns[0]].astype(str))


def main():
    for f in (PEAR, SPEAR, MAP):
        if not os.path.exists(f):
            sys.exit(f"MISSING: {f}\n  (run run_all.sh for Pearson, then run_spearman.sh)")
    pear, spear = maxabs(PEAR), maxabs(SPEAR)
    m = pd.read_csv(MAP, dtype=str); m["metric"] = m["metric"].astype(str)
    m["P"] = m["metric"].map(pear); m["S"] = m["metric"].map(spear)
    v = m.dropna(subset=["P", "S"])                      # 2,727 valid-CC basis
    ex = {}
    if os.path.exists(EXPLAIN):
        ex = pd.read_csv(EXPLAIN).set_index("metric")["explain"].to_dict()

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    HDR = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="305496")
    def hdr(ws, n):
        for c in range(1, n + 1):
            ws.cell(1, c).font = HDR; ws.cell(1, c).fill = HF
            ws.cell(1, c).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"

    # ---- Summary ----
    n = len(v)
    def pct(s, t): return round(100 * (s > t).mean(), 1)
    ws = wb.create_sheet("Summary")
    for r in [
        ["SCE Pearson vs Spearman comparison (weighted; max |CC| over all-age 2020-2024; N=%d)" % n, ""],
        ["", ""],
        ["Variables > 0.30 (moderate)  Pearson", f"{pct(v.P,.3)}%  ({int((v.P>.3).sum())})"],
        ["Variables > 0.30 (moderate)  Spearman", f"{pct(v.S,.3)}%  ({int((v.S>.3).sum())})"],
        ["Variables > 0.45 (strong)    Pearson", f"{pct(v.P,.45)}%  ({int((v.P>.45).sum())})"],
        ["Variables > 0.45 (strong)    Spearman", f"{pct(v.S,.45)}%  ({int((v.S>.45).sum())})"],
        ["Strongest |CC|  Pearson / Spearman", f"{v.P.max():.2f} / {v.S.max():.2f}"],
        ["Corr of per-variable max|CC| (P vs S)", round(float(v.P.corr(v.S)), 3)],
    ]:
        ws.append(r)
    ws["A1"].font = Font(bold=True); ws.column_dimensions["A"].width = 46; ws.column_dimensions["B"].width = 22

    # ---- TableS8: per super-cluster ----
    ws = wb.create_sheet("TableS8")
    ws.append(["Super-cluster", "N", "% mod (P)", "% mod (S)", "strong n (P)", "strong n (S)", "max |CC| (P)", "max |CC| (S)"])
    for sid, d in sorted(v.groupby("super_cluster_id"), key=lambda x: int(x[0])):
        ws.append([f"{sid}. {d['super_cluster_name'].iloc[0]}", len(d),
                   pct(d.P, .3), pct(d.S, .3),
                   int((d.P > .45).sum()), int((d.S > .45).sum()),
                   round(d.P.max(), 2), round(d.S.max(), 2)])
    hdr(ws, 8)
    for i, w in enumerate([34, 6, 10, 10, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- TopVariables: union of top-30 by each method ----
    topP = set(v.sort_values("P", ascending=False).head(30).metric)
    topS = set(v.sort_values("S", ascending=False).head(30).metric)
    u = v[v.metric.isin(topP | topS)].sort_values("P", ascending=False)
    ws = wb.create_sheet("TopVariables")
    ws.append(["Variable", "Description", "max|CC| (P)", "max|CC| (S)", "in P top30", "in S top30"])
    for _, r in u.iterrows():
        ws.append([r.metric, str(ex.get(r.metric, ""))[:60], round(r.P, 3), round(r.S, 3),
                   "✓" if r.metric in topP else "", "✓" if r.metric in topS else ""])
    hdr(ws, 6)
    for i, w in enumerate([12, 48, 12, 12, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  moderate (>0.30): Pearson {pct(v.P,.3)}%  ->  Spearman {pct(v.S,.3)}%")
    print(f"  strong (>0.45):   Pearson {int((v.P>.45).sum())}  ->  Spearman {int((v.S>.45).sum())}")
    print(f"  strongest |CC|:   {v.P.max():.2f}  ->  {v.S.max():.2f}")


if __name__ == "__main__":
    main()
