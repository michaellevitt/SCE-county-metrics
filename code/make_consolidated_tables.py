#!/usr/bin/env python3
"""
make_consolidated_tables.py  --  Step L of run_standard_k120_w1.0.sh

Consolidate the ESSENTIAL tabs from the three pipeline workbooks into a single
"paper tables" workbook, with an Index sheet documenting what each retained tab
does for the manuscript. All thresholds are |CC| effect-size bands
(moderate |CC| > 0.30, strong |CC| > 0.45), not statistical significance -- LP is retired.

Inputs (defaults match the standard w1.0 run):
  --master  master_sem_clusters_clean2_k120_w1.0.xlsx
  --extra   extra_tables_clean2_k120_w1.0.xlsx
  --best    sem_best_lp_clean2_k120_w1.0.xlsx     (sheet renamed SEM_Best_CC)
  --output  SCE_paper_tables_consolidated_clean2_k120_w1.0.xlsx

Dropped on purpose (redundant / internal / secondary -- see Index notes):
  XDE cross-tabs Table_3/Table_4 (secondary clustering, not used in paper),
  Table_5/Table_1_and_2/T6_Top30/Multi_Year/T4_Clusters (duplicate content),
  T1_Overview/T7_Patterns/Table_SI/Table_SC_Cluster_SI (QC / SI-pattern dists),
  T5_Domains (descriptive cross-cutting tag, not a hierarchy level).
"""
import argparse, re, sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _mv(s):
    """Whole-word metric->variable for display text (case-preserving)."""
    if not isinstance(s, str):
        return s
    mp = {"metrics": "variables", "metric": "variable", "Metrics": "Variables", "Metric": "Variable"}
    return re.sub(r"\b(Metrics|metrics|Metric|metric)\b", lambda m: mp[m.group(0)], s)


def build_spec(master, extra, best):
    """(new tab name, source path, source sheet, what-it-is, role-in-paper)."""
    return [
        ("Master", master, "Master",
         "Full per-metric dataset (2,745 metrics): super-cluster, k=120 cluster, metric code, "
         "plain-English description, max signed CC, n moderate-association years (|CC|>0.3), n strong "
         "years (|CC|>0.45), SI temporal pattern, sign, age dominance, extensive/intensive class, data year.",
         "THE master Supplementary Data table -- every other tab and every manuscript table derives from it."),
        ("SuperClusters", master, "T3_SuperCluster",
         "Per super-cluster: total metrics, n & % with |CC|>0.3, mean |CC|, and age dominance (% GE65 / % LT65).",
         "Main TABLE 1 (super-cluster association summary). The %GE65/%LT65 columns also underlie the "
         "Age-band divergence section and Table S8."),
        ("Clusters", master, "Clusters",
         "Per k=120 semantic cluster: super-cluster, cluster label, # metrics, # & % moderate (|CC|>0.3), mean |CC|.",
         "Basis for TABLE S1 (super-cluster sizes) and TABLE S6 (super-cluster -> constituent cluster names)."),
        ("Moderate_Variables", master, "Significant",
         "Every metric reaching |CC|>0.3 in any 2020-2024 year, with cluster, signed CC, n moderate-association years, SI pattern.",
         "Source for TABLE S7 (top-six per super-cluster) and the 'signal is sparse and concentrated' counts."),
        ("Best_Per_Cluster", best, "SEM_Best_CC",
         "One row per k=120 cluster: its strongest metric, ranked by max |CC| (signed CC, n moderate-association years, SI).",
         "Basis for the per-cluster Supplementary Data file (top_metric_per_cluster) and the representatives in TABLE S7."),
        ("Race_Ethnicity", extra, "Table_Race",
         "Race/ethnicity-group findings: # moderate-association metrics, mean |CC|, strongest +/- metric, interpretation, with caveats.",
         "Supporting evidence for the Discussion race paragraph (race = first-wave face of the SES axis) "
         "and the Limitations AHRF Black/AA caveat."),
    ]

INDEX_TITLE = ("Consolidated SCE paper tables -- standard run (2019 metrics | w1.0 population-weighting | "
               "k=120 clusters | |CC| effect-size bands: moderate >0.30, strong >0.45)")

# Columns omitted from every copied tab (matched exactly on header, case-insensitive,
# so dominant_age / first_sig_year / etc. are NOT affected): LP magnitude and the
# SI_near / sign / age helper columns are dropped as not paper-relevant.
DROP_COLS = {"sum_abs_lp", "si_near", "sign", "age"}

# Two alternating fills for the Moderate_by_SI tab (one band per shared SI value).
SI_BAND_COLORS = ["D9E1F2", "E2EFDA"]  # light blue / light green


# Tokens forced to all-caps in headers; everything else is Title-Cased word by word.
HEADER_ABBR = {"LP", "CC", "N", "IE", "EI", "ID", "SI", "SC", "US", "AHRF", "GE65", "LT65"}
# Whole-token rewrites (consistency / readability fixes).
HEADER_SPECIAL = {"pct": "%", "verysig": "Strong", "sig": "Moderate",
                  "metric": "Variable", "metrics": "Variables"}


def _cap_token(tok):
    if tok == "":
        return tok
    if tok.lower() in HEADER_SPECIAL:
        return HEADER_SPECIAL[tok.lower()]
    core = re.sub(r"[^A-Za-z0-9]", "", tok)
    if core and core.upper() in HEADER_ABBR:           # e.g. id, cc, |CC|, GE65
        return tok.replace(core, core.upper())
    if tok.isalpha():
        return tok[:1].upper() + tok[1:].lower()
    def _fix(m):                                       # mixed token (digits/symbols/newline)
        w = m.group(0)
        return w.upper() if w.upper() in HEADER_ABBR else (w[:1].upper() + w[1:].lower())
    return re.sub(r"[A-Za-z]+", _fix, tok)


def cap_header(h):
    """Title-case each word; abbreviations (LP, CC, N, IE/EI, ID, SI, SC, ...) all-caps."""
    if not isinstance(h, str):
        return h
    return " ".join(_cap_token(t) for t in h.split(" "))


def finalize_headers(ws, hdr_row):
    """Rename Ward100 -> Cluster ID, '_'->' ', consistent capitalization, rotate 90 deg."""
    rot = Alignment(textRotation=90, vertical="bottom", horizontal="center")
    for c in ws[hdr_row]:
        if isinstance(c.value, str) and c.value.strip():
            v = c.value.strip()
            if v == "Ward100":
                v = "Cluster ID"
            elif v == "SC":                       # unify with 'Super Cluster ID' used in other tabs
                v = "Super Cluster ID"
            hv = cap_header(v.replace("_", " ").replace("-", " "))
            hv = hv.replace("Abs CC", "|CC|")     # absolute-mean uses |CC| bars (matches 'Mean |CC|');
            c.value = hv                          # signed mean stays 'Cluster Mean CC' (no bars)
        c.alignment = rot
    ws.row_dimensions[hdr_row].height = 150


def band_by_col(ws, hdr_row, col_idx, colors):
    """Alternate fill for each run of equal values in column col_idx (1-based), rows below hdr_row."""
    prev = object(); band = -1
    for i in range(hdr_row + 1, ws.max_row + 1):
        v = ws.cell(row=i, column=col_idx).value
        if v in (None, ""):
            continue
        if v != prev:
            band += 1; prev = v
        fill = PatternFill("solid", fgColor=colors[band % 2])
        for c in ws[i]:
            c.fill = fill


# Column-formatting rules (matched on normalized header: newlines->space, lowercased).
_NUM_BLANKS = {"", "-", "---", "—", "nan", "na", "none"}
_FMT_4DP = {"cluster mean cc", "cluster mean |cc|", "max signed cc", "mean |cc|"}   # 4 decimals
_FMT_1DP = {"cluster % sig", "min lp", "% sig", "% ge65", "% lt65"}                 # 1 decimal
_CENTER_TEXT = {"ei class", "best death"}                                          # center though non-numeric


def _norm_hdr(h):
    return str(h).replace("\n", " ").strip().lower()


def _is_blank(v):
    return v is None or (isinstance(v, str) and v.strip().lower() in _NUM_BLANKS)


def _as_num(v):
    """Return float(v) if v is a number or a numeric string, else None."""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def format_sheet(ws, hdr_row):
    """Center purely-numeric columns (+ EI Class / Best Death); apply decimal formats.
    Numeric strings (e.g. text-stored '0.434') count as numeric and are coerced for the formats."""
    center = Alignment(horizontal="center", vertical="center")
    for j in range(1, ws.max_column + 1):
        hn = _norm_hdr(ws.cell(row=hdr_row, column=j).value)
        cells = [ws.cell(row=i, column=j) for i in range(hdr_row + 1, ws.max_row + 1)]
        nonblank = [c.value for c in cells if not _is_blank(c.value)]
        numeric = bool(nonblank) and all(_as_num(v) is not None for v in nonblank)
        if numeric or hn in _CENTER_TEXT:
            for c in cells:
                c.alignment = center
        fmt = "0.0000" if hn in _FMT_4DP else ("0.0" if hn in _FMT_1DP else None)
        if fmt:
            for c in cells:
                if _is_blank(c.value):
                    continue
                n = _as_num(c.value)
                if n is not None:
                    c.value = n              # coerce numeric strings so the number format renders
                    c.number_format = fmt


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--master", default="master_sem_clusters_clean2_k120_w1.0.xlsx")
    ap.add_argument("--extra",  default="extra_tables_clean2_k120_w1.0.xlsx")
    ap.add_argument("--best",   default="sem_best_lp_clean2_k120_w1.0.xlsx")
    ap.add_argument("--output", default="SCE_paper_tables_consolidated_clean2_k120_w1.0.xlsx")
    args = ap.parse_args()

    spec = build_spec(args.master, args.extra, args.best)
    src = {}
    for f in {args.master, args.extra, args.best}:
        src[f] = openpyxl.load_workbook(f, read_only=True, data_only=True)

    TNR  = Font(name="Times New Roman", size=11)
    TNRB = Font(name="Times New Roman", size=11, bold=True)
    HEAD = PatternFill("solid", fgColor="D5E8F0")

    out = Workbook(); out.remove(out.active)

    # ---- Index ----
    idx = out.create_sheet("Index")
    idx.append([_mv(INDEX_TITLE)]); idx.append([])
    idx.append(["Tab", "Source workbook", "What it contains", "Role in the paper"])
    for name, f, sheet, what, role in spec:
        idx.append([name, f, _mv(what), _mv(role)])
    for r in idx.iter_rows():
        for c in r:
            c.font = TNR; c.alignment = Alignment(vertical="top", wrap_text=True)
    idx["A1"].font = TNRB
    for c in idx[3]:
        c.font = TNRB; c.fill = HEAD
    idx.column_dimensions["A"].width = 20; idx.column_dimensions["B"].width = 42
    idx.column_dimensions["C"].width = 62; idx.column_dimensions["D"].width = 62
    idx.freeze_panes = "A4"

    # ---- copy each essential sheet (values) ----
    for name, f, sheet, what, role in spec:
        if sheet not in src[f].sheetnames:
            print(f"WARNING: sheet '{sheet}' not in {f}; skipping", file=sys.stderr); continue
        ws_src = src[f][sheet]; ws = out.create_sheet(name); maxw = {}
        rows_src = list(ws_src.iter_rows(values_only=True))
        # locate source header row and the column indices to drop (exact header match)
        hsrc = next((i for i, r in enumerate(rows_src)
                     if sum(1 for c in r if c not in (None, "")) >= 3), None)
        drop = set()
        if hsrc is not None:
            for j, v in enumerate(rows_src[hsrc]):
                if isinstance(v, str) and v.strip().lower() in DROP_COLS:
                    drop.add(j)
        for r in rows_src:
            ws.append([_mv(v) for j, v in enumerate(r) if j not in drop])
        rows = list(ws.iter_rows()); hdr_idx = None
        for i, r in enumerate(rows):
            if sum(1 for c in r if c.value not in (None, "")) >= 3:
                hdr_idx = i; break
        for i, r in enumerate(rows):
            is_hdr = (hdr_idx is not None and i == hdr_idx)
            for c in r:
                c.font = TNRB if is_hdr else TNR
                if is_hdr:
                    c.fill = HEAD
                if c.value is not None and not is_hdr:   # width from DATA only (rotated headers don't widen)
                    maxw[c.column_letter] = min(max(maxw.get(c.column_letter, 6), len(str(c.value)) + 2), 55)
        for col, w in maxw.items():
            ws.column_dimensions[col].width = w
        if hdr_idx is not None:
            hdr_row = hdr_idx + 1
            ward_col = next((c.column for c in ws[hdr_row]
                             if isinstance(c.value, str) and c.value.strip() == "Ward100"), None)
            if name in ("Master", "Moderate_Variables") and ward_col is not None:
                band_by_col(ws, hdr_row, ward_col, SI_BAND_COLORS)   # band by Cluster ID
            finalize_headers(ws, hdr_row)                            # Ward100->Cluster ID, _->space, rotate 90
            format_sheet(ws, hdr_row)                                # center numeric cols + decimal formats
            ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1).coordinate

    # ---- derived tab: Moderate_Variables sorted by SI (V-containing first, then Z-A), banded by SI group ----
    if "Moderate_Variables" in out.sheetnames:
        base = out["Moderate_Variables"]
        rows = list(base.iter_rows(values_only=True))
        h = next((i for i, r in enumerate(rows)
                  if sum(1 for c in r if c not in (None, "")) >= 3), None)
        if h is not None:
            hdr = rows[h]
            si_j = next((j for j, v in enumerate(hdr)
                         if isinstance(v, str) and v.strip() == "SI"), None)
            pre = [list(r) for r in rows[:h + 1]]
            data = [list(r) for r in rows[h + 1:] if any(c not in (None, "") for c in r)]
            if si_j is not None:
                data.sort(key=lambda r: str(r[si_j] or ""), reverse=True)            # SI Z-A
                data.sort(key=lambda r: 0 if "V" in str(r[si_j] or "") else 1)        # V-containing first (stable)
            ws = out.create_sheet("Moderate_by_SI")
            for r in pre + data:
                ws.append(r)
            for c in ws[h + 1]:                       # header row
                c.font = TNRB; c.fill = HEAD
            prev = object(); band = -1; maxw = {}
            for i in range(h + 2, ws.max_row + 1):    # data rows -> band by SI group
                si = ws.cell(row=i, column=si_j + 1).value if si_j is not None else None
                if si != prev:
                    band += 1; prev = si
                fill = PatternFill("solid", fgColor=SI_BAND_COLORS[band % 2])
                for c in ws[i]:
                    c.font = TNR; c.fill = fill
            for r in ws.iter_rows():
                for c in r:
                    if c.value is not None:
                        maxw[c.column_letter] = min(max(maxw.get(c.column_letter, 8), len(str(c.value)) + 2), 55)
            for col, w in maxw.items():
                ws.column_dimensions[col].width = w
            finalize_headers(ws, h + 1)              # _->space, rotate 90 (SI banding already applied)
            format_sheet(ws, h + 1)                  # center numeric cols + decimal formats
            ws.freeze_panes = ws.cell(row=h + 2, column=1).coordinate
            ix = out["Index"]
            ix.append(["Moderate_by_SI", "(derived from Moderate_Variables)",
                       "The same moderate-association rows, re-sorted by SI temporal pattern (V-containing patterns "
                       "first, then Z-A) with alternating fill banding per SI group.",
                       "Reading aid for the temporal-pattern (SI) structure; not a separate analysis."])
            for c in ix[ix.max_row]:
                c.font = TNR; c.alignment = Alignment(vertical="top", wrap_text=True)

    out.save(args.output)
    print("WROTE %s  (sheets: %s)" % (args.output, ", ".join(out.sheetnames)), file=sys.stderr)


if __name__ == "__main__":
    main()
