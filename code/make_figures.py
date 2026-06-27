import os
import sys
#!/usr/bin/env python3
"""
make_figures.py
===============
Regenerate all 7 publication figures from master_xde_clusters.xlsx
and the metric-death CC file.

Figures produced:
  fig1_table_risky_null_clusters.png  -- Tables 1+2+null SC as image
  fig2_temporal_trajectory.png        -- Temporal CC trajectory
  fig2b_age_dominance_by_year.png     -- Age dominance by year (stacked bar + % lines)
  fig3_year_supercluster_heatmap.png  -- Year x super-cluster CC heatmap
  fig4_supercluster_summary.png       -- Super-cluster 3-panel summary
  fig5_race_summary.png               -- Race/ethnicity group summary
  fig6_SI_patterns_updated.png        -- SI pattern distribution + table
  fig7_top30_forest_plot.png          -- Forest plot top 30 metrics
  Metric_Super-Cluster_Cluster.csv    -- Metric to SC/Cluster mapping (always generated)

Usage:
  python3 code/make_figures.py --master-xlsx master_xde_clusters_2745.xlsx --cc-file metric_x_death_cc_0_0_25_1.csv --out-dir figures_2745/ >& make_figures_2745.log
"""

import argparse
import os
import sys

import matplotlib
matplotlib.use('Agg')
import warnings
import matplotlib as mpl
warnings.filterwarnings('ignore', category=UserWarning)
mpl.rcParams['font.family']     = 'serif'
mpl.rcParams['font.serif']      = ['Times New Roman', 'Palatino', 'Georgia', 'DejaVu Serif', 'serif']
mpl.rcParams['axes.titlesize']  = 11
mpl.rcParams['axes.labelsize']  = 10
mpl.rcParams['xtick.labelsize'] = 9
mpl.rcParams['ytick.labelsize'] = 9
mpl.rcParams['legend.fontsize'] = 9
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
import numpy as np
import pandas as pd
from openpyxl import load_workbook

def _tee(path):
    """Print 'Saved <relpath>' to stdout and stderr."""
    _p = str(path)
    try:
        _rel = os.path.relpath(_p)
    except ValueError:
        _rel = _p
    msg = 'Saved ' + _rel
    print(msg)
    print(msg, file=sys.stderr, flush=True)


# ============================================================
# Helpers
# ============================================================

RISKY_COL = '#C0392B'
PROT_COL  = '#2980B9'
NAVY      = '#1F3864'

# SC names loaded from sem_sc_names.csv via --sc-names arg (no hardcoding)
SC_NAMES_FULL = {}

def load_sc_names(csv_path):
    """Load SC names from sem_sc_names.csv into module-level dicts."""
    global SC_NAMES_FULL
    import pandas as _pd
    if not csv_path or not os.path.exists(csv_path):
        print(f"  WARNING: sc_names file not found: {csv_path}", file=sys.stderr)
        return
    df = _pd.read_csv(csv_path)
    for _, r in df.iterrows():
        sc_id = int(r['sc_id'])
        name  = str(r.get('sc_name', f'SC{sc_id:02d}')).strip()
        short = str(r.get('sc_name_short', name)).strip()
        SC_NAMES_FULL[sc_id] = name
    print(f"  Loaded {len(SC_NAMES_FULL)} SC names from {csv_path}")


SC_COLORS = {
    9: RISKY_COL, 3: RISKY_COL, 4: '#1E8449',
    2: RISKY_COL, 5: '#AAB7B8', 10: '#7D6608',
    1: PROT_COL,  6: '#AAB7B8', 7: '#AAB7B8', 8: PROT_COL, 11: '#AAB7B8',
}  # SC01-SC11 only; key=super_cluster_id


def load_master(master_xlsx):
    """Load Multi_Year and Master tabs from master Excel."""
    wb = load_workbook(master_xlsx, data_only=True)

    def read_ws(name):
        ws = wb[name]
        from _workbook_style import detect_table_offset
        hdr_row, hdr_col = detect_table_offset(ws)
        hdr = []
        c = hdr_col
        while c <= ws.max_column:
            v = ws.cell(hdr_row, c).value
            if v is None:
                break
            hdr.append(str(v))
            c += 1
        n = len(hdr)
        rows = [{hdr[i]: ws.cell(r, hdr_col + i).value for i in range(n)}
                for r in range(hdr_row + 1, ws.max_row + 1)]
        return pd.DataFrame(rows)

    my_df  = read_ws('Multi_Year')
    all_df = read_ws('Master')
    for df in [my_df, all_df]:
        for col in ['max_signed_cc', 'n_sig_years', 'super_cluster_id',
                    'Ward100', 'rank_abs_cc']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
    return my_df, all_df


# --- Significance mode: 'lp' (p-value, default) or 'cc' (|CC| > CC_SIG) ---
SIG_MODE  = 'lp'
CC_SIG    = 0.3
SIG_LABEL = 'LP<=-13'

def _sig_count(lp_series, cc_series):
    """Count significant entries in matched LP/CC series per current mode."""
    if SIG_MODE == 'cc':
        return int((cc_series.abs() > CC_SIG).sum()) if len(cc_series) else 0
    return int((lp_series <= -13).sum()) if len(lp_series) else 0

def _sig_count_col(df, lp_col, cc_col):
    """Count significant rows of a dataframe column per current mode."""
    if SIG_MODE == 'cc':
        return int((df[cc_col].abs() > CC_SIG).sum()) if cc_col in df.columns else 0
    return int((df[lp_col] <= -13).sum()) if lp_col in df.columns else 0

def _sig_scalar(lp_val, cc_val):
    """Single-value significance test per current mode."""
    if SIG_MODE == 'cc':
        return pd.notna(cc_val) and abs(cc_val) > CC_SIG
    return pd.notna(lp_val) and lp_val <= -13


def yr_stats_from_cc(sub, cc_df, years):
    """Compute per-year CC stats for a subset of metrics."""
    out = {}
    for yr in years:
        cc_col = f'asedx_p_{yr}'
        lp_col = f'LP_asedx_p_{yr}'
        metrics = [m for m in sub['metric'] if m in cc_df.index]
        cs = cc_df.loc[metrics, cc_col].dropna() \
             if metrics and cc_col in cc_df.columns else pd.Series()
        ls = cc_df.loc[metrics, lp_col].dropna() \
             if metrics and lp_col in cc_df.columns else pd.Series()
        out[yr] = {
            'n_sig': _sig_count(ls, cs),
            'mean':  float(cs.mean())        if len(cs) else np.nan,
            'min':   float(cs.min())          if len(cs) else np.nan,
            'max':   float(cs.max())          if len(cs) else np.nan,
            'p5':    float(cs.quantile(.05))  if len(cs) else np.nan,
            'p95':   float(cs.quantile(.95))  if len(cs) else np.nan,
        }
    return out


def save(fig, path, dpi=200):
    """Save figure, adding the filename as a small label in the top margin."""
    fname = os.path.basename(path)
    fig.text(0.5, 0.995, fname, ha='center', va='top',
             fontsize=7, color='#888888', fontfamily='monospace',
             transform=fig.transFigure)
    fig.savefig(path, dpi=dpi, bbox_inches='tight', facecolor='white')
    try:
        _rel = os.path.relpath(path)
    except ValueError:
        _rel = path
    print('Saved ' + _rel, file=sys.stderr, flush=True)
    plt.close(fig)
    print(f"  Saved: {path}")


# ============================================================
# Figure 1 -- Table: top risky + protective + null SCs
# ============================================================

def make_fig1(my_df, all_df, out_dir):
    print("Fig 1: Table risky/protective/null...")

    top_risky = my_df[my_df['max_signed_cc'] > 0].nsmallest(20, 'rank_abs_cc').copy()
    top_prot  = my_df[my_df['max_signed_cc'] < 0].nsmallest(10, 'rank_abs_cc').copy()
    for df in [top_risky, top_prot]:
        df['explain'] = df['explain'].apply(
            lambda v: str(v).replace('_', ' ')
            if len(str(v)) > 50 else str(v).replace('_', ' '))

    # SC_NAMES_FULL is loaded at module level from --sc-names file

    null_scs = []
    for sc in sorted(all_df['super_cluster_id'].dropna().astype(int).unique()):
        s_sub = my_df[my_df['super_cluster_id'] == sc]
        a_sub = all_df[all_df['super_cluster_id'] == sc]
        if len(s_sub) > 0:
            continue
        csizes = a_sub.groupby('Ward100').size().sort_values(ascending=False)
        top2 = []
        for w, sz in csizes.head(2).items():
            lbl = a_sub[a_sub['Ward100'] == w]['cluster_label'].iloc[0] \
                  if len(a_sub[a_sub['Ward100'] == w]) else ''
            top2.append(f"W{int(w)}({sz}): {str(lbl)}")
        null_scs.append([f'SC{sc:02d}', SC_NAMES_FULL.get(sc, ''),
                         '  |  '.join(top2), str(len(a_sub)),
                         str(len(csizes)), '0'])

    fig = plt.figure(figsize=(22, 22), facecolor='white')
    gs  = fig.add_gridspec(3, 1, hspace=0.06, top=0.94, bottom=0.02,
                            left=0.02, right=0.98,
                            height_ratios=[20, 10, max(len(null_scs), 1)+1])

    def make_tbl(ax, data, col_labels, col_widths, title, title_bg,
                 row_colors=None, mono_cols=None):
        ax.axis('off')
        ax.set_title(title, fontsize=11, fontweight='bold', color='white',
                     backgroundcolor=title_bg, pad=6, loc='left', x=0.0)
        tbl = ax.table(cellText=data, colLabels=col_labels,
                       cellLoc='left', loc='center', colWidths=col_widths)
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)
        tbl.scale(1, 1.55)
        for j in range(len(col_labels)):
            tbl[(0, j)].set_facecolor(NAVY)
            tbl[(0, j)].set_text_props(color='white', fontweight='bold', fontsize=9)
            tbl[(0, j)].set_edgecolor('#FFFFFF')
        for i in range(1, len(data)+1):
            bg = row_colors[i-1] if row_colors else 'white'
            for j in range(len(col_labels)):
                tbl[(i, j)].set_facecolor(bg)
                tbl[(i, j)].set_edgecolor('#DDDDDD')
                if mono_cols and j in mono_cols:
                    tbl[(i, j)].set_text_props(fontfamily='monospace', fontsize=8.5)

    def shorten_sc(sc):
        return f"SC{int(sc):02d}: {SC_NAMES_FULL.get(int(sc), '')}" \
               if pd.notna(sc) else ''

    ax1 = fig.add_subplot(gs[0])
    risky_data, risky_colors = [], []
    for _, r in top_risky.iterrows():
        rk = int(r['rank_abs_cc'])
        risky_data.append([shorten_sc(r['super_cluster_id']),
                           str(r.get('cluster_label', '')),
                           f'#{rk}', r['explain'],
                           f"+{r['max_signed_cc']:.3f}", r['SI'], r['sign']])
        risky_colors.append('#FEF0EE' if rk <= 10 else '#FCE4D6')
    make_tbl(ax1, risky_data,
             ['Super-cluster', 'Cluster', 'Rank', 'Metric description', 'CC', 'SI', 'Sign'],
             [0.18, 0.24, 0.05, 0.30, 0.07, 0.07, 0.07],
             '  ^  TOP 20 RISKY METRICS', '#8B0000', risky_colors, mono_cols={5, 6})

    ax2 = fig.add_subplot(gs[1])
    prot_data, prot_colors = [], []
    for _, r in top_prot.iterrows():
        rk = int(r['rank_abs_cc'])
        prot_data.append([shorten_sc(r['super_cluster_id']),
                          str(r.get('cluster_label', '')),
                          f'#{rk}', r['explain'],
                          f"{r['max_signed_cc']:.3f}", r['SI'], r['sign']])
        prot_colors.append('#EBF5FB' if rk <= 15 else '#D6E4F0')
    make_tbl(ax2, prot_data,
             ['Super-cluster', 'Cluster', 'Rank', 'Metric description', 'CC', 'SI', 'Sign'],
             [0.18, 0.24, 0.05, 0.30, 0.07, 0.07, 0.07],
             '  v  TOP 10 PROTECTIVE METRICS', '#1F4E79', prot_colors, mono_cols={5, 6})

    ax3 = fig.add_subplot(gs[2])
    if null_scs:
        make_tbl(ax3, null_scs,
                 ['SC', 'Super-cluster name', 'Largest clusters (n metrics)',
                  'Total\nmetrics', 'Ward\nclusters', 'Multi-year\nsig'],
                 [0.05, 0.27, 0.40, 0.07, 0.07, 0.08],
                 '  X  NULL SUPER-CLUSTERS (0 significant metrics any year)',
                 '#37474F', ['#ECEFF1'] * len(null_scs))
    else:
        ax3.axis('off')
        ax3.text(0.5, 0.5, 'All super-clusters have at least one significant metric',
                 ha='center', va='center', fontsize=10, color='#37474F',
                 transform=ax3.transAxes)

    fig.suptitle('Table 1.  Strongest predictors of US county pandemic excess mortality 2020-2024\n'
                 'and super-clusters with zero significant associations',
                 fontsize=12, fontweight='bold', y=0.975)
    save(fig, os.path.join(out_dir, 'fig1_table_risky_null_clusters.png'))


# ============================================================
# Figure 2 -- Temporal trajectory
# ============================================================

def make_fig2(my_df, cc_df, years, out_dir):
    print("Fig 2: Temporal trajectory...")
    pos = my_df[my_df['max_signed_cc'] > 0]
    neg = my_df[my_df['max_signed_cc'] < 0]
    ps  = yr_stats_from_cc(pos, cc_df, years)
    ns  = yr_stats_from_cc(neg, cc_df, years)

    x       = np.arange(len(years))
    p_means = [ps[yr]['mean'] for yr in years]
    p_p5    = [ps[yr]['p5']   for yr in years]
    p_p95   = [ps[yr]['p95']  for yr in years]
    p_mins  = [ps[yr]['min']  for yr in years]
    p_maxs  = [ps[yr]['max']  for yr in years]
    n_means = [ns[yr]['mean'] for yr in years]
    n_p5    = [ns[yr]['p5']   for yr in years]
    n_p95   = [ns[yr]['p95']  for yr in years]
    n_mins  = [ns[yr]['min']  for yr in years]
    n_maxs  = [ns[yr]['max']  for yr in years]

    fig, ax = plt.subplots(figsize=(6.7, 6))
    ax.fill_between(x, p_mins, p_maxs, alpha=0.10, color=RISKY_COL)
    ax.fill_between(x, n_maxs, n_mins, alpha=0.10, color=PROT_COL)
    ax.fill_between(x, p_p5,  p_p95,  alpha=0.22, color=RISKY_COL, label='Risky 5th-95th %ile')
    ax.fill_between(x, n_p5,  n_p95,  alpha=0.22, color=PROT_COL,  label='Protective 5th-95th %ile')
    ax.plot(x, p_means, 'o-', color=RISKY_COL, lw=2.5, ms=9, zorder=5, label='Risky mean CC')
    ax.plot(x, n_means, 's-', color=PROT_COL,  lw=2.5, ms=9, zorder=5, label='Protective mean CC')
    ax.plot(x, p_maxs,  'v',  color=RISKY_COL, ms=7, alpha=0.55, label='Risky max/min CC')
    ax.plot(x, p_mins,  '^',  color=RISKY_COL, ms=7, alpha=0.55)
    ax.plot(x, n_mins,  'v',  color=PROT_COL,  ms=7, alpha=0.55, label='Protective max/min CC')
    ax.plot(x, n_maxs,  '^',  color=PROT_COL,  ms=7, alpha=0.55)
    for i, yr in enumerate(years):
        ax.annotate(f'n={ps[yr]["n_sig"]}', (x[i], p_means[i]),
                    textcoords='offset points', xytext=(0, 10),
                    ha='center', fontsize=8, color=RISKY_COL, fontweight='bold')
        ax.annotate(f'n={ns[yr]["n_sig"]}', (x[i], n_means[i]),
                    textcoords='offset points', xytext=(0, -15),
                    ha='center', fontsize=8, color=PROT_COL, fontweight='bold')
        ax.annotate(f'{p_maxs[i]:+.2f}', (x[i], p_maxs[i]),
                    textcoords='offset points', xytext=(5, 1),
                    ha='left', fontsize=7, color=RISKY_COL, alpha=0.85)
        ax.annotate(f'{p_mins[i]:+.2f}', (x[i], p_mins[i]),
                    textcoords='offset points', xytext=(5, 1),
                    ha='left', fontsize=7, color=RISKY_COL, alpha=0.85)
        ax.annotate(f'{n_mins[i]:+.2f}', (x[i], n_mins[i]),
                    textcoords='offset points', xytext=(5, 1),
                    ha='left', fontsize=7, color=PROT_COL, alpha=0.85)
        ax.annotate(f'{n_maxs[i]:+.2f}', (x[i], n_maxs[i]),
                    textcoords='offset points', xytext=(5, 1),
                    ha='left', fontsize=7, color=PROT_COL, alpha=0.85)
    ax.axhline(0, color='black', lw=0.7, ls='--', alpha=0.4)
    ax.set_xticks(x); ax.set_xticklabels(years, fontsize=11)
    ax.set_ylabel('Correlation coefficient (CC)', fontsize=11)
    ax.set_xlabel('Year', fontsize=11)
    ax.set_title('Temporal trajectory of CC by year\n'
                 'Lines=mean  |  Bands=5th-95th %ile  |  Markers=min/max', fontsize=10)
    ax.legend(fontsize=8, loc='upper right', framealpha=0.92)
    ax.grid(True, alpha=0.25)
    all_vals = p_mins + n_mins
    all_maxs = p_maxs + n_maxs
    ax.set_ylim(min(all_vals)*1.15, max(all_maxs)*1.22)
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig2_temporal_trajectory.png'))



# ============================================================
# Figure 2b -- Age dominance by year
# ============================================================

def make_fig2b(my_df, cc_df, years, out_dir):
    print("Fig 2b: Age dominance by year...")
    age_data = {}
    for yr in years:
        lp_ge = f'LP_asedx_p_{yr}_GE65'
        lp_lt = f'LP_asedx_p_{yr}_LT65'
        cc_ge = f'asedx_p_{yr}_GE65'
        cc_lt = f'asedx_p_{yr}_LT65'
        ge65 = lt65 = both = 0
        for m in my_df['metric']:
            if m not in cc_df.index: continue
            row  = cc_df.loc[m]
            sg   = _sig_scalar(row.get(lp_ge), row.get(cc_ge))
            sl   = _sig_scalar(row.get(lp_lt), row.get(cc_lt))
            cg   = row.get(cc_ge, np.nan)
            cl   = row.get(cc_lt, np.nan)
            if sg and sl:
                both += 1
            elif sg:
                ge65 += 1
            elif sl:
                lt65 += 1
        age_data[yr] = {'GE65': ge65, 'LT65': lt65, 'both': both}

    x  = np.arange(len(years))
    ge = [age_data[yr]['GE65'] for yr in years]
    lt = [age_data[yr]['LT65'] for yr in years]
    bt = [age_data[yr]['both'] for yr in years]
    tot = [ge[i]+lt[i]+bt[i] for i in range(len(years))]
    ge_pct = [ge[i]/tot[i]*100 if tot[i] else 0 for i in range(len(years))]
    lt_pct = [lt[i]/tot[i]*100 if tot[i] else 0 for i in range(len(years))]
    bt_pct = [bt[i]/tot[i]*100 if tot[i] else 0 for i in range(len(years))]

    C_GE = '#2980B9'; C_LT = '#E74C3C'; C_BT = '#8E44AD'

    # ---- Raw per-year sig counts across ALL metrics in cc_df (not just my_df) ----
    # These are the underlying signal counts the third panel visualises.
    raw_ge, raw_lt = [], []
    for yr in years:
        lp_ge_col = f'LP_asedx_p_{yr}_GE65'
        lp_lt_col = f'LP_asedx_p_{yr}_LT65'
        cc_ge_col = f'asedx_p_{yr}_GE65'
        cc_lt_col = f'asedx_p_{yr}_LT65'
        raw_ge.append(_sig_count_col(cc_df, lp_ge_col, cc_ge_col))
        raw_lt.append(_sig_count_col(cc_df, lp_lt_col, cc_lt_col))

    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(21, 7))

    # Left: stacked bar
    w = 0.6
    ax1.bar(x, ge, w, label='GE65 dominant (>)', color=C_GE, alpha=0.85)
    ax1.bar(x, lt, w, bottom=ge, label='LT65 dominant (<)', color=C_LT, alpha=0.85)
    ax1.bar(x, bt, w, bottom=[ge[i]+lt[i] for i in range(len(years))],
            label='Both age groups', color=C_BT, alpha=0.85)
    for i in range(len(years)):
        if ge[i] > 0:
            ax1.text(i, ge[i]/2, str(ge[i]), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        if lt[i] > 0:
            ax1.text(i, ge[i]+lt[i]/2, str(lt[i]), ha='center', va='center',
                     fontsize=10, color='white', fontweight='bold')
        if bt[i] > 0:
            ax1.text(i, ge[i]+lt[i]+bt[i]/2, str(bt[i]), ha='center', va='center',
                     fontsize=9, color='white', fontweight='bold')
        ax1.text(i, tot[i]+2, f'n={tot[i]}', ha='center', va='bottom', fontsize=9)
    ax1.set_xticks(x); ax1.set_xticklabels(years, fontsize=12)
    ax1.set_ylabel('Number of significant metrics (any year)', fontsize=11)
    ax1.set_title('Age group dominance by year\n(absolute counts)', fontsize=12, fontweight='bold')
    ax1.legend(fontsize=10, loc='upper right')
    ax1.grid(True, axis='y', alpha=0.25)

    # Right: line plot %
    ax2.plot(x, ge_pct, 'o-', color=C_GE, lw=2.5, ms=9, label='GE65 dominant')
    ax2.plot(x, lt_pct, 's-', color=C_LT, lw=2.5, ms=9, label='LT65 dominant')
    ax2.plot(x, bt_pct, '^--', color=C_BT, lw=1.5, ms=7, label='Both')
    for i in range(len(years)):
        ax2.annotate(f'{ge_pct[i]:.0f}%', (x[i], ge_pct[i]),
                     textcoords='offset points', xytext=(0, 9),
                     ha='center', fontsize=9, color=C_GE, fontweight='bold')
        if lt_pct[i] > 0:
            ax2.annotate(f'{lt_pct[i]:.0f}%', (x[i], lt_pct[i]),
                         textcoords='offset points', xytext=(0, -14),
                         ha='center', fontsize=9, color=C_LT, fontweight='bold')
    ax2.set_xticks(x); ax2.set_xticklabels(years, fontsize=12)
    ax2.set_ylabel('% of age-specific significant metrics', fontsize=11)
    ax2.set_title('Age dominance shift across pandemic waves\n'
                  '(% of age-specific significant metrics)', fontsize=12, fontweight='bold')
    ax2.legend(fontsize=10); ax2.set_ylim(0, 110)
    ax2.grid(True, alpha=0.25)
    # Wave shading
    ax2.axvspan(-0.4, 0.4, alpha=0.06, color='grey')
    ax2.axvspan(0.6,  1.4, alpha=0.06, color='red')
    ax2.axvspan(1.6,  2.4, alpha=0.06, color='blue')
    ax2.text(0,  103, 'COVID-19\n(2020)', ha='center', fontsize=7.5, color='grey')
    ax2.text(1,  103, 'Delta\n(2021)',    ha='center', fontsize=7.5, color='#B03A2E')
    ax2.text(2,  103, 'Omicron\n(2022)', ha='center', fontsize=7.5, color='#1A5276')

    # ---- Right panel: GE65 vs LT65 share crossover (raw cc_df, %) ----
    # Each year, GE65% = raw_ge / (raw_ge + raw_lt) and LT65% similarly,
    # so the two lines sum to 100% and the dominance flips show up as
    # crossovers. Uses raw counts from cc_df (every metric), not my_df, so
    # the 2023 LT65 spike (73%) and 2024 GE65 collapse (96%) are visible.
    ge_pct_raw, lt_pct_raw = [], []
    for i in range(len(years)):
        tot = raw_ge[i] + raw_lt[i]
        if tot == 0:
            ge_pct_raw.append(0.0); lt_pct_raw.append(0.0)
        else:
            ge_pct_raw.append(raw_ge[i] / tot * 100)
            lt_pct_raw.append(raw_lt[i] / tot * 100)
    total_sig = sum(raw_ge) + sum(raw_lt)

    # Wave shading (COVID-19 / Delta / Omicron)
    ax3.axvspan(-0.4, 0.4, alpha=0.06, color='grey')
    ax3.axvspan(0.6,  1.4, alpha=0.06, color=C_LT)
    ax3.axvspan(1.6,  2.4, alpha=0.06, color=C_GE)
    ax3.text(0,  108, 'COVID-19', ha='center', fontsize=8.5, color='grey')
    ax3.text(1,  108, 'Delta',    ha='center', fontsize=8.5,
             color='#B03A2E', fontweight='bold')
    ax3.text(2,  108, 'Omicron',  ha='center', fontsize=8.5,
             color='#1A5276', fontweight='bold')

    # 50% parity reference
    ax3.axhline(50, color='black', lw=0.7, ls='--', alpha=0.35)

    # The two crossover lines
    ax3.plot(x, ge_pct_raw, 'o-', color=C_GE, lw=3, ms=11,
             label='GE65 sig (% of age-specific sig)')
    ax3.plot(x, lt_pct_raw, 's-', color=C_LT, lw=3, ms=11,
             label='LT65 sig (% of age-specific sig)')

    # Per-year labels: GE65 above, LT65 below
    for i in range(len(years)):
        ax3.annotate(f'{ge_pct_raw[i]:.0f}%', (x[i], ge_pct_raw[i]),
                     textcoords='offset points', xytext=(0, 12),
                     ha='center', fontsize=11, color=C_GE, fontweight='bold')
        ax3.annotate(f'{lt_pct_raw[i]:.0f}%', (x[i], lt_pct_raw[i]),
                     textcoords='offset points', xytext=(0, -16),
                     ha='center', fontsize=11, color=C_LT, fontweight='bold')

    ax3.set_xticks(x)
    ax3.set_xticklabels(years, fontsize=12)
    ax3.set_ylabel('% of age-specific sig metrics', fontsize=11)
    ax3.set_title(f'GE65 vs LT65 share crossover\n'
                  f'({SIG_LABEL};  total age-specific sig = {total_sig})',
                  fontsize=12, fontweight='bold')
    ax3.legend(fontsize=9, loc='lower left')
    ax3.set_ylim(-5, 115)
    ax3.grid(True, alpha=0.25)

    fig.suptitle('Age dominance oscillates across pandemic phases:  '
                 'LT65 in 2020 / 2021 / 2023, GE65 in 2022 / 2024',
                 fontsize=13, fontweight='bold')
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig2b_age_dominance_by_year.png'))

# ============================================================
# Figure 3 -- Year x super-cluster heatmap
# ============================================================

def make_fig3(my_df, cc_df, years, out_dir):
    print("Fig 3: Year x super-cluster heatmap...")
    scs = sorted([int(s) for s in my_df['super_cluster_id'].dropna().unique()])
    heatmap_cc  = np.zeros((len(scs), len(years)))
    heatmap_sig = np.zeros((len(scs), len(years)), dtype=int)

    for si, sc in enumerate(scs):
        sub     = my_df[my_df['super_cluster_id'] == sc]
        metrics = [m for m in sub['metric'] if m in cc_df.index]
        for yi, yr in enumerate(years):
            cc_col = f'asedx_p_{yr}'
            lp_col = f'LP_asedx_p_{yr}'
            cs = cc_df.loc[metrics, cc_col].dropna() \
                 if cc_col in cc_df.columns and metrics else pd.Series()
            ls = cc_df.loc[metrics, lp_col].dropna() \
                 if lp_col in cc_df.columns and metrics else pd.Series()
            heatmap_cc[si,  yi] = cs.mean() if len(cs) else 0
            heatmap_sig[si, yi] = _sig_count(ls, cs)

    sc_labels = [f'SC{sc}: {SC_NAMES_FULL.get(sc, "")}' for sc in scs]
    cmap = mcolors.LinearSegmentedColormap.from_list(
        'rw', [(0, '#922B21'), (0.5, '#FDFEFE'), (1, '#1A5276')])

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 5))
    im1 = ax1.imshow(heatmap_cc,  aspect='auto', cmap=cmap,
                     vmin=-0.35, vmax=0.35, interpolation='none')
    im2 = ax2.imshow(heatmap_sig, aspect='auto', cmap=plt.cm.YlOrRd,
                     vmin=0, vmax=heatmap_sig.max(), interpolation='none')
    for ax, data, iscc in [(ax1, heatmap_cc, True), (ax2, heatmap_sig, False)]:
        ax.set_xticks(range(len(years))); ax.set_xticklabels(years, fontsize=11)
        ax.set_yticks(range(len(scs)));   ax.set_yticklabels(sc_labels, fontsize=10)
        for si in range(len(scs)):
            for yi in range(len(years)):
                v = data[si, yi]
                if iscc:
                    if abs(v) > 0.01:
                        col = 'white' if abs(v) > 0.22 else 'black'
                        ax.text(yi, si, f'{v:+.2f}', ha='center', va='center',
                                fontsize=6.5, color=col)
                else:
                    if v > 0:
                        col = 'white' if v > heatmap_sig.max()*0.6 else 'black'
                        ax.text(yi, si, str(int(v)), ha='center', va='center',
                                fontsize=7, color=col)
    ax1.set_title('Mean CC by super-cluster x year\n(red=risky, blue=protective)',
                  fontsize=11, fontweight='bold')
    ax2.set_title(f'n metrics significant ({SIG_LABEL})\nby super-cluster x year',
                  fontsize=11, fontweight='bold')
    fig.colorbar(im1, ax=ax1, fraction=0.02, pad=0.02).set_label('Mean CC', fontsize=9)
    fig.colorbar(im2, ax=ax2, fraction=0.02, pad=0.02).set_label('n significant', fontsize=9)
    fig.suptitle('Temporal x domain structure: SC8 persists; other SCs vary',
                 fontsize=12, fontweight='bold')
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig3_year_supercluster_heatmap.png'))


# ============================================================
# Figure 4 -- Super-cluster 3-panel summary
# ============================================================

def make_fig4(my_df, all_df, out_dir):
    print("Fig 4: Super-cluster summary...")
    sc_data = []
    for sc in sorted(all_df['super_cluster_id'].dropna().astype(int).unique()):
        a_sub = all_df[all_df['super_cluster_id'] == sc]
        s_sub = my_df[my_df['super_cluster_id'] == sc]
        n_all = len(a_sub); n_sig = len(s_sub)
        pct   = round(n_sig/n_all*100, 1) if n_all else 0
        n_pos = int((s_sub['max_signed_cc'] > 0).sum())
        n_neg = int((s_sub['max_signed_cc'] < 0).sum())
        mac   = float(s_sub['max_signed_cc'].abs().mean()) if n_sig else 0
        sc_data.append({'sc': sc, 'name': SC_NAMES_FULL.get(sc, ''), 'n_all': n_all,
                        'n_sig': n_sig, 'pct': pct, 'n_pos': n_pos,
                        'n_neg': n_neg, 'mac': mac})
    sc_df = pd.DataFrame(sc_data).sort_values('n_sig', ascending=True)
    y     = range(len(sc_df))

    colors1 = [RISKY_COL if r['n_pos'] > r['n_neg']
               else (PROT_COL if r['n_neg'] > r['n_pos'] else '#BDC3C7')
               for _, r in sc_df.iterrows()]

    fig, axes = plt.subplots(1, 3, figsize=(18, 7),
                              gridspec_kw={'width_ratios': [1.4, 1, 1]})
    ax = axes[0]
    ax.barh(list(y), sc_df['pct'].tolist(), color=colors1, alpha=0.85, edgecolor='white')
    ax.set_yticks(list(y))
    ax.set_yticklabels([f'SC{int(r.sc):02d}: {r["name"]}' for _, r in sc_df.iterrows()], fontsize=9.5)
    ax.set_xlabel('% metrics significant (any year)', fontsize=10)
    ax.set_title('Significance rate\nby super-cluster', fontsize=11, fontweight='bold')
    for i, (_, r) in enumerate(sc_df.iterrows()):
        if r['pct'] > 0:
            ax.text(r['pct']+0.5, i, f"{int(r['n_sig'])}/{int(r['n_all'])}  ({r['pct']:.0f}%)",
                    va='center', fontsize=8.5)
        else:
            ax.text(0.5, i, '0', va='center', fontsize=9, color='#E74C3C', fontweight='bold')
    ax.set_xlim(0, 120); ax.grid(True, axis='x', alpha=0.25)

    ax2 = axes[1]
    pv = sc_df['n_pos'].tolist(); nv = sc_df['n_neg'].tolist()
    ax2.barh(list(y), pv, color=RISKY_COL, alpha=0.85, label='Risky (+CC)', edgecolor='white')
    ax2.barh(list(y), nv, left=pv, color=PROT_COL, alpha=0.85,
             label='Protective (-CC)', edgecolor='white')
    ax2.set_yticks(list(y)); ax2.set_yticklabels(['']*len(sc_df))
    ax2.set_xlabel('Number of significant metrics', fontsize=10)
    ax2.set_title('Direction of\nassociation', fontsize=11, fontweight='bold')
    ax2.legend(fontsize=9, loc='lower right')
    for i, (_, r) in enumerate(sc_df.iterrows()):
        if r['n_pos'] > 0:
            ax2.text(r['n_pos']/2, i, str(int(r['n_pos'])), ha='center', va='center',
                     fontsize=8, color='white', fontweight='bold')
        if r['n_neg'] > 0:
            ax2.text(r['n_pos']+r['n_neg']/2, i, str(int(r['n_neg'])), ha='center',
                     va='center', fontsize=8, color='white', fontweight='bold')
    ax2.grid(True, axis='x', alpha=0.25)

    ax3 = axes[2]
    ax3.barh(list(y), sc_df['mac'].tolist(), color=colors1, alpha=0.85, edgecolor='white')
    ax3.set_yticks(list(y)); ax3.set_yticklabels(['']*len(sc_df))
    ax3.set_xlabel('Mean |CC|', fontsize=10)
    ax3.set_title('Signal strength\n(mean |CC|)', fontsize=11, fontweight='bold')
    for i, v in enumerate(sc_df['mac'].tolist()):
        if v > 0:
            ax3.text(v+0.003, i, f'{v:.3f}', va='center', fontsize=8.5)
    ax3.set_xlim(0, 0.42); ax3.grid(True, axis='x', alpha=0.25)

    rp  = mpatches.Patch(color=RISKY_COL, alpha=0.85, label='Predominantly risky')
    pp  = mpatches.Patch(color=PROT_COL,  alpha=0.85, label='Predominantly protective')
    np_ = mpatches.Patch(color='#BDC3C7', alpha=0.85, label='No significant metrics')
    fig.legend(handles=[rp, pp, np_], loc='lower center', ncol=3,
               fontsize=10, bbox_to_anchor=(0.5, -0.04))
    fig.suptitle('Super-cluster significance summary', fontsize=12, fontweight='bold')
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig4_supercluster_summary.png'))


# ============================================================
# Figure 5 -- Race group summary
# ============================================================

def make_fig5(my_df, out_dir):
    print("Fig 5: Race summary...")
    race_defs = [
        ('White non-Hispanic',
         r'[Ww]hite.?[Nn]on.?[Hh]ispanic|[Ww]hite_[Nn]on.?[Hh]ispanic', None),
        ('Hispanic / Latino',
         r'[Hh]ispanic|[Ll]atino',
         r'[Ww]hite.?[Nn]on.?[Hh]ispanic'),
        ('AIAN (American Indian / Alaska Native)',
         r'[Aa]merican.?[Ii]ndian|[Aa]laska.?[Nn]ative', None),
        ('Multiracial / Some Other Race',
         r'[Mm]ultiracial|[Ss]ome.?[Oo]ther.?[Rr]ace', None),
        ('Asian',
         r'\b[Aa]sian\b',
         r'[Hh]ispanic|[Ll]atino|[Ww]hite|[Bb]lack|[Nn]ative'),
        ('Black / African American',
         r'\b[Bb]lack\b|[Aa]frican.?[Aa]merican', None),
    ]
    race_stats = []
    for grp, pat, excl in race_defs:
        sub = my_df[my_df['explain'].str.contains(pat, na=False, regex=True, case=False)]
        if excl:
            sub = sub[~sub['explain'].str.contains(excl, na=False, regex=True, case=False)]
        n   = len(sub)
        np_ = int((sub['max_signed_cc'] > 0).sum())
        nn  = int((sub['max_signed_cc'] < 0).sum())
        mc  = round(float(sub['max_signed_cc'].mean()), 3) if n > 0 else None
        race_stats.append({'group': grp, 'n': n, 'n_pos': np_, 'n_neg': nn, 'mean_cc': mc})
    race_df = pd.DataFrame(race_stats)

    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    y = range(len(race_df))

    ax = axes[0]
    pos_vals = race_df['n_pos'].tolist(); neg_vals = race_df['n_neg'].tolist()
    ax.barh(list(y), pos_vals, color=RISKY_COL, alpha=0.85,
            label='Risky (+CC)', edgecolor='white')
    ax.barh(list(y), neg_vals, left=pos_vals, color=PROT_COL, alpha=0.85,
            label='Protective (-CC)', edgecolor='white')
    ax.set_yticks(list(y)); ax.set_yticklabels(race_df['group'].tolist(), fontsize=11)
    ax.set_xlabel('Number of significant metrics (any year)', fontsize=11)
    ax.set_title('Multi-year significant metrics\nby race/ethnicity group',
                 fontsize=12, fontweight='bold')
    ax.legend(fontsize=10)
    for i, (_, r) in enumerate(race_df.iterrows()):
        if r['n_pos'] > 0:
            ax.text(r['n_pos']/2, i, str(int(r['n_pos'])), ha='center', va='center',
                    fontsize=9, color='white', fontweight='bold')
        if r['n_neg'] > 0:
            ax.text(r['n_pos']+r['n_neg']/2, i, str(int(r['n_neg'])), ha='center',
                    va='center', fontsize=9, color='white', fontweight='bold')
        total = int(r['n_pos']+r['n_neg'])
        ax.text(total+0.3, i,
                f"n={total}  ({r['n_pos']/max(total,1)*100:.0f}% risky)",
                va='center', fontsize=8.5)
    ax.grid(True, axis='x', alpha=0.25)

    ax2 = axes[1]
    for i, (_, r) in enumerate(race_df.iterrows()):
        if r['mean_cc'] is None:
            continue
        col = RISKY_COL if r['mean_cc'] > 0 else PROT_COL
        n_  = max(r['n'], 1)
        ax2.scatter(r['mean_cc'], i, s=n_*8+30, color=col, alpha=0.85,
                    edgecolors='white', lw=1.5, zorder=5)
        ha = 'left' if r['mean_cc'] >= 0 else 'right'
        off = 0.005 if r['mean_cc'] >= 0 else -0.005
        ax2.text(r['mean_cc']+off, i, f"{r['mean_cc']:+.3f}",
                 va='center', ha=ha, fontsize=9.5, fontweight='bold')
    ax2.axvline(0, color='black', lw=1.0, ls='--', alpha=0.6)
    ax2.set_yticks(list(y))
    ax2.set_yticklabels(race_df['group'].tolist(), fontsize=11)
    ax2.set_xlabel('Mean signed CC', fontsize=10)
    ax2.set_title('Mean CC by race/ethnicity\n(bubble size proportional to n metrics)',
                  fontsize=12, fontweight='bold')
    ax2.grid(True, alpha=0.25); ax2.set_xlim(-0.35, 0.35)
    fig.suptitle('Race/ethnicity patterns: White non-Hispanic protective; AIAN risky',
                 fontsize=12, fontweight='bold')
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig5_race_summary.png'))


# ============================================================
# Figure 6 -- SI patterns
# ============================================================

def make_fig6(my_df, out_dir, all_df=None):
    """SI pattern horizontal bar chart only (PNG).
    Table goes to Excel via make_extra_tables.py Table_SI tab.
    Uses all metrics with at least 1 S in SI string (n_sig_years >= 1).
    """
    print("Fig 6: SI patterns (bar chart only)...")
    # Include any metric with at least one significant year
    df = (all_df if all_df is not None else my_df).copy()
    # Use a two-character I/S alphabet: merge very-significant ('V') into 'S'.
    # (Replace BEFORE filtering so metrics that are only ever very-significant are kept.)
    df['SI'] = df['SI'].astype(str).str.replace('V', 'S', regex=False)
    df = df[df['SI'].str.contains('S', na=False)].copy()

    # Count metrics per (sign-invariant) temporal SI pattern.  We deliberately do NOT split
    # by +CC/-CC: the sign of a metric's CC is a coding convention ("% with" vs "% without",
    # counts of advantaged vs disadvantaged groups), so an aggregate protective/risky count
    # is an artifact of metric definitions, not a property of the data.  The meaningful,
    # sign-invariant quantity is the temporal pattern itself.
    counts = df['SI'].value_counts()                 # descending by count
    bar_si_r = list(reversed(counts.index.tolist())) # largest at top for barh
    bar_v_r  = [int(counts[s]) for s in bar_si_r]

    fig, ax = plt.subplots(figsize=(4.5, 7), facecolor='white')
    y = np.arange(len(bar_si_r))
    ax.barh(y, bar_v_r, 0.6, color='#4F81BD', alpha=0.9)
    for i, v in enumerate(bar_v_r):
        if v > 0:
            ax.text(v+0.4, i, str(v), va='center', fontsize=8, fontweight='bold', color='#1F3864')

    ax.set_yticks(y)
    ax.set_yticklabels(bar_si_r, fontfamily='monospace', fontsize=9.5, fontweight='bold')
    ax.set_xlabel('Number of metrics', fontsize=10)
    ax.grid(True, axis='x', alpha=0.25)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig6_SI_patterns_updated.png'))



# ============================================================
# Figure 7 -- Forest plot top 30
# ============================================================

def make_fig7(my_df, out_dir):
    print("Fig 7: Forest plot top 30...")
    top30 = my_df.nsmallest(30, 'rank_abs_cc').iloc[::-1].copy()
    labels = []
    for _, r in top30.iterrows():
        e = str(r.get('explain', '')).replace('_', ' ').strip()
        labels.append(e)
    ccs    = top30['max_signed_cc'].values
    sc_ids = top30['super_cluster_id'].values
    colors = [SC_COLORS.get(int(s), '#888888') if pd.notna(s) else '#888888'
              for s in sc_ids]
    sis    = top30['SI'].values

    fig, ax = plt.subplots(figsize=(13, 10))
    y = np.arange(len(top30))
    ax.barh(y, ccs, color=colors, alpha=0.85, edgecolor='white', height=0.7)
    ax.axvline(0, color='black', lw=0.8)
    for i, (cc, si) in enumerate(zip(ccs, sis)):
        xoff = 0.004 if cc > 0 else -0.004
        ha   = 'left' if cc > 0 else 'right'
        ax.text(cc+xoff, i, f'{cc:+.3f}  {si}', va='center', fontsize=7.5,
                color='white' if abs(cc) > 0.22 else 'black', ha=ha,
                fontfamily='monospace')
    ax.set_yticks(y); ax.set_yticklabels(labels, fontsize=8.5)
    ax.set_xlabel('Correlation coefficient (CC)', fontsize=11)
    ax.set_title('Top 30 significant metrics by |CC| (any year)\n'
                 'Colour = super-cluster  |  Values show CC and SI pattern',
                 fontsize=11, fontweight='bold')
    ax.grid(True, axis='x', alpha=0.25)
    ax.set_xlim(min(ccs)*1.35, max(ccs)*1.35)
    present_scs = sorted(set([int(s) for s in sc_ids if pd.notna(s)]))
    handles = [mpatches.Patch(color=SC_COLORS.get(sc, '#888'), alpha=0.85,
                               label=f'SC{sc}: {SC_NAMES_FULL.get(sc, "")}')
               for sc in present_scs]
    ax.legend(handles=handles, fontsize=8.5, loc='lower right', framealpha=0.9)
    for i, (_, r) in enumerate(top30.iterrows()):
        rk = int(r['rank_abs_cc']) if pd.notna(r.get('rank_abs_cc')) else i+1
        ax.text(min(ccs)*1.32, i, f'#{rk}', va='center', fontsize=7.5,
                color='#666666', ha='left')
    plt.tight_layout()
    save(fig, os.path.join(out_dir, 'fig7_top30_forest_plot.png'))



# ============================================================
# CSV -- Metric / Super-Cluster / Cluster mapping
# ============================================================

def make_metric_sc_cluster_csv(my_df_all, out_dir):
    """Generate Metric_Super-Cluster_Cluster.csv from the Master tab data."""
    print("CSV: Metric_Super-Cluster_Cluster...")
    out = my_df_all[['metric','explain','super_cluster_id',
                      'super_cluster_name','Ward100','cluster_label']].copy()
    out = out.rename(columns={'Ward100': 'Cluster ID'})
    out['super_cluster_id'] = pd.to_numeric(out['super_cluster_id'],
                                             errors='coerce').astype('Int64')
    out['Cluster ID']       = pd.to_numeric(out['Cluster ID'],
                                             errors='coerce').astype('Int64')
    out = out.sort_values(['super_cluster_id','Cluster ID','metric']).reset_index(drop=True)
    path = os.path.join(out_dir, 'Metric_Super-Cluster_Cluster.csv')
    out.to_csv(path, index=False)
    _tee(path)
    print(f"  Saved: {path}  ({len(out)} rows)")
    return out

# ============================================================
# Main
# ============================================================

def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--sc-names',    required=True,
                   help='sem_sc_names.csv (ward_xde_*/sem_sc_names.csv)')
    p.add_argument('--master-xlsx', required=True,
                   help='master_xde_clusters.xlsx')
    p.add_argument('--cc-file',     required=True,
                   help='metric_x_death_cc_*.csv')
    p.add_argument('--out-dir',     default='figures/',
                   help='Output directory for PNG files (default: figures/)')
    p.add_argument('--figs',        default='1,2,3,4,5,6,7',
                   help='Comma-separated list of figures to make (default: all)')
    p.add_argument('--sig-mode', choices=['lp', 'cc'], default='cc',
                   help="Significance definition: 'cc' (|CC|>--cc-sig, default) or 'lp' (legacy)")
    p.add_argument('--cc-sig', type=float, default=0.3,
                   help="|CC| Significant threshold in cc mode (default 0.3)")
    return p.parse_args()


if __name__ == '__main__':
    args  = parse_args()
    SIG_MODE  = args.sig_mode
    CC_SIG    = args.cc_sig
    SIG_LABEL = (f'|CC|>{CC_SIG:g}' if SIG_MODE == 'cc' else 'LP<=-13')
    figs_to_make = set(int(x) for x in args.figs.split(','))

    os.makedirs(args.out_dir, exist_ok=True)

    print(f"Loading master Excel: {args.master_xlsx}")
    load_sc_names(args.sc_names)
    my_df, all_df = load_master(args.master_xlsx)
    print(f"  Multi_Year: {len(my_df)} rows")
    print(f"  Master:     {len(all_df)} rows")

    print(f"Loading CC file: {args.cc_file}")
    cc_df = pd.read_csv(args.cc_file).set_index('metric')
    print(f"  CC metrics: {len(cc_df)}")

    YEARS = ['2020', '2021', '2022', '2023', '2024']

    import traceback as _tb
    _failed = []
    def _run(n, fn, *a, **kw):
        try:
            fn(*a, **kw)
        except Exception as _e:
            print(f'ERROR in Fig {n}: {_e}', file=sys.stderr)
            _tb.print_exc(file=sys.stderr)
            _failed.append(n)

    if 1 in figs_to_make:  _run(1,  make_fig1, my_df, all_df, args.out_dir)
    if 2 in figs_to_make:  _run(2,  make_fig2, my_df, cc_df, YEARS, args.out_dir)
    if 2 in figs_to_make:  _run('2b', make_fig2b, my_df, cc_df, YEARS, args.out_dir)
    if 3 in figs_to_make:  _run(3,  make_fig3, my_df, cc_df, YEARS, args.out_dir)
    if 4 in figs_to_make:  _run(4,  make_fig4, my_df, all_df, args.out_dir)
    if 5 in figs_to_make:  _run(5,  make_fig5, my_df, args.out_dir)
    if 6 in figs_to_make:  _run(6,  make_fig6, my_df, args.out_dir, all_df=all_df)
    if 7 in figs_to_make:  _run(7,  make_fig7, my_df, args.out_dir)
    if _failed:
        print(f'WARNING: {len(_failed)} figure(s) failed: {_failed}', file=sys.stderr)
        sys.exit(1)

    # Always generate the metric/SC/cluster CSV
    make_metric_sc_cluster_csv(all_df, args.out_dir)

    print(f"\nAll done. Figures saved to: {args.out_dir}")
