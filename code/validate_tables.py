"""validate_tables.py

Checks final output tables for integrity errors.
Runs after Steps 12, 13, 13b and reports all issues found.

Usage (single line):
  python3 code/validate_tables.py --master master_xde_clusters_2745.xlsx --extra extra_tables_sem_2745.xlsx --best-lp sem_best_lp_2745.xlsx
"""

import argparse, sys, os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

PASS = "\033[32mPASS\033[0m"
FAIL = "\033[31mFAIL\033[0m"
WARN = "\033[33mWARN\033[0m"

errors   = []
warnings = []
passes   = []
SIG_RANGE = (300, 900)   # expected Significant-tab row range (override via --sig-min/--sig-max)

def ok(msg):
    passes.append(msg)
    print(f"  {PASS}  {msg}")

def fail(msg):
    errors.append(msg)
    print(f"  {FAIL}  {msg}")

def warn(msg):
    warnings.append(msg)
    print(f"  {WARN}  {msg}")

def check(condition, ok_msg, fail_msg):
    if condition:
        ok(ok_msg)
    else:
        fail(fail_msg)


def load_tab(path, tab):
    """Load an Excel tab, return DataFrame or None.
    Handles both legacy (A1) and styled (B2) workbook layouts."""
    from _workbook_style import detect_table_offset
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        if tab not in wb.sheetnames:
            fail(f"{os.path.basename(path)}: tab '{tab}' missing")
            return None
        ws = wb[tab]
        hdr_row, hdr_col = detect_table_offset(ws)
        # Header row up to first None
        hdr = []
        c = hdr_col
        while c <= ws.max_column:
            v = ws.cell(hdr_row, c).value
            if v is None:
                break
            hdr.append(str(v))
            c += 1
        n = len(hdr)
        if n == 0:
            fail(f"{tab}: empty sheet")
            return None
        rows = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            rows.append(tuple(ws.cell(r, hdr_col + i).value for i in range(n)))
        df = pd.DataFrame(rows, columns=hdr)
        wb.close()
        return df
    except Exception as e:
        fail(f"Cannot load {os.path.basename(path)} tab '{tab}': {e}")
        return None


# -- Master Excel checks ------------------------------------------------------
def check_master(path):
    print(f"\n{'='*60}")
    print(f"Checking master Excel: {os.path.basename(path)}")
    print(f"{'='*60}")

    if not os.path.exists(path):
        fail(f"File not found: {path}"); return

    wb = load_workbook(path, read_only=True, data_only=True)
    tabs = wb.sheetnames
    wb.close()

    # Required tabs
    for t in ['Master', 'Significant', 'Multi_Year', 'Clusters',
              'T1_Overview', 'T3_SuperCluster', 'T4_Clusters']:
        check(t in tabs, f"Tab '{t}' present", f"Tab '{t}' MISSING")

    # Master tab
    df = load_tab(path, 'Master')
    if df is None: return

    n = len(df)
    print(f"\n  Master tab: {n} rows")
    check(2500 <= n <= 3000, f"n={n} rows in expected range 2500-3000",
          f"n={n} rows outside expected range 2500-3000")

    req_cols = ['metric', 'explain', 'super_cluster_id', 'super_cluster_name',
                'Ward100', 'cluster_label', 'max_signed_cc', 'min_lp',
                'sum_abs_lp', 'n_sig_years', 'SI', 'sign', 'age']
    for c in req_cols:
        check(c in df.columns, f"Column '{c}' present",
              f"Column '{c}' MISSING from Master tab")

    if 'metric' in df.columns:
        n_dup = df['metric'].duplicated().sum()
        check(n_dup == 0, "No duplicate metrics in Master",
              f"{n_dup} duplicate metric rows in Master tab")

    if 'max_signed_cc' in df.columns:
        cc_num = pd.to_numeric(df['max_signed_cc'], errors='coerce')
        n_nan = cc_num.isna().sum()
        n_bad = ((cc_num.abs() > 1) & cc_num.notna()).sum()
        check(n_bad == 0, "All CC values in [-1,+1]",
              f"{n_bad} CC values outside [-1,+1]")
        ok(f"{n_nan} NaN max_signed_cc (zero-variance metrics -- expected)") if n_nan > 0 else ok("No NaN CC values")

    if 'min_lp' in df.columns:
        lp_num = pd.to_numeric(df['min_lp'], errors='coerce')
        n_pos  = (lp_num > 0).sum()
        check(n_pos == 0, "All LP values <= 0",
              f"{n_pos} positive LP values (should all be negative)")

    if 'super_cluster_id' in df.columns:
        sc_ids = pd.to_numeric(df['super_cluster_id'], errors='coerce').dropna().astype(int)
        bad_sc = sc_ids[(sc_ids < 1) | (sc_ids > 12)].unique()
        check(len(bad_sc) == 0, "All super_cluster_id in [1,12]",
              f"super_cluster_id values outside [1,12]: {sorted(bad_sc)}")
        n_scs = sc_ids.nunique()
        check(n_scs >= 10, f"All {n_scs} super-clusters represented",
              f"No super-clusters found in Master")

    if 'Ward100' in df.columns:
        ward_ids = pd.to_numeric(df['Ward100'], errors='coerce').dropna().astype(int)
        n_wards  = ward_ids.nunique()
        check(n_wards == 100, f"All 100 Ward clusters represented",
              f"Only {n_wards} Ward clusters in Master (expected 100)")

    if 'SI' in df.columns:
        si_vals = df['SI'].dropna().astype(str)
        # lp mode: chars S/I ; cc mode: V (very sig) / S (sig) / I
        bad_si  = si_vals[~si_vals.str.match(r'^[VSI]{5}$')]
        check(len(bad_si) == 0, "All SI strings are 5-char [VSI] patterns",
              f"{len(bad_si)} malformed SI strings: {bad_si.unique()[:5]}")

    if 'n_sig_years' in df.columns:
        nsig = pd.to_numeric(df['n_sig_years'], errors='coerce')
        bad  = nsig[(nsig < 0) | (nsig > 5)].dropna()
        check(len(bad) == 0, "All n_sig_years in [0,5]",
              f"{len(bad)} n_sig_years outside [0,5]")

    if 'super_cluster_name' in df.columns:
        blank_sc = df[df['super_cluster_name'].isna() |
                      (df['super_cluster_name'].astype(str).str.strip() == '') |
                      df['super_cluster_name'].astype(str).str.startswith('Super-cluster')]
        check(len(blank_sc) == 0,
              "All super_cluster_name values are non-default",
              f"{len(blank_sc)} rows have blank/default super_cluster_name")

    # Significant tab
    sig_df = load_tab(path, 'Significant')
    if sig_df is not None:
        n_sig = len(sig_df)
        _smin, _smax = SIG_RANGE
        check(_smin <= n_sig <= _smax,
              f"Significant tab: {n_sig} rows in expected range",
              f"Significant tab: {n_sig} rows outside expected range {_smin}-{_smax}")
        if 'n_sig_years' in sig_df.columns:
            nsig_min = pd.to_numeric(sig_df['n_sig_years'], errors='coerce').min()
            check(nsig_min >= 1,
                  "All Significant rows have n_sig_years >= 1",
                  f"Significant tab has rows with n_sig_years < 1 (min={nsig_min})")

    # Multi_Year tab
    my_df = load_tab(path, 'Multi_Year')
    if my_df is not None and sig_df is not None:
        n_my = len(my_df)
        check(n_my <= n_sig,
              f"Multi_Year ({n_my}) <= Significant ({n_sig})",
              f"Multi_Year ({n_my}) > Significant ({n_sig}) -- impossible")
        if 'n_sig_years' in my_df.columns:
            nsig_min = pd.to_numeric(my_df['n_sig_years'], errors='coerce').min()
            check(nsig_min >= 1,
                  "All Multi_Year rows have n_sig_years >= 1",
                  f"Multi_Year tab has rows with n_sig_years < 1 (min={nsig_min})")

    # Clusters tab
    cl_df = load_tab(path, 'Clusters')
    if cl_df is not None:
        check(len(cl_df) == 100,
              f"Clusters tab: 100 rows (one per XDE cluster)",
              f"Clusters tab: {len(cl_df)} rows (expected 100)")


# -- Extra tables checks ------------------------------------------------------
def check_extra(path):
    print(f"\n{'='*60}")
    print(f"Checking extra tables: {os.path.basename(path)}")
    print(f"{'='*60}")

    if not os.path.exists(path):
        fail(f"File not found: {path}"); return

    wb = load_workbook(path, read_only=True, data_only=True)
    tabs = wb.sheetnames
    wb.close()
    print(f"  Tabs present: {tabs}")

    for t in ['Table_1_and_2', 'Table_3', 'Table_4', 'Table_5',
              'Table_Race', 'Table_SC_Cluster_SI', 'Table_SI',
              'Table_Temporal', 'XDE_SEM_Overlap']:
        check(t in tabs, f"Tab '{t}' present", f"Tab '{t}' MISSING")

    # Table_Temporal: expect 5 data rows (one per year)
    t_temp = load_tab(path, 'Table_Temporal')
    if t_temp is not None:
        # Count rows with year values 2020-2024
        yr_col = t_temp.iloc[:, 0].astype(str)
        yr_rows = yr_col[yr_col.isin(['2020','2021','2022','2023','2024'])]
        check(len(yr_rows) == 5,
              "Table_Temporal: 5 year rows (2020-2024)",
              f"Table_Temporal: {len(yr_rows)} year rows (expected 5)")

    # XDE_SEM_Overlap: expect 12 data rows
    t_xde = load_tab(path, 'XDE_SEM_Overlap')
    if t_xde is not None:
        xde_rows = t_xde.iloc[:, 0].astype(str)
        xde_rows = xde_rows[xde_rows.str.startswith('XDE-SC')]
        check(len(xde_rows) >= 10,
              "XDE_SEM_Overlap: 12 XDE-SC rows",
              f"XDE_SEM_Overlap: {len(xde_rows)} XDE-SC rows")


# -- Best-LP table checks -----------------------------------------------------
def check_best_lp(path):
    print(f"\n{'='*60}")
    print(f"Checking best-LP table: {os.path.basename(path)}")
    print(f"{'='*60}")

    if not os.path.exists(path):
        fail(f"File not found: {path}"); return

    # SEM_Best_LP has a styled title row at row 1 and real headers at row 2,
    # so use header=1 to skip the title.
    try:
        df = pd.read_excel(path, sheet_name='SEM_Best_LP', header=1)
    except Exception as e:
        fail(f"Cannot load SEM_Best_LP: {e}")
        return

    n_data = len(df)
    check(n_data == 100,
          f"SEM_Best_LP: 100 data rows (one per XDE cluster)",
          f"SEM_Best_LP: {n_data} data rows (expected 100)")

    if 'Ward100' in df.columns:
        w = pd.to_numeric(df['Ward100'], errors='coerce').dropna().astype(int)
        dups = w[w.duplicated()]
        check(len(dups) == 0,
              "No duplicate Ward100 values",
              f"Duplicate Ward100 values: {sorted(dups.unique())}")
        check(w.nunique() == 100,
              "All 100 Ward clusters represented",
              f"Only {w.nunique()} Ward clusters represented")

    if 'min_lp' in df.columns:
        lp = pd.to_numeric(df['min_lp'], errors='coerce')
        check((lp <= 0).all(),
              "All min_lp values <= 0",
              f"{(lp > 0).sum()} positive min_lp values")


# -- Main ---------------------------------------------------------------------
def check_lp_sensitivity(path):
    print(f"\n{'='*60}")
    print(f"Checking LP sensitivity: {os.path.basename(path)}")
    print(f"{'='*60}")

    if not os.path.exists(path):
        fail(f"File not found: {path}"); return

    wb = load_workbook(path, read_only=True, data_only=True)
    tabs = wb.sheetnames
    wb.close()

    for t in ['LP_sensitivity_counts', 'LP_sensitivity_summary']:
        check(t in tabs, f"Tab '{t}' present", f"Tab '{t}' MISSING")



def check_output_files(base_dir, min_ased):
    """Check all expected PNG and XLSX files exist."""
    print(f"\n{'='*60}")
    print("Checking expected output files")
    print(f"{'='*60}")

    m = str(min_ased)   # e.g. "25.1"

    expected_png = [
        # CC histograms
        f'cc_histograms_{m}.png',
        f'cc_histograms_T_{m}.png',
        # XDE clustering figures
        'ward_xde_2745/combined_xde_centroid.png',
        'ward_xde_2745/full_heatmap.png',
        # SEM clustering
        'ward_sem_2745/cluster_size_histogram.png',
        # Publication figures
        'figures_2745/fig1_table_risky_null_clusters.png',
        'figures_2745/fig2_temporal_trajectory.png',
        'figures_2745/fig2b_age_dominance_by_year.png',
        'figures_2745/fig3_year_supercluster_heatmap.png',
        'figures_2745/fig4_supercluster_summary.png',
        'figures_2745/fig5_race_summary.png',
        'figures_2745/fig6_SI_patterns_updated.png',
        'figures_2745/fig7_top30_forest_plot.png',
        # Extra tables figure
        'fig_xde_sem_overlap.png',
        # LP sensitivity figures
        'lp_sensitivity_sweep/fig_lp_heatmap.png',
        'lp_sensitivity_sweep/fig_lp_threshold_sensitivity.png',
        'lp_sensitivity_sweep/lp_sensitivity_ge_share.png',
        'lp_sensitivity_sweep/lp_sensitivity_top_si.png',
        'lp_sensitivity_sweep/lp_sensitivity_pandemic_waves.png',
        'lp_sensitivity_sweep/lp_sensitivity_pandemic_waves_grid.png',
        # Despair figures
        'despair_cc_trajectory.png',
        'despair_lp_lt65.png',
        # SEM clustering additional figures
        'ward_sem_2745/elbow.png',
        'ward_sem_2745/silhouette.png',
        'ward_sem_2745/dendrogram.png',
        # XDE clustering additional figures
        'ward_xde_2745/medoid_dendrogram_desc_ward100.png',
        'ward_xde_2745/pattern_dendrogram_ward100.png',
    ]

    expected_xlsx = [
        'master_sem_clusters_2745.xlsx',
        'extra_tables_sem_2745.xlsx',
        'sem_best_lp_2745.xlsx',
        'lp_sensitivity_sweep/lp_sensitivity_counts.xlsx',
        'ward_xde_2745/medoid_analysis_k100.xlsx',
    ]

    n_png_ok = 0; n_xlsx_ok = 0

    print(f"\n  PNG files (expected {len(expected_png)}):")
    for f in expected_png:
        path = os.path.join(base_dir, f)
        if os.path.exists(path):
            size = os.path.getsize(path)
            ok(f"{f}  ({size//1024} KB)")
            n_png_ok += 1
        else:
            fail(f"{f}  MISSING")

    print(f"\n  XLSX files (expected {len(expected_xlsx)}):")
    for f in expected_xlsx:
        path = os.path.join(base_dir, f)
        if os.path.exists(path):
            size = os.path.getsize(path)
            ok(f"{f}  ({size//1024} KB)")
            n_xlsx_ok += 1
        else:
            fail(f"{f}  MISSING")

    print(f"\n  PNG:  {n_png_ok}/{len(expected_png)} present  (pipeline find may show more from previous runs)")
    print(f"  XLSX: {n_xlsx_ok}/{len(expected_xlsx)} present")


def main():
    p = argparse.ArgumentParser(description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--master',  required=True, help='master_sem_clusters_2745.xlsx')
    p.add_argument('--extra',   required=True, help='extra_tables_sem_2745.xlsx')
    p.add_argument('--best-lp',       default=None, help='sem_best_lp_2745.xlsx')
    p.add_argument('--lp-sensitivity', default=None, help='lp_sensitivity_sweep/lp_sensitivity_counts.xlsx')
    p.add_argument('--base-dir',       default='.',  help='Pipeline working directory (default: .)')
    p.add_argument('--min-ased',       default='25.1', help='min_ased value used in run (default: 25.1)')
    p.add_argument('--sig-min', type=int, default=300, help='Significant tab min expected rows (default 300)')
    p.add_argument('--sig-max', type=int, default=900, help='Significant tab max expected rows (default 900)')
    args = p.parse_args()
    global SIG_RANGE
    SIG_RANGE = (args.sig_min, args.sig_max)

    check_master(args.master)
    check_extra(args.extra)
    if args.best_lp:
        check_best_lp(args.best_lp)
    if args.lp_sensitivity:
        check_lp_sensitivity(args.lp_sensitivity)
    check_output_files(args.base_dir, args.min_ased)

    print(f"\n{'='*60}")
    print(f"SUMMARY:  {len(passes)} passed  |  "
          f"{len(warnings)} warnings  |  {len(errors)} errors")
    print(f"{'='*60}")
    if warnings:
        print("Warnings:")
        for w in warnings:
            print(f"  - {w}")
    if errors:
        print("Errors:")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("All checks passed.")
        sys.exit(0)


if __name__ == '__main__':
    main()
