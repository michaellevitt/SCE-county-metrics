#!/usr/bin/env python3
"""
make_master_excel_v12.py
=======================
Build master_xde_clusters.xlsx from source files.

Changes vs v10:
  - Added SI_near column using LP=-6.5 (Bonferroni threshold for 41,535 tests)
    SI_near uses 3 symbols: S=LP<=-13, s=-13<LP<=-6.5, I=LP>-6.5
  - Added n_sig_years_near, first_sig_near, last_sig_near columns
  - Master and Multi_Year tabs include SI_near (Multi_Year: n_sig_years >= 1)

Changes vs v9:
  - Compatible with ward_medoid_ccbs_v4 output file names
    (xde100_assignments.csv, medoid_list_xde100.json, Z_med_xde100.npy)
  - Labels file: reads label_proposed column from generate_cluster_labels.py output
  - --labels arg now also accepts label_proposed_v2 fallback column
  - --reps arg: reads Ward100 column dynamically (already robust)
  - --sc-names-file arg added: optional CSV to load SC names at runtime
    without editing the script (columns: sc_id, sc_name)
  - SC_NAMES loaded dynamically from --sc-names-file (required)
  - All v9 changes preserved (cluster_mean_abs_CC, T2_Temporal update)

Tabs produced:
  Master      -- all 2,602 metrics (25 columns)
  Clusters    -- one row per XDE100 cluster (9 columns, incl cluster_mean_abs_CC)
  Significant -- metrics with any S in SI string
  Multi_Year  -- metrics with n_sig_years >= 1 + rank_abs_cc column
  T1_Overview      -- overall summary statistics
  T2_Temporal      -- year-by-year CC stats table + updated line chart
  T3_SuperCluster  -- per super-cluster statistics + bar chart
  T4_Clusters      -- per XDE cluster statistics + scatter plot
  T5_Domains       -- domain label validation table
  T6_Top30         -- top 30 metrics by |CC| + bar chart
  T7_Patterns      -- SI pattern distribution + grouped bar chart

Usage:
  python3 make_master_excel_v12.py \
    --cc-file        metric_x_death_cc_0_0_25_1.csv \
    --assignments    xde100_assignments_ling.csv \
    --labels         xde100_labels_ling_v1.csv \
    --reps           xde100_reps_ling.csv \
    --extended-explain BEN_MERGED_MEASURES_explain_extended_2603.csv \
    --z-med          Z_med_ling.npy \
    --medoid-list    medoid_list_ling.json \
    --ei-file        hub_members_extensive_intensive.csv \
    --output         master_xde_clusters.xlsx
"""

import argparse
import json
import re
import os
import sys
import sys
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io

from scipy.cluster.hierarchy import dendrogram, fcluster
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

def _tee(path):
    """Print 'Saved <relpath>' to stdout and stderr."""
    _p = str(path)
    try:
        _rel = os.path.relpath(_p)
    except ValueError:
        _rel = _p
    msg = 'Saved ' + _rel
    print(msg)
    print(msg, file=sys.stderr, flush=True)


YEARS = ['2020', '2021', '2022', '2023', '2024']

# ============================================================
# Style constants
# ============================================================
NAVY   = '1F3864'
MID    = '2E75B6'
LIGHT  = 'D6E4F0'
WHITE  = 'FFFFFF'
GREEN  = 'E2EFDA'
AMBER  = 'FFF2CC'
RED_BG = 'FCE4D6'
GREY   = 'F2F2F2'
RISKY_COL = '#C0392B'
PROT_COL  = '#2980B9'
GREY_COL  = '#7F8C8D'

thin = Side(style='thin', color='CCCCCC')

# SC names loaded dynamically from sem_sc_names.csv
SC_NAMES       = {}  # populated dynamically by load_sc_names()

def load_sc_names(csv_path):
    """Populate SC_NAMES from sem_sc_names.csv."""
    global SC_NAMES
    import pandas as _pd
    if not csv_path or not os.path.exists(csv_path):
        print(f"  WARNING: sc_names file not found: {csv_path}", file=sys.stderr)
        return
    df = _pd.read_csv(csv_path)
    for _, r in df.iterrows():
        sc_id = int(r['sc_id'])
        SC_NAMES[sc_id]       = str(r.get('sc_name',       f'Super-cluster {sc_id:02d}')).strip()
    print(f"  Loaded {len(SC_NAMES)} SC names from {csv_path}")



def dat(ws, row, col, val, bg=WHITE, bold=False, left=False, courier=False):
    c = ws.cell(row=row, column=col, value=val)
    fname = 'Courier New' if courier else 'Arial'
    c.font = Font(name=fname, size=9, bold=bold)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left' if left else 'center', vertical='center')
    return c

def brd(c):
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def hdr(ws, row, col, val, bg=None, rotate=False, wrap=False):
    """Styled header cell (white bold text on navy/mid background)."""
    if bg is None:
        bg = NAVY
    cl = ws.cell(row=row, column=col, value=val)
    cl.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    cl.fill = PatternFill('solid', start_color=bg)
    cl.alignment = Alignment(
        horizontal='center', vertical='bottom' if rotate else 'center',
        text_rotation=60 if rotate else 0,
        wrap_text=wrap
    )
    return brd(cl)


def save_fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    plt.close(fig)
    return buf

def add_image_from_bytes(ws, buf, anchor):
    img = XLImage(buf)
    img.anchor = anchor
    ws.add_image(img)

# ============================================================
# Dendrogram helpers
# ============================================================
def swivel_to_ward_order(Z, n, ward_ids):
    Z2 = Z.copy()
    mw = {i: ward_ids[i] for i in range(n)}
    for step in range(len(Z2)):
        nid = n + step
        l, r = int(Z2[step, 0]), int(Z2[step, 1])
        lm, rm = mw.get(l, 0), mw.get(r, 0)
        if lm > rm:
            Z2[step, 0], Z2[step, 1] = Z2[step, 1], Z2[step, 0]
            lm, rm = rm, lm
        mw[nid] = lm
    return Z2

# ============================================================
# Build SI / sign / age strings
# ============================================================
LP_PRIMARY = -13.0   # main significance threshold (lp mode)
LP_NEAR    = -6.5    # Bonferroni threshold (alpha=0.05 / 41535 tests)

# Significance mode: 'lp' (p-value, default/canonical) or 'cc' (|CC| magnitude bands).
# In 'cc' mode a year is Significant if |CC| > CC_SIG and Very Significant if |CC| > CC_VERYSIG.
# SI string encodes V (>CC_VERYSIG) / S (CC_SIG..CC_VERYSIG] / I (<=CC_SIG).
SIG_MODE   = 'lp'
CC_SIG     = 0.3
CC_VERYSIG = 0.45

def build_pattern_strings(metrics_series, cc_df):
    si_list, si_near_list, sign_list, age_list = [], [], [], []
    for metric in metrics_series:
        if metric not in cc_df.index:
            si_list.append('IIIII'); si_near_list.append('IIIII')
            sign_list.append('_____'); age_list.append('_____')
            continue
        row = cc_df.loc[metric]
        si, si_near, sign, age = [], [], [], []
        for yr in YEARS:
            lp_all = row.get(f'LP_asedx_p_{yr}', np.nan)
            cc_all = row.get(f'asedx_p_{yr}', np.nan)
            if SIG_MODE == 'cc':
                acc         = abs(cc_all) if pd.notna(cc_all) else np.nan
                is_verysig  = pd.notna(acc) and acc > CC_VERYSIG
                is_sig      = pd.notna(acc) and acc > CC_SIG
                si.append('V' if is_verysig else ('S' if is_sig else 'I'))
                si_near.append('V' if is_verysig else ('S' if is_sig else 'I'))
            else:
                is_sig      = pd.notna(lp_all) and lp_all <= LP_PRIMARY
                is_near_sig = pd.notna(lp_all) and lp_all <= LP_NEAR
                si.append('S' if is_sig else 'I')
                # SI_near: S=LP<=-13, s=-13<LP<=-6.5, I=LP>-6.5
                if is_sig:            si_near.append('S')
                elif is_near_sig:     si_near.append('s')
                else:                 si_near.append('I')
            sign.append(('+' if cc_all >= 0 else '-') if (is_sig and pd.notna(cc_all)) else '_')
            lp_ge = row.get(f'LP_asedx_p_{yr}_GE65', np.nan)
            lp_lt = row.get(f'LP_asedx_p_{yr}_LT65', np.nan)
            cc_ge = row.get(f'asedx_p_{yr}_GE65', np.nan)
            cc_lt = row.get(f'asedx_p_{yr}_LT65', np.nan)
            if SIG_MODE == 'cc':
                either = (pd.notna(cc_ge) and abs(cc_ge) > CC_SIG) or \
                         (pd.notna(cc_lt) and abs(cc_lt) > CC_SIG)
            else:
                either = (pd.notna(lp_ge) and lp_ge <= LP_PRIMARY) or \
                         (pd.notna(lp_lt) and lp_lt <= LP_PRIMARY)
            if either and pd.notna(cc_ge) and pd.notna(cc_lt):
                age.append('>' if abs(cc_ge) > abs(cc_lt) else ('<' if abs(cc_lt) > abs(cc_ge) else '='))
            else:
                age.append('_')
        si_list.append(''.join(si)); si_near_list.append(''.join(si_near))
        sign_list.append(''.join(sign)); age_list.append(''.join(age))
    return si_list, si_near_list, sign_list, age_list

# ============================================================
# Build master dataframe
# ============================================================
def build_master(args):
    print("Loading source files...")

    load_sc_names(args.sc_names_file)

    # Load semantic cluster assignments if provided
    sem_map_id  = {}  # metric -> semantic_cluster_id
    sem_map_lbl = {}  # metric -> semantic_cluster_label
    if args.semantic_clusters and os.path.exists(args.semantic_clusters):
        sem_df = pd.read_csv(args.semantic_clusters)
        if 'semantic_cluster_id' in sem_df.columns:
            sem_map_id  = dict(zip(sem_df['metric'], sem_df['semantic_cluster_id']))
            sem_map_lbl = dict(zip(sem_df['metric'], sem_df['semantic_cluster_label']))
            print(f'Loaded semantic clusters for {len(sem_map_id)} metrics')

    cc_df   = pd.read_csv(args.cc_file).set_index('metric')
    assign  = pd.read_csv(args.sem_assignments)
    labels  = pd.read_csv(args.sem_labels)
    reps    = pd.read_csv(args.sem_reps)
    ext     = pd.read_csv(args.extended_explain)
    ei      = pd.read_csv(args.ei_file)
    Z_med   = np.load(args.sem_z_med)
    with open(args.sem_medoid_list) as f:
        medoid_list = json.load(f)

    ward_col = next((c for c in reps.columns
                    if c.startswith('Ward') or c in ('sem100','cluster_id')), None)
    if ward_col is None:
        sys.exit('ERROR: no cluster ID column in reps CSV')
    assign_col = next((c for c in assign.columns
                       if c.startswith('Ward') or c in ('sem100','cluster_id',
                                                         'semantic_cluster_id')), None)
    if assign_col is None:
        sys.exit('ERROR: no cluster ID column in assignments CSV')
    m2w = dict(zip(reps['medoid'], reps[ward_col]))
    n = len(medoid_list)

    # Dendrogram leaf order
    ward_ids = [m2w.get(m, 999) for m in medoid_list]
    Z_plot   = swivel_to_ward_order(Z_med, n, ward_ids)
    dn       = dendrogram(Z_plot, no_plot=True)
    leaf_order = dn['leaves']
    ward_leaf  = [m2w.get(medoid_list[i], 0) for i in leaf_order]

    # Super-cluster assignment
    # If --sc-assignments provided, use it (from cluster_labels_to_superclusters.py)
    # Otherwise fall back to k=N cut of Z_med dendrogram
    if args.sc_assignments and os.path.exists(args.sc_assignments):
        sc_assign_df  = pd.read_csv(args.sc_assignments)
        sc_col = 'sem100' if 'sem100' in sc_assign_df.columns else 'Ward100'
        ward_to_super = dict(zip(sc_assign_df[sc_col].astype(int),
                                 sc_assign_df['sc_id'].astype(int)))
        print(f"  Loaded SC assignments from {args.sc_assignments}: "
              f"{len(ward_to_super)} Ward->SC mappings")
    else:
        super_raw = fcluster(Z_med, t=12, criterion='maxclust')
        seen_sc, new_id_sc = {}, 1
        _dn2 = dendrogram(Z_plot, no_plot=True)
        for li in _dn2['leaves']:
            sc = int(super_raw[li])
            if sc not in seen_sc:
                seen_sc[sc] = new_id_sc; new_id_sc += 1
        ward_to_super = {m2w.get(medoid_list[i], 0): seen_sc[int(super_raw[i])] for i in range(n)}
        print(f"  Using k=12 dendrogram cut for SC assignments")

    # Sort master by dendrogram order
    # Accept label_proposed or label_proposed_v2 (second-pass improvement)
    lbl_col   = 'label_proposed_v2' if 'label_proposed_v2' in labels.columns else 'label_proposed'
    label_map = dict(zip(labels['cluster_id'], labels[lbl_col]))
    print(f"  Using label column: {lbl_col}")
    medoid_set = set(reps['medoid'])
    ward_to_pos = {wid: pos for pos, wid in enumerate(ward_leaf)}

    assign[assign_col] = assign[assign_col].astype(int)
    assign['_sort']    = assign[assign_col].map(ward_to_pos)

    # Rename assign cluster col to ward_col for uniform downstream use
    if assign_col != ward_col:
        assign = assign.rename(columns={assign_col: ward_col})
    # Drop columns from assign that will come from ext_sub to avoid _x/_y suffixes
    drop_cols = [c for c in assign.columns
                 if c in ('explain','linguistic_label','domain_label','k400_label',
                          'semantic_cluster_label','semantic_medoid_metric',
                          'is_singleton','centroid_distance','semantic_cluster_id',
                          'cluster_label','_sort') and c != ward_col]
    assign = assign[['metric', ward_col]].copy()
    # Normalise cluster col name to 'Ward100' for downstream compatibility
    if ward_col != 'Ward100':
        assign = assign.rename(columns={ward_col: 'Ward100'})
        ward_col = 'Ward100'

    # Explain and metadata
    ext_sub = ext[['metric'] + [c for c in ['explain','linguistic_label','domain_label','k400_label'] if c in ext.columns]].copy()
    ei_map  = dict(zip(ei['metric'], ei['classification']))

    # CC summary (excluding 2020-2023)
    cc_cols = [c for c in cc_df.columns if c.startswith('asedx_p_') and '2020-2023' not in c]
    lp_cols = [c for c in cc_df.columns if c.startswith('LP_asedx_p_') and '2020-2023' not in c]
    abs_cc      = cc_df[cc_cols].abs()
    best_idx    = abs_cc.values.argmax(axis=1)
    max_signed  = cc_df[cc_cols].values[np.arange(len(cc_df)), best_idx]
    best_death  = [cc_cols[i] for i in best_idx]
    min_lp      = cc_df[lp_cols].min(axis=1)
    sum_abs_lp  = cc_df[lp_cols].abs().sum(axis=1)

    cc_summary = pd.DataFrame({
        'metric': cc_df.index,
        'max_signed_cc': np.round(max_signed, 4),
        'best_death': best_death,
        'min_lp': np.round(min_lp.values, 2),
        'sum_abs_lp': np.round(sum_abs_lp.values, 2),
    }).set_index('metric')

    # Merge
    df = assign.copy()
    df = df.merge(ext_sub, on='metric', how='left')
    df = df.merge(cc_summary, on='metric', how='left')
    df['cluster_label'] = df[ward_col].map(label_map)
    df['is_medoid']     = df['metric'].isin(medoid_set).map({True: 'T', False: 'F'})

    # Pattern strings
    print("Building SI/sign/age strings...")
    si, si_near, sign, age = build_pattern_strings(df['metric'], cc_df)
    df['SI'] = si; df['SI_near'] = si_near; df['sign'] = sign; df['age'] = age

    # Derived columns
    # A year is "significant" if its SI char is 'S' or 'V' (V = very significant, cc mode).
    def _sigpos(s):
        return [i for i, ch in enumerate(s) if ch in 'SV']
    df['n_sig_years']     = df['SI'].apply(lambda s: s.count('S') + s.count('V'))
    if SIG_MODE == 'cc':   # extra tier column only in |CC| mode (keeps lp schema unchanged)
        df['n_verysig_years'] = df['SI'].apply(lambda s: s.count('V'))
    df['n_sig_years_near'] = df['SI_near'].apply(lambda s: s.count('S') + s.count('s') + s.count('V'))
    df['first_sig_year'] = df['SI'].apply(lambda s: int(YEARS[_sigpos(s)[0]]) if _sigpos(s) else None)
    df['last_sig_year']  = df['SI'].apply(lambda s: int(YEARS[_sigpos(s)[-1]]) if _sigpos(s) else None)
    df['dominant_age']   = df['age'].apply(lambda a: 'GE65' if a.count('>')>a.count('<') else ('LT65' if a.count('<')>a.count('>') else ('both' if a.count('>')+a.count('<')>0 else 'none')))

    # Cluster-level columns
    cluster_sizes = df.groupby(ward_col).size().to_dict()
    cluster_n_sig = df[df['n_sig_years']>=1].groupby(ward_col).size().to_dict()
    df['cluster_n_metrics'] = df[ward_col].map(cluster_sizes)
    df['cluster_n_sig']     = df[ward_col].map(cluster_n_sig).fillna(0).astype(int)
    df['cluster_pct_sig']   = (df['cluster_n_sig'] / df['cluster_n_metrics'] * 100).round(1)
    df['cluster_mean_cc']   = df.groupby(ward_col)['max_signed_cc'].transform('mean').round(4)
    df['super_cluster_id']  = df[ward_col].map(ward_to_super)
    df['super_cluster_name']= df['super_cluster_id'].map(SC_NAMES)

    # Data year and temporal flag
    def get_year(expl):
        m = re.search(r'=(\d{4})', str(expl))
        if m:
            yr = int(m.group(1))
            return yr if 1960 <= yr <= 2024 else None
        return None
    def get_base(metric):
        m = re.match(r'^(f\d+)\d{2}$', str(metric))
        return m.group(1) if m else None

    df['data_year']   = df['explain'].apply(get_year)
    all_metrics_list  = df['metric'].tolist()
    base_counts       = {}
    for m in all_metrics_list:
        b = get_base(m)
        if b: base_counts[b] = base_counts.get(b, 0) + 1
    df['is_temporal'] = df['metric'].apply(lambda m: base_counts.get(get_base(m), 0) > 1).map({True: 'T', False: 'F'})

    # EI class -- authoritative classification file first; otherwise a description heuristic.
    # Default is Extensive (most AHRF metrics are population-scaling counts); rate/share/ratio/
    # median/average/index/per-capita language -> Intensive. Never leaves a metric UNKNOWN.
    INTENSIVE_KW = [r'percent', r'pct', r'\brate\b', r'density', r'per[_ ]?capita', r'ratio',
                    r'median', r'average', r'\bmean\b', r'%', r'_per_', r'\bper\b', r'\bindex\b']
    def classify_ei(metric, explain):
        c = ei_map.get(metric)
        if isinstance(c, str) and c.strip():
            return c.strip().upper()                     # EXTENSIVE / INTENSIVE / META / LAND
        el = explain.lower() if isinstance(explain, str) else ''
        if any(re.search(kw, el) for kw in INTENSIVE_KW):
            return 'INTENSIVE'
        return 'EXTENSIVE'
    df['ei_class'] = df.apply(lambda r: classify_ei(r['metric'], r['explain']), axis=1)
    df['ei_class'] = df['ei_class'].replace({'EXTENSIVE': 'Extensive', 'INTENSIVE': 'Intensive',
                                             'META': 'Meta', 'LAND': 'Land'})

    # Strip year suffix from explain
    df['explain'] = df['explain'].apply(
        lambda v: re.sub(r'=\d{4}(-\d{2,4})?$', '', str(v)).rstrip('_').strip() if isinstance(v, str) else v)

    # Prefix '=' cells
    for col in df.select_dtypes(include=['object','string']).columns:
        df[col] = df[col].apply(lambda v: "'" + str(v) if isinstance(v, str) and str(v).startswith('=') else v)

    # Final column order
    col_order = [
        'super_cluster_id', 'super_cluster_name', ward_col, 'cluster_label',
        'cluster_n_metrics', 'cluster_n_sig', 'cluster_pct_sig', 'cluster_mean_cc',
        'metric', 'explain', 'data_year', 'is_temporal', 'ei_class',
        'max_signed_cc', 'best_death', 'min_lp', 'sum_abs_lp',
        'n_sig_years', 'n_verysig_years', 'first_sig_year', 'last_sig_year',
        'SI', 'SI_near', 'sign', 'age', 'dominant_age', 'is_medoid',
        'semantic_cluster_id', 'semantic_cluster_label',
    ]
    df = df[[c for c in col_order if c in df.columns]]
    df = df.sort_values(['_sort' if '_sort' in df.columns else ward_col,
                         'sum_abs_lp'], ascending=[True, False]).reset_index(drop=True)
    if '_sort' in df.columns: df = df.drop(columns=['_sort'])

    print(f"Master dataframe: {df.shape}")
    return df, cc_df, ward_col

# ============================================================
# Write a formatted data tab
# ============================================================
from _workbook_style import clean_style_workbook  # noqa: E402


def write_top_per_cluster_tab(wb, df, ward_col):
    """Top_Per_Cluster tab: one row per Ward100 cluster, the metric with the
    lowest (most significant) min_lp.

    Sort: n_sig_years descending (5,4,3,2,1,0), then min_lp ascending so the
    strongest, most-consistent signals lead.

    Style matches the screenshot template: Times New Roman, bold-black
    column headers (rotated for narrow cols, horizontal for wide text cols),
    thin black borders on all cells, no decorative fills.
    """
    print("Writing Top_Per_Cluster tab...")
    df_v = df.dropna(subset=['min_lp', ward_col]).copy()
    top  = df_v.loc[df_v.groupby(ward_col)['min_lp'].idxmin()].copy()
    top  = top.sort_values(['n_sig_years', 'min_lp'],
                           ascending=[False, True]).reset_index(drop=True)

    cols_in_order = [
        'super_cluster_id', 'super_cluster_name', ward_col, 'cluster_label',
        'metric', 'explain',
        'max_signed_cc', 'min_lp', 'sum_abs_lp', 'n_sig_years',
        'SI', 'SI_near', 'sign', 'age', 'dominant_age',
    ]
    cols = [c for c in cols_in_order if c in top.columns]
    top  = top[cols]

    HORIZ_HDR = {'super_cluster_name', 'cluster_label', 'metric', 'explain'}
    LEFT_DATA = {'super_cluster_name', 'cluster_label', 'explain'}
    BOLD_DATA = {'max_signed_cc', 'min_lp', ward_col, 'super_cluster_id'}
    MONO_DATA = {'SI', 'SI_near', 'sign', 'age', 'dominant_age', 'metric'}
    WIDTHS    = {
        'super_cluster_id': 5, 'super_cluster_name': 38, ward_col: 5,
        'cluster_label': 35, 'metric': 11, 'explain': 60,
        'max_signed_cc': 6, 'min_lp': 7, 'sum_abs_lp': 8, 'n_sig_years': 5,
        'SI': 6, 'SI_near': 7, 'sign': 6, 'age': 5, 'dominant_age': 8,
    }
    NUM_FMT   = {
        'max_signed_cc': '+0.00;-0.00;0.00',
        'min_lp':        '0.0',
        'sum_abs_lp':    '0.0',
        'n_sig_years':   '0',
        'super_cluster_id': '0',
        ward_col:        '0',
    }

    ws = wb.create_sheet('Top_Per_Cluster')
    blk_thin = Side(style='thin', color='000000')
    border   = Border(left=blk_thin, right=blk_thin, top=blk_thin, bottom=blk_thin)

    # ---- Header row ----
    for ci, col in enumerate(cols, 1):
        cl = ws.cell(row=1, column=ci, value=col)
        cl.font = Font(name='Times New Roman', bold=True, size=10)
        cl.fill = PatternFill(fill_type=None)
        if col in HORIZ_HDR:
            cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                     wrap_text=False)
        else:
            cl.alignment = Alignment(horizontal='center', vertical='bottom',
                                     text_rotation=90)
        cl.border = border
    ws.row_dimensions[1].height = 90

    # ---- Data rows ----
    for ri, row in enumerate(top.itertuples(index=False), 2):
        for ci, col in enumerate(cols, 1):
            val = getattr(row, col)
            if not isinstance(val, str) and pd.isna(val):
                val = ''
            cl = ws.cell(row=ri, column=ci, value=val)
            cl.font = Font(
                name='Courier New' if col in MONO_DATA else 'Times New Roman',
                size=10, bold=(col in BOLD_DATA))
            cl.alignment = Alignment(
                horizontal='left' if col in LEFT_DATA else 'center',
                vertical='center')
            cl.border = border
            if col in NUM_FMT and isinstance(val, (int, float)):
                cl.number_format = NUM_FMT[col]

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(col, 10)

    ws.freeze_panes = 'C2'
    return ws


def write_data_tab(wb, name, df, ward_col):
    ws = wb.create_sheet(name)
    headers = list(df.columns)

    BOLD_COLS    = {'Ward100', 'super_cluster_id', 'metric', 'SI', 'sign', 'age', ward_col}
    COURIER_COLS = {'SI', 'SI_near', 'sign', 'age'}
    LEFT_COLS    = {'cluster_label', 'super_cluster_name', 'explain'}
    COL_WIDTHS   = {
        'super_cluster_id': 6, 'super_cluster_name': 40, ward_col: 7, 'cluster_label': 35,
        'cluster_n_metrics': 8, 'cluster_n_sig': 8, 'cluster_pct_sig': 8, 'cluster_mean_cc': 10,
        'metric': 12, 'explain': 45, 'data_year': 8, 'is_temporal': 8, 'ei_class': 6,
        'max_signed_cc': 12, 'best_death': 20, 'min_lp': 8, 'sum_abs_lp': 10,
        'n_sig_years': 8, 'first_sig_year': 10, 'last_sig_year': 10,
        'SI': 8, 'SI_near': 8, 'sign': 8, 'age': 8, 'dominant_age': 10, 'is_medoid': 7,
        'semantic_cluster_id': 8, 'semantic_cluster_label': 35,
        'rank_abs_cc': 8,
    }

    for ci, col in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='bottom',
                                text_rotation=90, wrap_text=False)
    ws.row_dimensions[1].height = 80

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, col in enumerate(headers, 1):
            val = getattr(row, col)
            if not isinstance(val, str) and pd.isna(val): val = ''
            c = ws.cell(row=ri, column=ci, value=val)
            is_courier = col in COURIER_COLS
            c.font = Font(name='Courier New' if is_courier else 'Arial',
                          size=9, bold=(col in BOLD_COLS))
            c.alignment = Alignment(
                horizontal='left' if col in LEFT_COLS else 'center',
                vertical='center')

    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 10)

    ws.freeze_panes = 'A2'
    return ws

# ============================================================
# Analysis tabs
# ============================================================
def write_analysis_tabs(wb, df, cc_df, ward_col):
    my_df = df[df['n_sig_years'] >= 1].copy()
    pos   = my_df[my_df['max_signed_cc'] > 0]
    neg   = my_df[my_df['max_signed_cc'] < 0]

    ext = pd.read_csv(args.extended_explain) if hasattr(args, 'extended_explain') else None

    # helper to get year CC stats
    def yr_stats(sub):
        means, nsig = [], []
        for yr in YEARS:
            cc_col = f'asedx_p_{yr}'; lp_col = f'LP_asedx_p_{yr}'
            m = [x for x in sub['metric'] if x in cc_df.index]
            cs = cc_df.loc[m, cc_col].dropna() if cc_col in cc_df.columns and m else pd.Series()
            ls = cc_df.loc[[x for x in m if x in cc_df.index], lp_col].dropna() if lp_col in cc_df.columns and m else pd.Series()
            means.append(float(cs.abs().mean()) if len(cs) else 0)
            nsig.append(int((ls <= -13).sum()) if len(ls) else 0)
        return means, nsig

    # ---- T1 Overview ----
    ws1 = wb.create_sheet('T1_Overview')
    ws1.column_dimensions['A'].width = 44
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 34
    hdr(ws1, 1, 1, 'Summary statistic'); hdr(ws1, 1, 2, 'Value', bg=MID); hdr(ws1, 1, 3, 'Notes', bg=MID)
    ws1.row_dimensions[1].height = 22
    rows_t1 = [
        ('Total predictor metrics analysed', 2602, ''),
        ('Significant metrics any year (n_sig_years >= 1)', len(my_df), ''),
        ('  CC range', f'{my_df["max_signed_cc"].abs().min():.3f} - {my_df["max_signed_cc"].abs().max():.3f}', ''),
        ('  Mean |CC|', f'{my_df["max_signed_cc"].abs().mean():.3f}', ''),
        ('', '', ''),
        ('Any-year significant (594)', df[df['n_sig_years']>=1].shape[0], ''),
        ('Never significant', df[df['n_sig_years']==0].shape[0], 'IIIII pattern'),
        ('', '', ''),
        ('XDE clusters with >= 1 significant metric (any year)', my_df[ward_col].nunique(), 'of 100'),
        ('XDE clusters with 0 significant metrics (any year)', 100 - my_df[ward_col].nunique(), ''),
        ('', '', ''),
        ('SSSSS (all 5 years sig)', len(my_df[my_df['SI']=='SSSSS']), 'all 5 years significant'),
        ('SSIII (2020-21 only)', len(my_df[my_df['SI']=='SSIII']), '2020-21 only'),
        ('ISSSS (2021-24)', len(my_df[my_df['SI']=='ISSSS']), '2021-24'),
        ('', '', ''),
        ('GE65 dominant age effect', len(my_df[my_df['dominant_age']=='GE65']), ''),
        ('LT65 dominant age effect', len(my_df[my_df['dominant_age']=='LT65']), ''),
        ('Both age groups', len(my_df[my_df['dominant_age']=='both']), ''),
        ('', '', ''),
        ('Strongest +CC metric |CC|', f"+{pos['max_signed_cc'].max():.3f}", 'Uninsured <65 <=138% FPL'),
        ('Strongest −CC metric |CC|', f"{neg['max_signed_cc'].min():.3f}", 'White non-Hisp income >$100K'),
    ]
    for ri, (label, val, note) in enumerate(rows_t1, 2):
        bold = label != '' and not label.startswith('  ')
        bg   = LIGHT if bold else WHITE
        if label == '': bg = GREY
        brd(dat(ws1, ri, 1, label, bg=bg, bold=bold, left=True))
        brd(dat(ws1, ri, 2, val,   bg=bg, bold=bold))
        brd(dat(ws1, ri, 3, note,  bg=bg, left=True))

    # ---- T2 Temporal (sign-invariant) ----
    # Columns: Year | n sig | Mean |CC| | Min |CC| | Max |CC|
    # The sign of a metric's CC is a coding convention, so we pool ALL significant
    # metrics and report |CC| statistics rather than splitting into risky/protective.
    ws2 = wb.create_sheet('T2_Temporal')
    col_widths_t2 = [8, 10, 12, 12, 12]
    for ci, w in enumerate(col_widths_t2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    NEUTRAL_COL = '#4F81BD'

    # Single header row
    ws2.row_dimensions[1].height = 60
    hdrs_t2 = ['Year', 'n sig', 'Mean |CC|', 'Min |CC|', 'Max |CC|']
    for ci, h in enumerate(hdrs_t2, 1):
        c = ws2.cell(1, ci, h)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='bottom',
                                text_rotation=(90 if ci > 1 else 0))
        brd(c)

    # Per-year |CC| statistics over ALL significant metrics (my_df)
    def yr_abs_stats(sub, yr):
        m = [x for x in sub['metric'] if x in cc_df.index]
        cc_col = f'asedx_p_{yr}'; lp_col = f'LP_asedx_p_{yr}'
        cs = cc_df.loc[m, cc_col].dropna().abs() if m and cc_col in cc_df.columns else pd.Series()
        ls = cc_df.loc[[x for x in m if x in cc_df.index], lp_col].dropna() if m and lp_col in cc_df.columns else pd.Series()
        n_sig = int((ls <= -13).sum()) if len(ls) else 0
        mean  = round(float(cs.mean()), 4) if len(cs) else None
        mn    = round(float(cs.min()),  4) if len(cs) else None
        mx    = round(float(cs.max()),  4) if len(cs) else None
        return n_sig, mean, mn, mx

    for ri, yr in enumerate(YEARS, 2):
        bg = LIGHT if ri % 2 == 0 else WHITE
        ws2.row_dimensions[ri].height = 15
        n_sig, mean, mn, mx = yr_abs_stats(my_df, yr)
        for ci, v in enumerate([yr, n_sig, mean, mn, mx], 1):
            brd(dat(ws2, ri, ci, v, bg=bg))

    ws2.freeze_panes = 'A2'

    # Sign-invariant temporal chart: single mean-|CC| line + 5th-95th %ile band
    def full_abs_yr_stats(sub):
        out = {}
        for yr in YEARS:
            cc_col = f'asedx_p_{yr}'; lp_col = f'LP_asedx_p_{yr}'
            m  = [x for x in sub['metric'] if x in cc_df.index]
            cs = cc_df.loc[m, cc_col].dropna().abs() if m and cc_col in cc_df.columns else pd.Series()
            ls = cc_df.loc[[x for x in m if x in cc_df.index], lp_col].dropna() if m and lp_col in cc_df.columns else pd.Series()
            out[yr] = {
                'n_sig': int((ls <= -13).sum()) if len(ls) else 0,
                'mean':  float(cs.mean())        if len(cs) else np.nan,
                'p5':    float(cs.quantile(.05))  if len(cs) else np.nan,
                'p95':   float(cs.quantile(.95))  if len(cs) else np.nan,
            }
        return out

    ms = full_abs_yr_stats(my_df)
    x  = np.arange(len(YEARS))
    means = [ms[yr]['mean'] for yr in YEARS]
    p5    = [ms[yr]['p5']   for yr in YEARS]
    p95   = [ms[yr]['p95']  for yr in YEARS]

    fig, ax = plt.subplots(figsize=(6.7, 6))
    ax.fill_between(x, p5, p95, alpha=0.22, color=NEUTRAL_COL, label='5th-95th %ile of |CC|')
    ax.plot(x, means, 'o-', color=NEUTRAL_COL, lw=2.5, ms=9, zorder=5, label='Mean |CC|')
    for i, yr in enumerate(YEARS):
        ax.annotate(f'n={ms[yr]["n_sig"]}', (x[i], means[i]), textcoords='offset points',
                    xytext=(0, 10), ha='center', fontsize=8, color=NEUTRAL_COL, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(YEARS, fontsize=11)
    ax.set_ylabel('Absolute correlation coefficient |CC|', fontsize=11)
    ax.set_xlabel('Year', fontsize=11)
    ax.set_title('Temporal trajectory of |CC| by year\nLine=mean  |  Band=5th-95th %ile', fontsize=10)
    ax.legend(fontsize=9, loc='upper right', framealpha=0.92)
    ax.grid(True, alpha=0.25)
    ax.set_ylim(0, max([v for v in p95 if not np.isnan(v)] + [0.1]) * 1.18)
    plt.tight_layout()
    ws2.column_dimensions['G'].width = 58
    add_image_from_bytes(ws2, save_fig_to_bytes(fig), 'G1')

    # ---- T3 SuperCluster ----
    ws3 = wb.create_sheet('T3_SuperCluster')
    ws3.column_dimensions['A'].width = 6; ws3.column_dimensions['B'].width = 46
    for ci, w in enumerate([8,8,8,10,10,8,8,8,10,10], 3):
        ws3.column_dimensions[get_column_letter(ci+2)].width = w
    h3 = ['SC','Super-cluster name','n total','n sig','% sig','Mean |CC|','% GE65','% LT65']
    for ci, h in enumerate(h3, 1): brd(hdr(ws3, 1, ci, h, rotate=(ci>2), wrap=(ci<=2)))
    ws3.row_dimensions[1].height = 80
    for ri, sc in enumerate(sorted(df['super_cluster_id'].dropna().astype(int).unique()), 2):
        a_sub = df[df['super_cluster_id']==sc]; s_sub = my_df[my_df['super_cluster_id']==sc]
        nt = len(a_sub); ns = len(s_sub); pct = round(ns/nt*100,1) if nt else 0
        mac = round(s_sub['max_signed_cc'].abs().mean(),3) if ns else None
        ge = round((s_sub['dominant_age']=='GE65').mean()*100,1) if ns else None
        lt = round((s_sub['dominant_age']=='LT65').mean()*100,1) if ns else None
        bg = RED_BG if ns==0 else (GREEN if pct>=80 else (AMBER if pct>=30 else WHITE))
        for ci, v in enumerate([sc, SC_NAMES.get(sc,''), nt, ns, pct, mac, ge, lt], 1):
            brd(dat(ws3, ri, ci, v, bg=bg, left=(ci==2)))

    # SuperCluster bar chart
    sc_data = [(sc, SC_NAMES.get(sc,''), len(df[df['super_cluster_id']==sc]),
                len(my_df[my_df['super_cluster_id']==sc]),
                my_df[my_df['super_cluster_id']==sc]['max_signed_cc'].mean() if len(my_df[my_df['super_cluster_id']==sc]) else 0)
               for sc in sorted(df['super_cluster_id'].dropna().astype(int).unique())]
    sc_df2 = pd.DataFrame(sc_data, columns=['sc','name','n_tot','n_sig','mc'])
    sc_df2['pct'] = sc_df2.apply(lambda r: r.n_sig/r.n_tot*100 if r.n_tot else 0, axis=1)
    sc_df2 = sc_df2.sort_values('pct')
    fig2, ax2 = plt.subplots(figsize=(9, 6))
    ax2.barh(range(len(sc_df2)), sc_df2['pct'], color='#4F81BD', alpha=0.85, edgecolor='white')
    ax2.set_yticks(range(len(sc_df2)))
    ax2.set_yticklabels([f"SC{int(r.sc):2d}: {r['name']}" for _, r in sc_df2.iterrows()], fontsize=9)
    ax2.set_xlabel('% of cluster metrics that are significant (any year)', fontsize=11)
    ax2.set_title('Super-cluster significance rate', fontsize=12)
    for i, row in enumerate(sc_df2.itertuples()):
        ax2.text(row.pct+0.5, i, f'{row.n_sig}/{row.n_tot} ({row.pct:.0f}%)', va='center', fontsize=8)
    ax2.set_xlim(0, 115); ax2.grid(True, axis='x', alpha=0.3)
    ws3.column_dimensions['N'].width = 75
    add_image_from_bytes(ws3, save_fig_to_bytes(fig2), 'N1')

    # ---- T4 Clusters ----
    ws4 = wb.create_sheet('T4_Clusters')
    ws4.column_dimensions['A'].width = 6; ws4.column_dimensions['B'].width = 46
    for ci, w in enumerate([8,8,8,10,10,10,10,8], 3):
        ws4.column_dimensions[get_column_letter(ci+2)].width = w
    h4 = ['W','Cluster label','SC','n total','n sig','% sig','Mean |CC|','Top SI']
    for ci, h in enumerate(h4, 1): brd(hdr(ws4, 1, ci, h, rotate=(ci>2), wrap=(ci<=2)))
    ws4.row_dimensions[1].height = 80
    cluster_stats = df.groupby([ward_col,'cluster_label','super_cluster_id']).agg(n_total=('metric','count')).reset_index()
    sig_counts2 = my_df.groupby(ward_col).size().reset_index(name='n_sig')
    cluster_stats = cluster_stats.merge(sig_counts2, on=ward_col, how='left')
    cluster_stats['n_sig'] = cluster_stats['n_sig'].fillna(0).astype(int)
    cluster_stats['pct']   = (cluster_stats['n_sig']/cluster_stats['n_total']*100).round(1)
    cluster_stats['super_cluster_id'] = pd.to_numeric(cluster_stats['super_cluster_id'], errors='coerce')
    cluster_stats = cluster_stats.sort_values(['super_cluster_id', ward_col])
    for ri, (_, r) in enumerate(cluster_stats.iterrows(), 2):
        wid = int(r[ward_col]); s_sub = my_df[my_df[ward_col]==wid]
        mac = round(s_sub['max_signed_cc'].abs().mean(),3) if len(s_sub) else None
        tsi = s_sub['SI'].value_counts().index[0] if len(s_sub) else ''
        pct = r['pct']
        bg  = RED_BG if pct==0 else (GREEN if pct>=80 else (AMBER if pct>=30 else WHITE))
        sc  = int(r['super_cluster_id']) if pd.notna(r['super_cluster_id']) else ''
        for ci, v in enumerate([wid, r['cluster_label'], sc, int(r['n_total']), int(r['n_sig']), pct,
                                  mac, tsi], 1):
            brd(dat(ws4, ri, ci, v, bg=bg, left=(ci==2), bold=(ci==1)))

    # Cluster scatter
    fig3, ax3 = plt.subplots(figsize=(9, 7))
    SC_COLORS = plt.cm.tab20(np.linspace(0, 1, 12))
    for _, r in cluster_stats.iterrows():
        sc = int(r['super_cluster_id']) if pd.notna(r['super_cluster_id']) else 0
        col = SC_COLORS[(sc-1)%12] if sc > 0 else 'grey'
        ax3.scatter(r['n_total'], r['n_sig'], color=col, s=60, alpha=0.8, edgecolors='white', lw=0.5)
        if r['n_sig'] >= 10:
            ax3.annotate(f"W{int(r[ward_col])}", (r['n_total'], r['n_sig']), fontsize=6.5,
                         xytext=(3,2), textcoords='offset points')
    ax3.plot([0,140],[0,140],'--',color='grey',lw=0.8,alpha=0.5,label='100% sig line')
    ax3.set_xlabel('Total metrics in cluster', fontsize=12)
    ax3.set_ylabel('Multi-year significant metrics', fontsize=12)
    ax3.set_title('Cluster redundancy: total vs significant metrics\n(labelled if >= 10 significant)', fontsize=12)
    patches3 = [mpatches.Patch(color=SC_COLORS[i], label=f'SC{i+1}') for i in range(12)]
    ax3.legend(handles=patches3, fontsize=7, ncol=2, loc='upper left')
    ax3.grid(True, alpha=0.3)
    ws4.column_dimensions['N'].width = 75
    add_image_from_bytes(ws4, save_fig_to_bytes(fig3), 'N1')

    # ---- T5 Domains ----
    ws5 = wb.create_sheet('T5_Domains')
    ws5.column_dimensions['A'].width = 56
    for ci, w in enumerate([8,8,8,10,8,8], 2):
        ws5.column_dimensions[get_column_letter(ci)].width = w
    h5 = ['Domain label','n total','n sig','% sig','Mean |CC|','% GE65']
    for ci, h in enumerate(h5, 1): brd(hdr(ws5, 1, ci, h, rotate=(ci>1), wrap=(ci==1)))
    ws5.row_dimensions[1].height = 80
    domain_map2 = dict(zip(pd.read_csv(args.extended_explain)['metric'],
                           pd.read_csv(args.extended_explain)['domain_label'].fillna(''))) if hasattr(args,'extended_explain') else {}
    df['domain_label_tmp']    = df['metric'].map(domain_map2).fillna('')
    my_df2 = my_df.copy(); my_df2['domain_label_tmp'] = my_df2['metric'].map(domain_map2).fillna('')
    has_all = df[df['domain_label_tmp']!='']
    has_sig = my_df2[my_df2['domain_label_tmp']!='']
    dom_stats = []
    for dom in sorted([d for d in has_all['domain_label_tmp'].unique() if isinstance(d, str) and d != '']):
        na = (has_all['domain_label_tmp']==dom).sum()
        s  = has_sig[has_sig['domain_label_tmp']==dom]
        ns = len(s); pct = round(ns/na*100,1) if na else 0
        mac = round(s['max_signed_cc'].abs().mean(),3) if ns else None
        ge  = round((s['dominant_age']=='GE65').mean()*100,1) if ns else None
        dom_stats.append((dom, na, ns, pct, mac, ge))
    dom_stats.sort(key=lambda x: -x[3])
    for ri, (dom, na, ns, pct, mac, ge) in enumerate(dom_stats, 2):
        bg = RED_BG if pct==0 else (GREEN if pct==100 else (AMBER if pct>=50 else WHITE))
        for ci, v in enumerate([dom, na, ns, pct, mac, ge], 1):
            brd(dat(ws5, ri, ci, v, bg=bg, left=(ci==1)))

    # ---- T6 Top30 ----
    ws6 = wb.create_sheet('T6_Top30')
    ws6.column_dimensions['A'].width = 6; ws6.column_dimensions['B'].width = 50
    ws6.column_dimensions['C'].width = 36
    for ci, w in enumerate([8,8,8,8,8,8], 4):
        ws6.column_dimensions[get_column_letter(ci+3)].width = w
    h6 = ['Rank','Explain','Cluster label','CC','SI','Sign','Age','Dom age','SC']
    for ci, h in enumerate(h6, 1): brd(hdr(ws6, 1, ci, h, rotate=(ci>3), wrap=(ci<=3)))
    ws6.row_dimensions[1].height = 80
    if 'rank_abs_cc' not in my_df.columns:
        my_df = my_df.copy()
        my_df['rank_abs_cc'] = my_df['max_signed_cc'].abs().rank(ascending=False).astype(int)
    top30 = my_df.nsmallest(30, 'rank_abs_cc')
    if 'rank_abs_cc' not in my_df.columns:
        my_df2b = my_df.copy()
        my_df2b['_abs'] = my_df2b['max_signed_cc'].abs()
        my_df2b['rank_abs_cc'] = my_df2b['_abs'].rank(ascending=False).astype(int)
        top30 = my_df2b.nsmallest(30, 'rank_abs_cc')
    top30 = top30.iloc[::-1]
    for ri, (_, r) in enumerate(top30.iterrows(), 2):
        bg = RED_BG if r['max_signed_cc']>0 else LIGHT
        rk = int(r['rank_abs_cc']) if 'rank_abs_cc' in r else ''
        sc = int(r['super_cluster_id']) if pd.notna(r.get('super_cluster_id','')) else ''
        for ci, v in enumerate([rk, r['explain'], r['cluster_label'], round(r['max_signed_cc'],3),
                                  r['SI'], r['sign'], r['age'], r['dominant_age'], sc], 1):
            brd(dat(ws6, ri, ci, v, bg=bg, left=(ci in [2,3]), bold=(ci==1),
                    courier=(ci in [5,6,7])))

    # Top30 bar chart
    top30_fwd = top30.iloc[::-1]
    labels_b = [str(e) for e in top30_fwd['explain']]
    ccs_b    = top30_fwd['max_signed_cc'].values
    fig4, ax4 = plt.subplots(figsize=(10, 9))
    ax4.barh(range(len(top30_fwd)), ccs_b,
             color=[RISKY_COL if c>0 else PROT_COL for c in ccs_b], alpha=0.85, edgecolor='white')
    ax4.set_yticks(range(len(top30_fwd))); ax4.set_yticklabels(labels_b, fontsize=8)
    ax4.axvline(0, color='black', lw=0.8)
    ax4.set_xlabel('Signed correlation coefficient (CC)', fontsize=11)
    ax4.set_title('Top 30 significant metrics by |CC| (any year)\n(red = +CC, blue = −CC)', fontsize=12)
    for i, (cc2, row) in enumerate(zip(ccs_b, top30_fwd.itertuples())):
        ax4.text(cc2+(0.003 if cc2>0 else -0.003), i, f' {row.SI}', va='center',
                 fontsize=6.5, color='white' if abs(cc2)>0.2 else 'black',
                 ha='left' if cc2>0 else 'right')
    ax4.grid(True, axis='x', alpha=0.3)
    ax4.set_xlim(min(ccs_b)*1.15, max(ccs_b)*1.15)
    ws6.column_dimensions['K'].width = 80
    add_image_from_bytes(ws6, save_fig_to_bytes(fig4), 'K1')

    # ---- T7 Patterns ----
    # Shows both SI (LP<=-13) and SI_near (LP<=-6.5) pattern distributions
    # SI_near includes metrics with >= 2 years at LP<=-6.5 (Bonferroni)
    ws7 = wb.create_sheet('T7_Patterns')

    # --- Primary SI (LP<=-13, multi-year) ---
    # Sign-invariant: a single bar per SI pattern = total count over all significant metrics.
    # The sign of a metric's CC is a coding convention, so pooling +CC/-CC counts would
    # be an artifact; the temporal SI pattern itself is the meaningful quantity.
    si_pos = pos['SI'].value_counts()
    si_neg = neg['SI'].value_counts()
    all_si2 = sorted(set(si_pos.index)|set(si_neg.index),
                     key=lambda x: -(si_pos.get(x,0)+si_neg.get(x,0)))
    tot_vals2 = [si_pos.get(s,0)+si_neg.get(s,0) for s in all_si2]

    # --- SI_near (LP<=-6.5, multi-year including s) ---
    # Use SI_near column where available; s/S both count as "significant"
    has_near = 'SI_near' in my_df.columns
    if has_near:
        my_near = my_df[my_df['SI_near'].apply(
            lambda x: (str(x).count('S')+str(x).count('s')) > 1 if pd.notna(x) else False)].copy()
        pos_near = my_near[my_near['max_signed_cc'] > 0]
        neg_near = my_near[my_near['max_signed_cc'] < 0]
        si_pos_near = pos_near['SI_near'].value_counts()
        si_neg_near = neg_near['SI_near'].value_counts()
        all_si_near = sorted(set(si_pos_near.index)|set(si_neg_near.index),
                             key=lambda x: -(si_pos_near.get(x,0)+si_neg_near.get(x,0)))
        # Only patterns that contain at least one 's' (new near-sig ones)
        near_only = [s for s in all_si_near if 's' in s]
        tot_near = [si_pos_near.get(s,0)+si_neg_near.get(s,0) for s in near_only]

    # --- Figure: two panels if SI_near available, one otherwise ---
    NEUTRAL_COL = '#4F81BD'
    ncols_fig = 2 if has_near else 1
    fig5, axes5 = plt.subplots(1, ncols_fig, figsize=(9*ncols_fig, 7))
    if ncols_fig == 1:
        axes5 = [axes5]

    ax5 = axes5[0]
    xb = np.arange(len(all_si2)); wb2 = 0.6
    b1 = ax5.bar(xb, tot_vals2, wb2, color=NEUTRAL_COL, alpha=0.85, edgecolor='white')
    for bar, val in zip(b1, tot_vals2):
        if val > 0:
            ax5.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                     str(val), ha='center', va='bottom', fontsize=8)
    ax5.set_xticks(xb)
    ax5.set_xticklabels(all_si2, fontsize=9, fontfamily='monospace', rotation=30, ha='right')
    ax5.set_ylabel('Number of metrics', fontsize=11)
    ax5.set_title(f'Primary SI patterns (LP<=-13)\nn={len(my_df)} significant metrics (any year)',
                  fontsize=11, fontweight='bold')
    ax5.grid(True, axis='y', alpha=0.3)

    if has_near and len(near_only) > 0:
        ax6 = axes5[1]
        xb2 = np.arange(len(near_only))
        b3 = ax6.bar(xb2, tot_near, wb2, color='#E59866', alpha=0.85, edgecolor='white')
        for bar, val in zip(b3, tot_near):
            if val > 0:
                ax6.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
                         str(val), ha='center', va='bottom', fontsize=8)
        ax6.set_xticks(xb2)
        ax6.set_xticklabels(near_only, fontsize=9, fontfamily='monospace', rotation=30, ha='right')
        ax6.set_ylabel('Number of metrics', fontsize=11)
        ax6.set_title(f'Near-sig SI_near patterns containing s (LP<=-6.5)\n'
                      f'n={len(my_near)} metrics  |  S=LP<=-13, s=-13<LP<=-6.5, I=not sig',
                      fontsize=11, fontweight='bold')
        ax6.grid(True, axis='y', alpha=0.3)

    fig5.suptitle('SI temporal pattern distributions\n'
                  'Left: primary (LP<=-13)   Right: near-sig patterns with s symbol (LP<=-6.5)',
                  fontsize=12, fontweight='bold')
    plt.tight_layout()
    ws7.column_dimensions['A'].width = 75 if not has_near else 140
    add_image_from_bytes(ws7, save_fig_to_bytes(fig5), 'A1')

    # --- Summary tables ---
    # Table 1: primary SI
    row_start = 42
    thin7 = Side(style='thin', color='CCCCCC')
    def brd7(c):
        c.border = Border(left=thin7, right=thin7, top=thin7, bottom=thin7); return c

    for ci, w in enumerate([12,8,12], 1):
        ws7.column_dimensions[get_column_letter(ci)].width = w

    ws7.cell(row_start, 1,
             ('Primary SI (|CC|>0.30)  n_sig_years >= 1' if SIG_MODE == 'cc'
              else 'Primary SI (LP<=-13)  n_sig_years >= 1')).font = \
        Font(name='Arial', bold=True, size=10)
    row_start += 1
    hdrs7 = ['SI Pattern','Total','Mean |CC|']
    for ci, h in enumerate(hdrs7, 1):
        c = brd7(ws7.cell(row=row_start, column=ci, value=h))
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri2, si2 in enumerate(all_si2, row_start+1):
        s2 = my_df[my_df['SI']==si2]
        tot = len(s2)
        mac2 = round(float(s2['max_signed_cc'].abs().mean()),3) if tot else None
        bg7 = 'F9F9F9'
        for ci, v in enumerate([si2, tot, mac2], 1):
            c = brd7(ws7.cell(row=ri2, column=ci, value=v))
            c.font = Font(name='Courier New' if ci==1 else 'Arial',
                         bold=(ci==1), size=9)
            c.fill = PatternFill('solid', start_color=bg7)
            c.alignment = Alignment(horizontal='center')

    # Table 2: SI_near near-only patterns
    if has_near and len(near_only) > 0:
        row_near = row_start + len(all_si2) + 3
        ws7.cell(row_near, 1,
                 'Near-sig SI_near (LP<=-6.5)  patterns with s symbol  '
                 '(S=LP<=-13, s=-13<LP<=-6.5, I=LP>-6.5)').font = \
            Font(name='Arial', bold=True, size=10)
        row_near += 1
        for ci, h in enumerate(hdrs7, 1):
            c = brd7(ws7.cell(row=row_near, column=ci, value=h))
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            c.fill = PatternFill('solid', start_color='37474F')
            c.alignment = Alignment(horizontal='center', vertical='center')
        for ri2, si2 in enumerate(near_only, row_near+1):
            s2 = my_near[my_near['SI_near']==si2]
            tot = len(s2)
            mac2 = round(float(s2['max_signed_cc'].abs().mean()),3) if tot else None
            bg7 = 'F5F5F5'
            for ci, v in enumerate([si2, tot, mac2], 1):
                c = brd7(ws7.cell(row=ri2, column=ci, value=v))
                c.font = Font(name='Courier New' if ci==1 else 'Arial',
                             bold=(ci==1), size=9)
                c.fill = PatternFill('solid', start_color=bg7)
                c.alignment = Alignment(horizontal='center')

    print("Analysis tabs written.")

# ============================================================
# Main
# ============================================================
def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--cc-file',          required=True)
    p.add_argument('--sem-assignments',  required=True,
                   help='ward_sem_metrics.csv')
    p.add_argument('--sem-labels',        required=True,
                   help='sem100_labels.csv from generate_cluster_labels.py')
    p.add_argument('--sem-reps',          required=True,
                   help='ward_sem_reps.csv')
    p.add_argument('--extended-explain', required=True)
    p.add_argument('--sem-z-med',         required=True,
                   help='Z_med_sem100.npy')
    p.add_argument('--sem-medoid-list',   required=True,
                   help='medoid_list_sem100.json')
    p.add_argument('--ei-file',          required=True)
    p.add_argument('--output',           default='master_sem_clusters_2745.xlsx')
    p.add_argument('--sc-names-file',    default=None,
                   help='sem_sc_names.csv (ward_xde_*/sem_sc_names.csv)')
    p.add_argument('--sc-assignments',   default=None,
                   help='Optional sem_sc_assignments.csv: Ward100->sc_id from cluster_labels_to_superclusters.py')
    p.add_argument('--semantic-clusters', default=None,
                   help='Optional ward_sem_metrics.csv to add semantic_cluster_id/label columns')
    p.add_argument('--sig-mode', choices=['lp', 'cc'], default='cc',
                   help="Significance definition: 'cc' (|CC| bands, default) or 'lp' (legacy p-value)")
    p.add_argument('--cc-sig', type=float, default=0.3,
                   help="|CC| threshold for Significant in cc mode (default 0.3)")
    p.add_argument('--cc-verysig', type=float, default=0.45,
                   help="|CC| threshold for Very Significant in cc mode (default 0.45)")
    return p.parse_args()

if __name__ == '__main__':
    args = parse_args()
    SIG_MODE   = args.sig_mode
    CC_SIG     = args.cc_sig
    CC_VERYSIG = args.cc_verysig
    print(f"Significance mode: {SIG_MODE}" +
          (f"  (Significant |CC|>{CC_SIG}, Very Significant |CC|>{CC_VERYSIG})"
           if SIG_MODE == 'cc' else f"  (LP<={LP_PRIMARY})"))
    df, cc_df, ward_col = build_master(args)

    wb = Workbook()
    wb.remove(wb.active)

    # Data tabs
    print("Writing Master tab...")
    write_data_tab(wb, 'Master', df, ward_col)

    print("Writing Clusters tab...")
    cl_df = df.drop_duplicates(subset=[ward_col]).iloc[:, :8].copy()
    # Add cluster_mean_abs_CC: mean of |max_signed_cc| across all metrics in each cluster
    cl_df['cluster_mean_abs_CC'] = cl_df[ward_col].map(
        df.groupby(ward_col)['max_signed_cc'].apply(lambda x: round(x.abs().mean(), 4)))
    ws_cl = write_data_tab(wb, 'Clusters', cl_df, ward_col)

    print("Writing Significant tab...")
    write_data_tab(wb, 'Significant', df[df['n_sig_years'] >= 1].copy(), ward_col)

    print("Writing Multi_Year tab...")
    my_df_out = df[df['n_sig_years'] >= 1].copy()
    # Add rank_abs_cc
    my_df_out['rank_abs_cc'] = my_df_out['max_signed_cc'].abs().rank(ascending=False).astype(int)
    write_data_tab(wb, 'Multi_Year', my_df_out, ward_col)

    # Top metric per cluster (one row per Ward100 cluster)
    write_top_per_cluster_tab(wb, df, ward_col)

    # Analysis tabs
    print("Writing analysis tabs...")
    write_analysis_tabs(wb, df, cc_df, ward_col)

    # Uniform style pass: shift all sheets to start at B2, Times New Roman
    # 11pt black on white, no fills, no freeze panes.
    print("Applying uniform black-on-white Times Roman 11pt style to all tabs...")
    clean_style_workbook(wb)

    wb.save(args.output)
    _tee(args.output)
    try:
        _r_args_output = os.path.relpath(args.output)
    except ValueError:
        _r_args_output = args.output
    print('Saved ' + _r_args_output, file=sys.stderr, flush=True)
    print(f"Saved: {args.output}")
