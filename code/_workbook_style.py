"""Shared workbook-styling helper.

Used by make_master_excel_v12.py and make_extra_tables_v5.py to apply a
uniform clean style across every sheet of a workbook:
  - shift content so the table starts at cell B2 (row 1 + col A blank margin)
  - Times New Roman 11pt black on white, preserving bold/italic
  - strip every PatternFill
  - remove freeze panes
  - shift embedded image anchors and conditional-formatting ranges so they
    move with the content
Borders are left untouched.
"""
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# ============================================================================
# Read-side helpers (consumers of styled workbooks)
# ============================================================================

def detect_table_offset(ws, max_scan=4):
    """Return (row, col) 1-indexed where the table starts.
    Legacy tabs return (1, 1); styled tabs return (2, 2)."""
    for r in range(1, max_scan + 1):
        for c in range(1, max_scan + 1):
            if ws.cell(r, c).value is not None:
                return (r, c)
    return (1, 1)


def read_styled_tab(ws):
    """Read a worksheet (legacy A1 or styled B2 layout) into a list-of-dicts.
    Returns (headers, list_of_dicts)."""
    hdr_row, hdr_col = detect_table_offset(ws)
    headers = []
    c = hdr_col
    while c <= ws.max_column:
        v = ws.cell(hdr_row, c).value
        if v is None:
            break
        headers.append(str(v))
        c += 1
    n_cols = len(headers)
    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, hdr_col + i).value for i in range(n_cols)}
        rows.append(row)
    return headers, rows


def read_styled_excel(path, sheet_name):
    """pandas.read_excel that handles both legacy (A1) and styled (B2) layouts.
    Returns a DataFrame with leading blank padding columns dropped."""
    import pandas as pd
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    hdr_row, hdr_col = detect_table_offset(ws)
    wb.close()
    df = pd.read_excel(path, sheet_name=sheet_name, header=hdr_row - 1)
    # Drop the padding column(s) on the left and any trailing all-NaN columns
    if hdr_col > 1:
        df = df.iloc[:, hdr_col - 1:].reset_index(drop=True)
    df = df.loc[:, ~df.columns.astype(str).str.startswith('Unnamed:')]
    return df


def _shift_anchor(anchor_str, dr=1, dc=1):
    """Shift a cell anchor like 'K1' or 'A10' by dr rows and dc columns.
    Returns the original string if it doesn't match the pattern."""
    m = re.match(r'^([A-Z]+)(\d+)$', str(anchor_str))
    if not m:
        return anchor_str
    col_letter, row_str = m.groups()
    new_col_idx = column_index_from_string(col_letter) + dc
    new_row     = int(row_str) + dr
    return f'{get_column_letter(new_col_idx)}{new_row}'


def _shift_range(range_str, dr=1, dc=1):
    """Shift an Excel range like 'F5:F10' or 'A1:Z100' by dr rows / dc cols.
    Multiple space-separated ranges are handled."""
    parts = []
    for part in str(range_str).split():
        cells = part.split(':')
        parts.append(':'.join(_shift_anchor(c, dr, dc) for c in cells))
    return ' '.join(parts)


def clean_style_workbook(wb, padding_row_height=6, padding_col_width=3):
    """Apply uniform clean style to every worksheet (see module docstring)."""
    for ws in wb.worksheets:
        # ---- Capture pre-shift dimensions ----
        old_widths  = {col: dim.width for col, dim in ws.column_dimensions.items()
                       if dim.width is not None}
        old_heights = {r: dim.height for r, dim in ws.row_dimensions.items()
                       if dim.height is not None}

        # ---- Capture conditional formatting (must shift its ranges) ----
        cf_to_readd = []
        try:
            for rng, rules in list(ws.conditional_formatting._cf_rules.items()):
                rng_str = rng.sqref if hasattr(rng, 'sqref') else str(rng)
                for rule in rules:
                    cf_to_readd.append((rng_str, rule))
            ws.conditional_formatting._cf_rules.clear()
        except Exception:
            pass

        # ---- Shift content ----
        ws.insert_rows(1)
        ws.insert_cols(1)

        # ---- Re-key column widths ----
        for col_letter in list(old_widths.keys()):
            try:
                ws.column_dimensions[col_letter].width = None
            except Exception:
                pass
        for col_letter, w in old_widths.items():
            try:
                new_idx = column_index_from_string(col_letter) + 1
                ws.column_dimensions[get_column_letter(new_idx)].width = w
            except Exception:
                pass
        ws.column_dimensions['A'].width = padding_col_width

        # ---- Re-key row heights ----
        for r in list(old_heights.keys()):
            try:
                ws.row_dimensions[r].height = None
            except Exception:
                pass
        for r, h in old_heights.items():
            try:
                ws.row_dimensions[r + 1].height = h
            except Exception:
                pass
        ws.row_dimensions[1].height = padding_row_height

        # ---- Reset font + strip fill on every populated cell ----
        for row in ws.iter_rows():
            for cell in row:
                if not cell.has_style and cell.value is None:
                    continue
                old = cell.font
                cell.font = Font(
                    name='Times New Roman',
                    size=11,
                    color='000000',
                    bold=bool(old and old.bold),
                    italic=bool(old and old.italic),
                )
                cell.fill = PatternFill(fill_type=None)

        # ---- Remove freeze panes ----
        ws.freeze_panes = None

        # ---- Re-add conditional formatting with shifted ranges ----
        for old_rng_str, rule in cf_to_readd:
            try:
                ws.conditional_formatting.add(_shift_range(old_rng_str, 1, 1), rule)
            except Exception:
                pass

        # ---- Shift image anchors ----
        if getattr(ws, '_images', None):
            for img in ws._images:
                a = img.anchor
                if isinstance(a, str):
                    img.anchor = _shift_anchor(a, dr=1, dc=1)
                elif hasattr(a, '_from') and a._from is not None:
                    a._from.row += 1
                    a._from.col += 1
                    if hasattr(a, 'to') and a.to is not None:
                        a.to.row += 1
                        a.to.col += 1
